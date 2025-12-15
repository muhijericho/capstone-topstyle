import io
import json
import logging
import re
from datetime import datetime, timedelta

import qrcode
import requests
from django.conf import settings
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.core.exceptions import ValidationError
from django.core.files.base import ContentFile
from django.core.mail import send_mail
from django.core.paginator import EmptyPage, PageNotAnInteger, Paginator
from django.db.models import Case, Count, DecimalField, F, Q, Sum, When
from django.db.models.functions import TruncDate
from django.http import HttpResponse, JsonResponse
from django.middleware.csrf import get_token
from django.shortcuts import get_object_or_404, redirect, render
from django.template.loader import render_to_string
from django.urls import reverse
from django.utils import timezone
from django.utils.crypto import get_random_string
from django.views.decorators.csrf import csrf_protect, ensure_csrf_cookie, csrf_exempt
from django.views.decorators.http import require_http_methods
from openpyxl import Workbook
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from requests.auth import HTTPBasicAuth
from twilio.base.exceptions import TwilioRestException
from twilio.rest import Client

from .forms import *
from .models import *


def offline_view(request):
    """Offline page for when user loses internet connection"""
    return render(request, 'business/offline.html')

def chrome_devtools_handler(request):
    """Handle Chrome DevTools well-known requests to suppress 404 warnings"""
    from django.http import HttpResponse
    return HttpResponse(status=204)  # No Content

def landing_page(request):
    """Landing page view with image support"""
    from .models import LandingPageImage
    
    # Get all active images
    images = {}
    for img_type, _ in LandingPageImage.IMAGE_TYPE_CHOICES:
        try:
            img = LandingPageImage.objects.get(image_type=img_type, is_active=True)
            images[img_type] = img
        except LandingPageImage.DoesNotExist:
            images[img_type] = None
    
    context = {
        'hero_image': images.get('hero'),
        'service_repair': images.get('service_repair'),
        'service_customize': images.get('service_customize'),
        'service_uniform': images.get('service_uniform'),
        'service_patches': images.get('service_patches'),
        'service_rental': images.get('service_rental'),
        'about_workshop': images.get('about_workshop'),
        'about_tailor': images.get('about_tailor'),
        'fabric_cotton': images.get('fabric_cotton'),
        'fabric_linen': images.get('fabric_linen'),
        'fabric_wool': images.get('fabric_wool'),
        'fabric_leather': images.get('fabric_leather'),
        'fabric_silk': images.get('fabric_silk'),
        'shop_exterior': images.get('shop_exterior'),
        'shop_exterior_2': images.get('shop_exterior_2'),
        'shop_interior': images.get('shop_interior'),
        'shop_interior_2': images.get('shop_interior_2'),
    }
    
    return render(request, 'business/landing.html', context)

def public_track_order(request):
    """Public track order view - no login required"""
    from .models import Order
    from django.core.exceptions import ValidationError
    
    if request.method == 'POST':
        order_id = request.POST.get('order_id', '').strip()
        decoded_order_id = request.POST.get('decoded_order_id', '').strip()
        qr_image = request.FILES.get('qr_image')
        
        # Use decoded order ID from client-side QR decoding if available
        if decoded_order_id and not order_id:
            order_id = decoded_order_id
        
        # Handle QR code image upload (server-side fallback)
        if qr_image and not order_id:
            try:
                # Decode QR code from image
                from PIL import Image
                import io
                import json
                
                try:
                    from pyzbar.pyzbar import decode as decode_qr
                except ImportError:
                    decode_qr = None
                
                if decode_qr:
                    image = Image.open(io.BytesIO(qr_image.read()))
                    decoded_objects = decode_qr(image)
                    
                    if decoded_objects:
                        qr_data = decoded_objects[0].data.decode('utf-8')
                        try:
                            qr_json = json.loads(qr_data)
                            order_id = (qr_json.get('id') or qr_json.get('orderId') or 
                                       qr_json.get('order_id') or qr_json.get('order_identifier'))
                        except json.JSONDecodeError:
                            order_id = qr_data
                    else:
                        messages.error(request, 'No QR code found in the uploaded image. Please try again or enter your Order ID manually.')
                        return render(request, 'business/public_track_order.html')
                else:
                    messages.error(request, 'QR code decoding is not available. Please enter your Order ID manually.')
                    return render(request, 'business/public_track_order.html')
            except Exception as e:
                messages.error(request, f'Error processing QR code: {str(e)}. Please try entering your Order ID manually.')
                return render(request, 'business/public_track_order.html')
        
        if not order_id:
            messages.error(request, 'Please enter an Order ID or upload a QR code image.')
            return render(request, 'business/public_track_order.html')
        
        # Search for order (same logic as track_order view)
        order = None
        search_variants = [order_id]
        
        if '-' in order_id:
            parts = order_id.split('-')
            if len(parts) == 2:
                prefix = parts[0]
                number_part = parts[1]
                number_without_leading_zeros = str(int(number_part)) if number_part.isdigit() else number_part
                
                if not 'O' in parts[1].upper() and number_part.isdigit():
                    search_variants.append(f"{prefix}-O{number_part}")
                    search_variants.append(f"{prefix}-O{number_without_leading_zeros}")
                    if len(number_without_leading_zeros) == 1:
                        search_variants.append(f"{prefix}-O0{number_without_leading_zeros}")
                    elif len(number_without_leading_zeros) == 2:
                        search_variants.append(f"{prefix}-O{number_without_leading_zeros}")
                elif 'O' in parts[1].upper() and number_part.replace('O', '').replace('o', '').isdigit():
                    number_only = number_part.replace('O', '').replace('o', '')
                    search_variants.append(f"{prefix}-{number_only}")
                    if len(number_only) == 2:
                        search_variants.append(f"{prefix}-0{number_only}")
                    elif len(number_only) == 1:
                        search_variants.append(f"{prefix}-00{number_only}")
        
        for variant in search_variants:
            try:
                order = Order.objects.get(order_identifier=variant)
                break
            except Order.DoesNotExist:
                continue
        
        if not order:
            for variant in search_variants:
                try:
                    order = Order.objects.get(order_identifier__iexact=variant)
                    break
                except Order.DoesNotExist:
                    continue
        
        if not order:
            try:
                order = Order.objects.get(order_id=order_id)
            except (Order.DoesNotExist, ValidationError):
                try:
                    if order_id.isdigit():
                        order = Order.objects.get(id=int(order_id))
                    else:
                        orders = Order.objects.filter(order_identifier__icontains=order_id)
                        if orders.exists():
                            for variant in search_variants:
                                exact_match = orders.filter(order_identifier__iexact=variant).first()
                                if exact_match:
                                    order = exact_match
                                    break
                            if not order:
                                for variant in search_variants:
                                    starts_match = orders.filter(order_identifier__istartswith=variant).first()
                                    if starts_match:
                                        order = starts_match
                                        break
                            if not order:
                                order = orders.order_by('-created_at').first()
                        else:
                            raise Order.DoesNotExist
                except (Order.DoesNotExist, ValueError):
                    messages.error(request, f'Order "{order_id}" not found. Please check your Order ID and try again.')
                    return render(request, 'business/public_track_order.html', {'order_id': order_id})
        
        if not order:
            messages.error(request, f'Order "{order_id}" not found. Please check your Order ID and try again.')
            return render(request, 'business/public_track_order.html', {'order_id': order_id})
        
        # Load order with related data
        order = Order.objects.select_related('customer').prefetch_related('items__product', 'items__product__category').get(id=order.id)
        
        return render(request, 'business/public_track_result.html', {
            'order': order,
            'found': True,
            'order_id': order_id
        })
    
    return render(request, 'business/public_track_order.html')

def login_view(request):
    if request.user.is_authenticated:
        return redirect('dashboard')
    
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect('dashboard')
        else:
            messages.error(request, 'Invalid username or password.')
    
    return render(request, 'business/login.html')


def forgot_password(request):
    """Allow users to request a password reset verification code via email or username."""
    import random
    
    if request.method == 'POST':
        identifier = request.POST.get('email', '').strip()
        
        if not identifier:
            messages.error(request, 'Email or username is required.')
            return redirect('forgot_password')
        
        # Try active accounts first; fall back to any account to give better guidance
        active_query = User.objects.filter(
            Q(email__iexact=identifier) | Q(username__iexact=identifier),
            is_active=True
        )
        any_query = User.objects.filter(Q(email__iexact=identifier) | Q(username__iexact=identifier))
        
        if active_query.exists():
            user = active_query.filter(email__iexact=identifier).first() or active_query.first()
        elif any_query.exists():
            user = any_query.filter(email__iexact=identifier).first() or any_query.first()
            messages.error(request, 'This account is inactive. Please contact an administrator to activate it before resetting the password.')
            return redirect('forgot_password')
        else:
            messages.error(request, 'No account found with that email or username.')
            return redirect('forgot_password')
        
        if not user.email:
            messages.error(request, 'This account has no email on file. Please contact an administrator to add an email before resetting the password.')
            return redirect('forgot_password')
        
        # Generate a 6-digit verification code
        verification_code = ''.join([str(random.randint(0, 9)) for _ in range(6)])
        
        # Store code in session with expiration (15 minutes)
        request.session['reset_code'] = verification_code
        request.session['reset_user_id'] = user.id
        request.session['reset_email'] = user.email
        request.session['reset_code_expires'] = (timezone.now() + timedelta(minutes=15)).isoformat()
        
        login_url = request.build_absolute_uri(reverse('login'))
        subject = 'TopStyle Business - Password Reset Verification Code'
        
        # Create email body with verification code
        email_body = f"""Hello {user.username},

You have requested to reset your password for TopStyle Business Management System.

Your verification code is: {verification_code}

This code will expire in 15 minutes.

Enter this code on the verification page to reset your password.

If you did not request this password reset, please ignore this email.

Best regards,
TopStyle Business Team

Login URL: {login_url}
"""
        
        from_email = settings.DEFAULT_FROM_EMAIL or settings.EMAIL_HOST_USER
        if not from_email:
            messages.error(request, 'Email is not configured. Please contact an administrator.')
            return redirect('forgot_password')
        
        # Send to user's email and also to admin email (ltv75850@gmail.com)
        recipient_emails = [user.email, 'ltv75850@gmail.com']
        
        try:
            send_mail(
                subject,
                email_body,
                from_email,
                recipient_emails,
                fail_silently=False,
            )
            messages.success(request, f'Verification code has been sent to {user.email} and ltv75850@gmail.com. Please check your email (including spam folder).')
            # Redirect to verify code page
            return redirect(f"{reverse('verify_reset_code')}?email={user.email}")
        except Exception as exc:
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f'Failed to send verification code email: {exc}')
            messages.error(request, f'Failed to send email: {exc}. Please check your email configuration.')
            return redirect('forgot_password')
    
    return render(request, 'business/forgot_password.html')


def verify_reset_code(request):
    """Verify the password reset code sent via email"""
    email = request.GET.get('email', '')
    
    if request.method == 'POST':
        code = request.POST.get('code', '').strip()
        email = request.POST.get('email', email).strip()
        
        if not code or not email:
            messages.error(request, 'Both code and email are required.')
            return render(request, 'business/verify_reset_code.html', {'email': email})
        
        if len(code) != 6 or not code.isdigit():
            messages.error(request, 'Invalid code format. Please enter a 6-digit code.')
            return render(request, 'business/verify_reset_code.html', {'email': email})
        
        # Get stored code from session
        stored_code = request.session.get('reset_code')
        stored_user_id = request.session.get('reset_user_id')
        stored_email = request.session.get('reset_email')
        expires_str = request.session.get('reset_code_expires')
        
        if not stored_code or not stored_user_id:
            messages.error(request, 'No verification code found. Please request a new one.')
            return redirect('forgot_password')
        
        # Check expiration
        if expires_str:
            try:
                # Parse ISO format datetime string
                expires_at = datetime.fromisoformat(expires_str.replace('Z', '+00:00'))
                # Make timezone-aware if needed
                if timezone.is_naive(expires_at):
                    expires_at = timezone.make_aware(expires_at)
                if timezone.now() > expires_at:
                    messages.error(request, 'Verification code has expired. Please request a new one.')
                    # Clear session
                    request.session.pop('reset_code', None)
                    request.session.pop('reset_user_id', None)
                    request.session.pop('reset_email', None)
                    request.session.pop('reset_code_expires', None)
                    return redirect('forgot_password')
            except (ValueError, TypeError) as e:
                import logging
                logger = logging.getLogger(__name__)
                logger.warning(f'Error parsing expiration date: {e}')
                pass
        
        # Verify code matches
        if code != stored_code:
            messages.error(request, 'Invalid verification code. Please try again.')
            return render(request, 'business/verify_reset_code.html', {'email': email})
        
        # Verify email matches
        if email.lower() != stored_email.lower():
            messages.error(request, 'Email does not match. Please use the correct email.')
            return render(request, 'business/verify_reset_code.html', {'email': stored_email})
        
        # Code is valid, redirect to reset password
        messages.success(request, 'Verification code confirmed. Please set your new password.')
        return redirect('reset_password')
    
    return render(request, 'business/verify_reset_code.html', {'email': email})


def reset_password(request):
    """Reset password after verification code is confirmed"""
    user_id = request.session.get('reset_user_id')
    
    if not user_id:
        messages.error(request, 'Please verify your code first.')
        return redirect('forgot_password')
    
    try:
        user = User.objects.get(id=user_id, is_active=True)
    except User.DoesNotExist:
        messages.error(request, 'Invalid session. Please start over.')
        request.session.pop('reset_code', None)
        request.session.pop('reset_user_id', None)
        request.session.pop('reset_email', None)
        request.session.pop('reset_code_expires', None)
        return redirect('forgot_password')
    
    if request.method == 'POST':
        new_username = request.POST.get('username', '').strip()
        new_password = request.POST.get('password', '').strip()
        confirm_password = request.POST.get('confirm_password', '').strip()
        
        # Validate username if provided
        if new_username and new_username.lower() != user.username.lower():
            if len(new_username) < 3:
                messages.error(request, 'Username must be at least 3 characters long.')
                return render(request, 'business/reset_password.html', {'email': user.email, 'current_username': user.username})
            if User.objects.filter(username__iexact=new_username).exclude(id=user.id).exists():
                messages.error(request, 'Username is already taken. Please choose a different username.')
                return render(request, 'business/reset_password.html', {'email': user.email, 'current_username': user.username})
        
        if not new_password:
            messages.error(request, 'Password is required.')
            return render(request, 'business/reset_password.html', {'email': user.email})
        
        if len(new_password) < 8:
            messages.error(request, 'Password must be at least 8 characters long.')
            return render(request, 'business/reset_password.html', {'email': user.email})
        
        if new_password != confirm_password:
            messages.error(request, 'Passwords do not match.')
            return render(request, 'business/reset_password.html', {'email': user.email})
        
        # Update username if requested
        if new_username and new_username.lower() != user.username.lower():
            user.username = new_username
        
        # Set new password
        user.set_password(new_password)
        user.save()
        
        # Clear session
        request.session.pop('reset_code', None)
        request.session.pop('reset_user_id', None)
        request.session.pop('reset_email', None)
        request.session.pop('reset_code_expires', None)
        
        messages.success(request, 'Password has been reset successfully. You can now log in with your new password.')
        return redirect('login')
    
    return render(request, 'business/reset_password.html', {'email': user.email, 'current_username': user.username})


@login_required
def logout_view(request):
    logout(request)
    return redirect('login')


@login_required
def dashboard(request):
    # Get real-time statistics using database connection functions
    inventory_status = get_inventory_status()
    sales_data = calculate_actual_sales()
    
    # Get order statistics
    total_orders = Order.objects.count()
    pending_orders = Order.objects.filter(status='pending').count()
    completed_orders = Order.objects.filter(status='completed').count()
    in_progress_orders = Order.objects.filter(status='in_progress').count()
    total_products = inventory_status['total_products']
    low_stock_products = inventory_status['low_stock']
    
    # Sales data for the last 30 days (only completed orders)
    thirty_days_ago = timezone.now() - timedelta(days=30)
    recent_sales = Sales.objects.filter(
        created_at__gte=thirty_days_ago,
        order__status='completed'
    ).aggregate(total=Sum('amount'))['total'] or 0
    
    # Monthly sales data for chart (only completed orders)
    monthly_sales = []
    for i in range(12):
        month_start = timezone.now().replace(day=1) - timedelta(days=30*i)
        month_end = month_start + timedelta(days=30)
        month_sales = Sales.objects.filter(
            created_at__gte=month_start,
            created_at__lt=month_end,
            order__status='completed'
        ).aggregate(total=Sum('amount'))['total'] or 0
        monthly_sales.append({
            'month': month_start.strftime('%b'),
            'amount': float(month_sales)
        })
    
    # Recent orders
    recent_orders = Order.objects.order_by('-created_at')[:5]
    
    # Low stock products
    low_stock = Product.objects.filter(quantity__lte=models.F('min_quantity'))[:5]
    
    # Rental status data
    rental_available = inventory_status['rental_available']
    rental_rented = inventory_status['rental_rented']
    rental_overdue = inventory_status['rental_overdue']
    
    context = {
        'total_orders': total_orders,
        'pending_orders': pending_orders,
        'completed_orders': completed_orders,
        'in_progress_orders': in_progress_orders,
        'total_products': total_products,
        'low_stock_products': low_stock_products,
        'out_of_stock_products': inventory_status['out_of_stock'],
        'recent_sales': recent_sales,
        'monthly_sales': json.dumps(monthly_sales),
        'recent_orders': recent_orders,
        'low_stock': low_stock,
        'rental_available': rental_available,
        'rental_rented': rental_rented,
        'rental_overdue': rental_overdue,
        'inventory_status': inventory_status,
        'sales_data': sales_data,
    }
    
    return render(request, 'business/dashboard.html', context)


@login_required
def orders_list(request):
    """Orders list with real-time database connections and comprehensive tracking"""
    from django.db.models import Case, IntegerField, When

    # Get filter parameter
    order_type_filter = request.GET.get('type', 'all')
    
    # Base queryset with related data for better performance (exclude archived orders)
    # Custom ordering: Pending orders first, then active orders, then completed at bottom
    orders = Order.objects.filter(is_archived=False).select_related('customer', 'sales').prefetch_related('items__product', 'items__product__category').annotate(
        status_priority=Case(
            # Pending orders get priority 1 (appear first)
            When(status='pending', then=1),
            # Active/In-progress orders get priority 2
            When(status__in=['in_progress', 'rented', 'almost_due', 'due', 'overdue', 'ready_to_pick_up', 'repair_done', 'returned'], then=2),
            # Completed orders get priority 3 (appear at bottom)
            When(status='completed', then=3),
            # Cancelled orders get priority 4 (appear at very bottom)
            When(status='cancelled', then=4),
            default=2,  # Any other status gets priority 2
            output_field=IntegerField(),
        )
    ).order_by('status_priority', '-created_at')  # Sort by priority first, then by creation date (newest first within each priority)
    
    # Apply filter if specified
    # Map 'rental' to 'rent' since the model uses 'rent' as the order_type value
    if order_type_filter != 'all':
        if order_type_filter == 'rental':
            orders = orders.filter(order_type='rent')
        else:
            orders = orders.filter(order_type=order_type_filter)
    
    # Calculate statistics for pie chart
    completed_orders = orders.filter(status='completed').count()
    pending_orders = orders.filter(status='pending').count()
    in_progress_orders = orders.filter(status='in_progress').count()
    cancelled_orders = orders.filter(status='cancelled').count()
    
    # Calculate revenue statistics
    from datetime import timedelta

    from django.db.models import Sum
    
    total_revenue = orders.aggregate(total=Sum('total_amount'))['total'] or 0
    
    # Monthly revenue (current month) - use timezone-aware datetime
    current_month = timezone.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    monthly_revenue = orders.filter(
        created_at__gte=current_month
    ).aggregate(total=Sum('total_amount'))['total'] or 0
    
    # Daily revenue (today) - use timezone-aware datetime
    today_start = timezone.now().replace(hour=0, minute=0, second=0, microsecond=0)
    daily_revenue = orders.filter(
        created_at__gte=today_start
    ).aggregate(total=Sum('total_amount'))['total'] or 0
    
    # Weekly revenue (this week - Monday to Sunday) - use timezone-aware datetime
    today_date = timezone.now().date()
    days_since_monday = today_date.weekday()  # Monday is 0, Sunday is 6
    start_of_week_date = today_date - timedelta(days=days_since_monday)
    start_of_week_datetime = timezone.make_aware(
        datetime.combine(start_of_week_date, datetime.min.time())
    )
    weekly_revenue = orders.filter(
        created_at__gte=start_of_week_datetime
    ).aggregate(total=Sum('total_amount'))['total'] or 0
    
    # Yearly revenue (current year) - use timezone-aware datetime
    current_year = timezone.now().replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
    yearly_revenue = orders.filter(
        created_at__gte=current_year
    ).aggregate(total=Sum('total_amount'))['total'] or 0
    
    # Calculate order type statistics - get total counts before pagination
    # Note: Model uses 'rent' not 'rental' as the order_type value
    all_orders_count = Order.objects.filter(is_archived=False).count()
    rental_orders = Order.objects.filter(order_type='rent', is_archived=False).count()
    repair_orders = Order.objects.filter(order_type='repair', is_archived=False).count()
    customize_orders = Order.objects.filter(order_type='customize', is_archived=False).count()
    
    # Get filtered order count for active filter (total count before pagination)
    filtered_order_count = orders.count()
    
    # Get inventory status for order-related products
    inventory_status = get_inventory_status()
    
    # Get orders with inventory issues
    orders_with_issues = []
    for order in orders[:10]:  # Check first 10 orders for performance
        for item in order.items.all():
            if item.product.is_low_stock or (item.product.product_type == 'rental' and item.product.is_overdue):
                orders_with_issues.append({
                    'order': order,
                    'issue': 'Low stock' if item.product.is_low_stock else 'Overdue rental',
                    'product': item.product
                })
    
    # Get recent activity from activity log
    recent_activities = ActivityLog.objects.filter(
        activity_type__in=['order_created', 'order_updated', 'order_completed']
    ).order_by('-created_at')[:5]
    
    # Aggregate categories for each order (before pagination)
    # Convert queryset to list first to avoid multiple DB queries during iteration
    orders_list = list(orders)
    
    # Check and update overdue status for rent orders
    now = timezone.now()
    for order in orders_list:
        if order.order_type in ['rent', 'rental'] and order.status not in ['completed', 'returned', 'cancelled']:
            if order.due_date:
                # Check if order is overdue
                if now > order.due_date and order.status != 'overdue':
                    # Mark as overdue
                    order.status = 'overdue'
                    order.save(update_fields=['status'])
                # Check if order is due today
                elif order.is_due_today and order.status not in ['due', 'overdue']:
                    order.status = 'due'
                    order.save(update_fields=['status'])
                # Check if order is almost due (1 day before)
                elif order.is_one_day_before_due and order.status not in ['almost_due', 'due', 'overdue']:
                    order.status = 'almost_due'
                    order.save(update_fields=['status'])
    
    for order in orders_list:
        if order.order_type == 'repair':
            # Aggregate repair categories
            categories = []
            quantities = []
            
            # Helper function to format repair type name (must be defined before use)
            def format_repair_type(repair_type_str):
                """Format repair type from snake_case to Title Case"""
                if not repair_type_str:
                    return None
                # Remove "Repair - " prefix if present
                if "Repair - " in repair_type_str:
                    repair_type_str = repair_type_str[10:].strip()
                # Remove "repair - " prefix if present (lowercase)
                if "repair - " in repair_type_str.lower():
                    repair_type_str = repair_type_str[repair_type_str.lower().find("repair - ") + 9:].strip()
                # Remove "Repair Service" if present (shouldn't be, but just in case)
                repair_type_str = repair_type_str.replace("Repair Service", "").replace("repair service", "").strip()
                # Remove class suffix like "(Class standard)"
                if " (Class " in repair_type_str:
                    repair_type_str = repair_type_str.split(" (Class ")[0].strip()
                # Handle common repair types
                repair_type_lower = repair_type_str.lower().strip()
                # If it's empty after cleaning, return None
                if not repair_type_lower:
                    return None
                repair_type_map = {
                    'zipper': 'Zipper',
                    'zipper_repair': 'Zipper',
                    'zipper_replacement': 'Zipper',
                    'buttons': 'Buttons',
                    'button': 'Buttons',
                    'buttons_repair': 'Buttons',
                    'patch': 'Patch',
                    'patches': 'Patch',
                    'patch_repair': 'Patch',
                    'lock': 'Lock',
                    'locks': 'Lock',
                    'lock_repair': 'Lock',
                    'garter': 'Garter',
                    'garter_repair': 'Garter',
                    'elastic': 'Elastic',
                    'elastic_repair': 'Elastic',
                    'bewang': 'Bewang',
                    'bewang_repair': 'Bewang',
                    'putol': 'Putol',
                    'baston': 'Baston',
                    'suklot': 'Suklot',
                    'baston_suklot': 'Baston Suklot',
                    'baston_putol': 'Baston Putol',
                    'ambel': 'Ambel',
                    'pasada': 'Pasada',
                }
                if repair_type_lower in repair_type_map:
                    return repair_type_map[repair_type_lower]
                # Check if it contains any of the repair types
                for key, value in repair_type_map.items():
                    if key in repair_type_lower:
                        return value
                # Replace underscores with spaces and title case
                formatted = repair_type_str.replace('_', ' ').title().strip()
                # Return the formatted string (don't return "Repair Service" - that's a fallback)
                if formatted and formatted.lower() != "repair service" and formatted.lower() != "service":
                    return formatted
                # If it's empty or just "Service", return None so we can try other methods
                return None
            
            # FIRST: Try to extract repair type from inventory transaction notes (most reliable for historical data)
            repair_type_from_transactions = None
            try:
                from .models import InventoryTransaction
                transactions = InventoryTransaction.objects.filter(reference_order=order).order_by('-created_at')[:5]
                for transaction in transactions:
                    if transaction.notes:
                        notes_lower = transaction.notes.lower()
                        # Look for patterns like "used for zipper repair" or "for zipper repair"
                        repair_types_in_transactions = ['zipper', 'buttons', 'button', 'patch', 'patches', 'lock', 'locks', 'garter', 'elastic', 'bewang', 'zipper_replacement', 'button_repair', 'lock_repair', 'putol', 'baston', 'suklot', 'ambel', 'pasada', 'general_repair', 'general_tshirt_repair']
                        for repair_type in repair_types_in_transactions:
                            if repair_type in notes_lower:
                                repair_type_from_transactions = repair_type
                                break
                        if repair_type_from_transactions:
                            break
            except Exception:
                repair_type_from_transactions = None
            
            # Format the extracted repair type
            if repair_type_from_transactions:
                repair_type_from_transactions = format_repair_type(repair_type_from_transactions)
            
            # SECOND: Try to extract repair type from order notes
            repair_type_from_notes = None
            if order.notes:
                notes_lower = order.notes.lower()
                # Check for common repair types in notes
                repair_types_in_notes = ['zipper', 'buttons', 'button', 'patch', 'patches', 'lock', 'locks', 'garter', 'elastic', 'bewang', 'zipper_replacement', 'button_repair', 'lock_repair', 'putol', 'baston', 'suklot', 'ambel', 'pasada', 'general_repair', 'general_tshirt_repair']
                for repair_type in repair_types_in_notes:
                    if repair_type in notes_lower:
                        repair_type_from_notes = format_repair_type(repair_type)
                        break
            
            for item in order.items.all():
                category_name = None
                
                # METHOD 1: Try to extract from inventory transactions first (most reliable)
                if repair_type_from_transactions:
                    category_name = repair_type_from_transactions
                # METHOD 2: Try to extract from order notes
                elif repair_type_from_notes:
                    category_name = repair_type_from_notes
                else:
                    # METHOD 2: Try to get repair type from product name
                    if item.product and item.product.name:
                        product_name = item.product.name
                        product_name_lower = product_name.lower()
                        
                        # Check for "Repair - " or "repair - " prefix
                        if "Repair - " in product_name or "repair - " in product_name_lower:
                            # Extract the repair type part
                            if "Repair - " in product_name:
                                repair_part = product_name.split("Repair - ")[1].strip()
                            else:
                                repair_part = product_name.split("repair - ")[1].strip()
                            # Remove class suffix if present
                            if " (Class " in repair_part:
                                repair_part = repair_part.split(" (Class ")[0].strip()
                            # Remove timestamp if present
                            if " - " in repair_part and len(repair_part.split(" - ")[-1]) > 10:
                                repair_part = repair_part.split(" - ")[0].strip()
                            category_name = format_repair_type(repair_part)
                        else:
                            # Check if product name itself is a repair type (common repair types)
                            repair_types_in_name = ['zipper', 'buttons', 'button', 'patch', 'patches', 'lock', 'locks', 'garter', 'elastic', 'bewang', 'zipper_replacement', 'button_repair', 'lock_repair', 'putol', 'baston', 'suklot', 'ambel', 'pasada', 'general_repair', 'general_tshirt_repair']
                            for repair_type in repair_types_in_name:
                                if repair_type in product_name_lower:
                                    category_name = format_repair_type(repair_type)
                                    break
                            
                            # If not found, try to format the entire product name (might be just the repair type)
                            if not category_name:
                                # Remove common prefixes/suffixes and format
                                cleaned_name = product_name.strip()
                                # Remove "Repair Service" if present
                                cleaned_name = cleaned_name.replace("Repair Service", "").replace("repair service", "").strip()
                                # Remove "Service" if at the end
                                if cleaned_name.lower().endswith(" service"):
                                    cleaned_name = cleaned_name[:-8].strip()
                                if cleaned_name:
                                    category_name = format_repair_type(cleaned_name)
                                else:
                                    category_name = format_repair_type(product_name)
                    
                    # METHOD 3: If still not found, try product description
                    if not category_name or category_name == "Repair Service":
                        if item.product and item.product.description:
                            desc = item.product.description.lower()
                            # Check for repair type in description
                            repair_types = ['zipper', 'buttons', 'button', 'patch', 'patches', 'lock', 'locks', 'garter', 'elastic', 'bewang', 'zipper_replacement', 'button_repair', 'lock_repair', 'putol', 'baston', 'suklot', 'ambel', 'pasada', 'general_repair', 'general_tshirt_repair']
                            for repair_type in repair_types:
                                if repair_type in desc:
                                    category_name = format_repair_type(repair_type)
                                    break
                    
                    # METHOD 4: Try product category
                    if not category_name or category_name == "Repair Service":
                        if item.product and item.product.category:
                            cat_name = item.product.category.name
                            # Check if category name contains repair type
                            cat_name_lower = cat_name.lower()
                            repair_types = ['zipper', 'buttons', 'button', 'patch', 'patches', 'lock', 'locks', 'garter', 'elastic', 'bewang', 'putol', 'baston', 'suklot', 'ambel', 'pasada', 'general_repair', 'general_tshirt_repair']
                            for repair_type in repair_types:
                                if repair_type in cat_name_lower:
                                    category_name = format_repair_type(repair_type)
                                    break
                            if not category_name or category_name == "Repair Service":
                                category_name = format_repair_type(cat_name)
                    
                    # METHOD 5: Check order notes if not already checked
                    if (not category_name or category_name == "Repair Service") and order.notes:
                        notes_lower = order.notes.lower()
                        repair_types = ['zipper', 'buttons', 'button', 'patch', 'patches', 'lock', 'locks', 'garter', 'elastic', 'bewang', 'zipper_replacement', 'button_repair', 'lock_repair', 'putol', 'baston', 'suklot', 'ambel', 'pasada', 'general_repair', 'general_tshirt_repair']
                        for repair_type in repair_types:
                            if repair_type in notes_lower:
                                category_name = format_repair_type(repair_type)
                                break
                
                # Final fallback - try to extract from product name directly
                if not category_name or category_name == "Repair Service":
                    if item.product and item.product.name:
                        product_name = item.product.name
                        # Try to extract anything after "Repair - " or "repair - "
                        if "Repair - " in product_name:
                            extracted = product_name.split("Repair - ")[1].strip()
                            # Remove class suffix and timestamp
                            if " (Class " in extracted:
                                extracted = extracted.split(" (Class ")[0].strip()
                            if " - " in extracted and len(extracted.split(" - ")[-1]) > 10:
                                extracted = extracted.split(" - ")[0].strip()
                            if extracted and extracted != "Repair Service":
                                category_name = format_repair_type(extracted)
                        elif "repair - " in product_name.lower():
                            extracted = product_name.lower().split("repair - ")[1].strip()
                            if " (class " in extracted:
                                extracted = extracted.split(" (class ")[0].strip()
                            if " - " in extracted and len(extracted.split(" - ")[-1]) > 10:
                                extracted = extracted.split(" - ")[0].strip()
                            if extracted and extracted != "repair service":
                                category_name = format_repair_type(extracted)
                
                # Only use "Repair Service" as absolute last resort
                if not category_name:
                    category_name = "Repair Service"
                
                if category_name:
                    # Group by category and sum quantities
                    found = False
                    for i, cat in enumerate(categories):
                        if cat == category_name:
                            quantities[i] += item.quantity
                            found = True
                            break
                    if not found:
                        categories.append(category_name)
                        quantities.append(item.quantity)
            
            # Format as "Category1 (qty), Category2 (qty)"
            if categories:
                category_parts = []
                for cat, qty in zip(categories, quantities):
                    if qty > 1:
                        category_parts.append(f"{cat} ({qty})")
                    else:
                        category_parts.append(cat)
                order.repair_categories_display = ", ".join(category_parts)
                order.repair_quantity_display = sum(quantities)
            else:
                order.repair_categories_display = "Repair Service"
                order.repair_quantity_display = 1
        
        elif order.order_type == 'customize':
            # Aggregate customize categories
            categories = []
            quantities = []
            
            # Helper function to format customize type name (must be defined before use)
            def format_customize_type(customize_type_str):
                """Format customize type name"""
                if not customize_type_str:
                    return None
                # Remove "Customize - " prefix if present
                if "Customize - " in customize_type_str:
                    customize_type_str = customize_type_str[11:].strip()
                # Remove "customize - " prefix if present (lowercase)
                if "customize - " in customize_type_str.lower():
                    customize_type_str = customize_type_str[customize_type_str.lower().find("customize - ") + 12:].strip()
                # Remove "Customize Service" if present (shouldn't be, but just in case)
                customize_type_str = customize_type_str.replace("Customize Service", "").replace("customize service", "").strip()
                # Remove class suffix like "(Class standard)"
                if " (Class " in customize_type_str:
                    customize_type_str = customize_type_str.split(" (Class ")[0].strip()
                # Remove timestamp suffix if present (e.g., " - 20251117 121237")
                if " - " in customize_type_str and len(customize_type_str.split(" - ")[-1]) > 10:
                    customize_type_str = customize_type_str.split(" - ")[0].strip()
                # Handle common customize types
                customize_type_lower = customize_type_str.lower().strip()
                # If it's empty after cleaning, return None
                if not customize_type_lower:
                    return None
                customize_type_map = {
                    'polo': 'Polo',
                    'polo_shirt': 'Polo Shirt',
                    'blouse': 'Blouse',
                    'pants': 'Pants',
                    'shorts': 'Shorts',
                    'skirt_palda': 'Skirt/Palda',
                    'skirt': 'Skirt/Palda',
                    'palda': 'Skirt/Palda',
                    'uniform': 'Uniform',
                    'pe': 'PE',
                }
                if customize_type_lower in customize_type_map:
                    return customize_type_map[customize_type_lower]
                # Check if it contains any of the customize types
                for key, value in customize_type_map.items():
                    if key in customize_type_lower:
                        return value
                # Replace underscores with spaces and title case
                formatted = customize_type_str.replace('_', ' ').title().strip()
                # Return the formatted string (don't return "Customize Service" - that's a fallback)
                if formatted and formatted.lower() != "customize service" and formatted.lower() != "service":
                    return formatted
                # If it's empty or just "Service", return None so we can try other methods
                return None
            
            # FIRST: Try to extract customize type from order notes
            customize_type_from_notes = None
            if order.notes:
                notes_lower = order.notes.lower()
                # Check for common customize types in notes
                customize_types_in_notes = ['polo', 'polo_shirt', 'blouse', 'pants', 'shorts', 'skirt', 'palda', 'skirt_palda', 'uniform', 'pe']
                for customize_type in customize_types_in_notes:
                    if customize_type in notes_lower:
                        customize_type_from_notes = format_customize_type(customize_type)
                        break
            
            for item in order.items.all():
                category_name = None
                
                # METHOD 1: Try to extract from order notes first
                if customize_type_from_notes:
                    category_name = customize_type_from_notes
                else:
                    # METHOD 2: Try to get customize type from product name
                    if item.product and item.product.name:
                        product_name = item.product.name
                        product_name_lower = product_name.lower()
                        
                        # Check for "Customize - " or "customize - " prefix
                        if "Customize - " in product_name or "customize - " in product_name_lower:
                            # Extract the customize type part
                            if "Customize - " in product_name:
                                customize_part = product_name.split("Customize - ")[1].strip()
                            else:
                                customize_part = product_name.split("customize - ")[1].strip()
                            # Remove class suffix if present
                            if " (Class " in customize_part:
                                customize_part = customize_part.split(" (Class ")[0].strip()
                            # Remove timestamp suffix if present
                            if " - " in customize_part and len(customize_part.split(" - ")[-1]) > 10:
                                customize_part = customize_part.split(" - ")[0].strip()
                            category_name = format_customize_type(customize_part)
                        else:
                            # Check if product name itself contains customize type
                            customize_types_in_name = ['polo', 'polo_shirt', 'blouse', 'pants', 'shorts', 'skirt', 'palda', 'skirt_palda', 'uniform', 'pe']
                            for customize_type in customize_types_in_name:
                                if customize_type in product_name_lower:
                                    category_name = format_customize_type(customize_type)
                                    break
                            
                            # If not found, try to format the entire product name (might be just the customize type)
                            if not category_name:
                                # Remove common prefixes/suffixes and format
                                cleaned_name = product_name.strip()
                                # Remove "Customize Service" if present
                                cleaned_name = cleaned_name.replace("Customize Service", "").replace("customize service", "").strip()
                                # Remove "Service" if at the end
                                if cleaned_name.lower().endswith(" service"):
                                    cleaned_name = cleaned_name[:-8].strip()
                                if cleaned_name:
                                    category_name = format_customize_type(cleaned_name)
                                else:
                                    category_name = format_customize_type(product_name)
                    
                    # METHOD 3: Try product description
                    if not category_name or category_name == "Customize Service":
                        if item.product and item.product.description:
                            desc = item.product.description.lower()
                            # Check for "Type: " in description
                            if "type: " in desc or "type:" in desc:
                                if "type: uniform" in desc:
                                    category_name = "Uniform"
                                elif "type: pe" in desc or "type:pe" in desc:
                                    category_name = "PE"
                                else:
                                    # Try to extract type from description
                                    customize_types = ['polo', 'polo_shirt', 'blouse', 'pants', 'shorts', 'skirt', 'palda', 'skirt_palda', 'uniform', 'pe']
                                    for customize_type in customize_types:
                                        if customize_type in desc:
                                            category_name = format_customize_type(customize_type)
                                            break
                    
                    # METHOD 4: Try product category
                    if not category_name or category_name == "Customize Service":
                        if item.product and item.product.category:
                            cat_name = item.product.category.name
                            cat_name_lower = cat_name.lower()
                            # Check if category name contains customize type
                            customize_types = ['polo', 'polo_shirt', 'blouse', 'pants', 'shorts', 'skirt', 'palda', 'skirt_palda', 'uniform', 'pe']
                            for customize_type in customize_types:
                                if customize_type in cat_name_lower:
                                    category_name = format_customize_type(customize_type)
                                    break
                            if not category_name or category_name == "Customize Service":
                                category_name = format_customize_type(cat_name)
                    
                    # METHOD 5: Check order notes if not already checked
                    if (not category_name or category_name == "Customize Service") and order.notes:
                        notes_lower = order.notes.lower()
                        customize_types = ['polo', 'polo_shirt', 'blouse', 'pants', 'shorts', 'skirt', 'palda', 'skirt_palda', 'uniform', 'pe']
                        for customize_type in customize_types:
                            if customize_type in notes_lower:
                                category_name = format_customize_type(customize_type)
                                break
                
                # Final fallback - try to extract from product name directly
                if not category_name or category_name == "Customize Service":
                    if item.product and item.product.name:
                        product_name = item.product.name
                        # Try to extract anything after "Customize - " or "customize - "
                        if "Customize - " in product_name:
                            extracted = product_name.split("Customize - ")[1].strip()
                            # Remove class suffix and timestamp
                            if " (Class " in extracted:
                                extracted = extracted.split(" (Class ")[0].strip()
                            if " - " in extracted and len(extracted.split(" - ")[-1]) > 10:
                                extracted = extracted.split(" - ")[0].strip()
                            if extracted and extracted != "Customize Service":
                                category_name = format_customize_type(extracted)
                        elif "customize - " in product_name.lower():
                            extracted = product_name.lower().split("customize - ")[1].strip()
                            if " (class " in extracted:
                                extracted = extracted.split(" (class ")[0].strip()
                            if " - " in extracted and len(extracted.split(" - ")[-1]) > 10:
                                extracted = extracted.split(" - ")[0].strip()
                            if extracted and extracted != "customize service":
                                category_name = format_customize_type(extracted)
                
                # Only use "Customize Service" as absolute last resort
                if not category_name:
                    category_name = "Customize Service"
                
                if category_name:
                    # Group by category and sum quantities
                    found = False
                    for i, cat in enumerate(categories):
                        if cat == category_name:
                            quantities[i] += item.quantity
                            found = True
                            break
                    if not found:
                        categories.append(category_name)
                        quantities.append(item.quantity)
            
            # Format as "Category1 (qty), Category2 (qty)"
            if categories:
                category_parts = []
                for cat, qty in zip(categories, quantities):
                    if qty > 1:
                        category_parts.append(f"{cat} ({qty})")
                    else:
                        category_parts.append(cat)
                order.customize_categories_display = ", ".join(category_parts)
                order.customize_quantity_display = sum(quantities)
            else:
                order.customize_categories_display = "Customize Service"
                order.customize_quantity_display = 1
        
        elif order.order_type == 'rent' or order.order_type == 'rental':
            # Aggregate rental categories from actual order items
            categories = []
            quantities = []
            
            for item in order.items.all():
                category_name = None
                
                # Get category from product
                if item.product and item.product.category:
                    category_name = item.product.category.name
                elif item.product and item.product.name:
                    # Try to infer from product name
                    product_name = item.product.name
                    # Check if it's a rental product with category info
                    if item.product.product_type == 'rental':
                        # Use product name as fallback
                        category_name = product_name
                    else:
                        category_name = "Rental Service"
                else:
                    category_name = "Rental Service"
                
                # Ensure category_name is never None or empty
                if not category_name or category_name.strip() == '' or category_name == 'N/A':
                    category_name = "Rental Service"
                
                # Group by category and sum quantities
                found = False
                for i, cat in enumerate(categories):
                    if cat == category_name:
                        quantities[i] += item.quantity
                        found = True
                        break
                if not found:
                    categories.append(category_name)
                    quantities.append(item.quantity)
            
            # Format as "Category1, Category2" or "Category1 (qty), Category2 (qty)"
            if categories:
                category_parts = []
                for cat, qty in zip(categories, quantities):
                    if qty > 1:
                        category_parts.append(f"{cat} ({qty})")
                    else:
                        category_parts.append(cat)
                order.rental_categories_display = ", ".join(category_parts)
                order.rental_quantity_display = sum(quantities)
            else:
                # Fallback: check if order has any items at all
                if order.items.exists():
                    order.rental_categories_display = "Rental Service"
                    order.rental_quantity_display = sum(item.quantity for item in order.items.all())
                else:
                    order.rental_categories_display = "Rental Service"
                    order.rental_quantity_display = 1
    
    # Pagination for orders
    paginator = Paginator(orders_list, 10)  # Show 10 orders per page
    page = request.GET.get('page')
    try:
        orders = paginator.page(page)
    except PageNotAnInteger:
        orders = paginator.page(1)
    except EmptyPage:
        orders = paginator.page(paginator.num_pages)
    
    # Get staff members for assignment (only active staff users)
    from django.contrib.auth.models import User
    staff_members = User.objects.filter(is_staff=True, is_active=True).order_by('first_name', 'username')
    
    # Calculate busy status for each staff member
    now = timezone.now()
    staff_with_status = []
    for staff in staff_members:
        active_orders = Order.objects.filter(
            assigned_staff=staff
        ).exclude(status__in=['completed', 'cancelled'])
        
        is_busy = active_orders.exists()
        time_since_assignment = None
        if is_busy:
            most_recent_order = active_orders.order_by('-staff_assigned_at').first()
            if most_recent_order and most_recent_order.staff_assigned_at:
                time_diff = now - most_recent_order.staff_assigned_at
                time_since_assignment = int(time_diff.total_seconds() / 60)
        
        staff_with_status.append({
            'staff': staff,
            'is_busy': is_busy,
            'time_since_assignment': time_since_assignment,
            'active_order_count': active_orders.count(),
        })
    
    context = {
        'orders': orders,
        'completed_orders': completed_orders,
        'pending_orders': pending_orders,
        'in_progress_orders': in_progress_orders,
        'cancelled_orders': cancelled_orders,
        'total_revenue': total_revenue,
        'yearly_revenue': yearly_revenue,
        'monthly_revenue': monthly_revenue,
        'staff_with_status': staff_with_status,
        'weekly_revenue': weekly_revenue,
        'daily_revenue': daily_revenue,
        'all_orders_count': all_orders_count,
        'rental_orders': rental_orders,
        'repair_orders': repair_orders,
        'customize_orders': customize_orders,
        'filtered_order_count': filtered_order_count,
        'current_filter': order_type_filter,
        'inventory_status': inventory_status,
        'orders_with_issues': orders_with_issues,
        'recent_activities': recent_activities,
        'staff_members': staff_members,
    }
    
    return render(request, 'business/orders.html', context)


@login_required
def order_detail(request, order_id):
    """Display order details"""
    from django.shortcuts import get_object_or_404

    from .rental_manager import RentalStatusManager
    
    order = get_object_or_404(Order, id=order_id)
    order_items = order.items.all().select_related('product')
    
    # Check if order has rental items
    has_rental_items = any(item.product.product_type == 'rental' for item in order_items)
    
    # Get rental status if applicable
    rental_status = None
    if has_rental_items:
        rental_status = RentalStatusManager.get_order_rental_status(order)
    
    # Get order activities
    order_activities = ActivityLog.objects.filter(
        order=order
    ).order_by('-created_at')[:20]
    
    # Get staff members for assignment (only active staff users)
    from django.contrib.auth.models import User
    staff_members = User.objects.filter(is_staff=True, is_active=True).order_by('first_name', 'username')
    
    context = {
        'order': order,
        'order_items': order_items,
        'total_items': order_items.count(),
        'has_rental_items': has_rental_items,
        'rental_status': rental_status or {'overdue_items': 0, 'total_items': 0},
        'order_activities': order_activities,
        'staff_members': staff_members,
    }
    
    return render(request, 'business/order_detail.html', context)

@login_required
@require_http_methods(["POST"])
def assign_staff_to_order(request, order_id):
    """Assign staff to a repair or customize order"""
    try:
        import json

        from django.contrib.auth.models import User
        from django.shortcuts import get_object_or_404
        
        order = get_object_or_404(Order, id=order_id)
        
        # Only allow assignment for repair and customize orders
        if order.order_type not in ['repair', 'customize']:
            return JsonResponse({
                'success': False,
                'error': 'Staff can only be assigned to Repair or Customize orders'
            })
        
        # Parse request data - handle both JSON and form data
        if request.content_type and 'application/json' in request.content_type:
            try:
                data = json.loads(request.body)
                staff_id = data.get('staff_id')
                notes = data.get('notes', '')
            except (json.JSONDecodeError, ValueError) as e:
                return JsonResponse({
                    'success': False,
                    'error': f'Invalid JSON data: {str(e)}'
                })
        else:
            # Handle form data (application/x-www-form-urlencoded)
            staff_id = request.POST.get('staff_id')
            notes = request.POST.get('notes', '')
        
        if not staff_id:
            return JsonResponse({
                'success': False,
                'error': 'Staff ID is required'
            })
        
        # Get staff user
        staff_user = get_object_or_404(User, id=staff_id, is_staff=True, is_active=True)
        
        # Check if staff is already busy (has active assigned orders)
        active_orders = Order.objects.filter(
            assigned_staff=staff_user
        ).exclude(status__in=['completed', 'cancelled'])
        
        if active_orders.exists():
            return JsonResponse({
                'success': False,
                'error': f'Staff {staff_user.get_full_name() or staff_user.username} is currently busy with {active_orders.count()} active order(s). Please wait for them to complete their current assignment or assign to another available staff member.'
            })
        
        # Assign staff to order
        order.assigned_staff = staff_user
        order.staff_assigned_at = timezone.now()  # Track assignment time
        order.save()
        
        # Log the assignment
        ActivityLog.objects.create(
            order=order,
            activity_type='staff_assigned',
            description=f'Staff {staff_user.get_full_name() or staff_user.username} assigned to order {order.order_identifier or order.id}' + (f' - Note: {notes}' if notes else '')
        )
        
        return JsonResponse({
            'success': True,
            'message': f'Staff {staff_user.get_full_name() or staff_user.username} assigned successfully',
            'staff_name': staff_user.get_full_name() or staff_user.username
        })
        
    except User.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Selected staff member not found'
        })
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error assigning staff to order: {e}")
        return JsonResponse({
            'success': False,
            'error': str(e)
        })

@login_required
@require_http_methods(["POST"])
def mark_order_done_by_staff(request, order_id):
    """Mark an order as done by the assigned staff member"""
    if request.method != 'POST':
        return JsonResponse({
            'success': False,
            'error': 'Only POST method is allowed'
        }, status=405)
    
    try:
        order = get_object_or_404(Order, id=order_id)
        
        # Check if order is already completed
        if order.status == 'completed':
            return JsonResponse({
                'success': False,
                'error': 'This order is already marked as completed'
            }, status=400)
        
        # Check if order has assigned staff
        if not order.assigned_staff:
            return JsonResponse({
                'success': False,
                'error': 'No staff assigned to this order'
            }, status=400)
        
        # Only allow for repair and customize orders
        if order.order_type not in ['repair', 'customize']:
            return JsonResponse({
                'success': False,
                'error': 'This action is only available for Repair or Customize orders'
            }, status=400)
        
        # Mark order as done by staff
        order.staff_completed_at = timezone.now()
        # Mark order as completed (this will trigger sales record creation via signal)
        order.status = 'completed'
        order.save()
        
        # Log the completion
        try:
            ActivityLog.objects.create(
                order=order,
                activity_type='staff_completed',
                description=f'Staff {order.assigned_staff.get_full_name() or order.assigned_staff.username} marked order {order.order_identifier or order.id} as done'
            )
        except Exception as log_error:
            import logging
            logger = logging.getLogger(__name__)
            logger.warning(f"Failed to create activity log: {log_error}")
        
        # Calculate time taken
        time_taken = None
        if order.staff_assigned_at and order.staff_completed_at:
            time_diff = order.staff_completed_at - order.staff_assigned_at
            time_taken_minutes = int(time_diff.total_seconds() / 60)
            time_taken = f"{time_taken_minutes} minutes"
        
        # Calculate updated salary for the staff member
        from decimal import Decimal

        from django.db.models import Sum
        
        try:
            completed_repair_customize_orders = Order.objects.filter(
                assigned_staff=order.assigned_staff,
                status='completed',
                order_type__in=['repair', 'customize']
            )
            total_revenue = completed_repair_customize_orders.aggregate(
                total=Sum('total_amount')
            )['total'] or Decimal('0')
            staff_salary = total_revenue * Decimal('0.4')  # 40% to staff
            owner_share = total_revenue * Decimal('0.6')  # 60% to owner
            completed_count = completed_repair_customize_orders.count()
        except Exception as calc_error:
            import logging
            logger = logging.getLogger(__name__)
            logger.warning(f"Failed to calculate salary: {calc_error}")
            total_revenue = Decimal('0')
            staff_salary = Decimal('0')
            owner_share = Decimal('0')
            completed_count = 0
        
        return JsonResponse({
            'success': True,
            'message': f'Order marked as done by {order.assigned_staff.get_full_name() or order.assigned_staff.username}',
            'time_taken': time_taken,
            'staff_id': order.assigned_staff.id,
            'is_available': True,
            'salary_data': {
                'completed_orders': completed_count,
                'total_revenue': float(total_revenue),
                'staff_salary': float(staff_salary),
                'owner_share': float(owner_share)
            }
        })
        
    except Order.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Order not found'
        }, status=404)
    except Exception as e:
        import logging
        import traceback
        logger = logging.getLogger(__name__)
        logger.error(f"Error marking order as done: {e}")
        logger.error(traceback.format_exc())
        return JsonResponse({
            'success': False,
            'error': f'An error occurred: {str(e)}'
        }, status=500)

@login_required
def create_order(request):
    if request.method == 'POST':
        form = OrderForm(request.POST)
        if form.is_valid():
            order = form.save(commit=False)
            order.save()
            return redirect('order_items', order_id=order.id)
    else:
        form = OrderForm()
    
    return render(request, 'business/create_order.html', {'form': form})


@login_required
def payment_method(request):
    return render(request, 'business/payment_method.html')

@login_required
def payment_process(request):
    """View for payment processing (like grocery store POS)"""
    return render(request, 'business/payment_process.html')

@login_required
def track_order_qr(request):
    """View for tracking order via QR code scan"""
    if request.method == 'POST':
        try:
            import json
            qr_data = json.loads(request.body)
            
            # Extract order ID from QR data - support multiple formats
            order_id = (qr_data.get('id') or qr_data.get('orderId') or 
                       qr_data.get('order_id') or qr_data.get('order_identifier'))
            
            if order_id:
                # Use database connection function to get comprehensive tracking data
                tracking_data = get_order_tracking_data(order_id)
                if tracking_data:
                    context = {
                        'tracking_data': tracking_data,
                        'qr_data': qr_data, 
                        'found': True,
                        'order': tracking_data['order'],
                        'customer': tracking_data['customer'],
                        'items': tracking_data['items'],
                        'inventory_status': tracking_data['inventory_status'],
                        'sales_info': tracking_data['sales_info']
                    }
                else:
                    # Order not found in database, show QR data only
                    context = {
                        'qr_data': qr_data,
                        'found': False,
                        'order_id': order_id
                    }
            else:
                context = {
                    'error': 'Invalid QR code data',
                    'found': False
                }
                
        except json.JSONDecodeError:
            context = {
                'error': 'Invalid QR code format',
                'found': False
            }
        except Exception as e:
            context = {
                'error': f'Error processing QR code: {str(e)}',
                'found': False
            }
    else:
        context = {
            'error': 'No QR code data provided',
            'found': False
        }
    
    return render(request, 'business/track_order.html', context)


@login_required
def api_decode_qr_image(request):
    """API endpoint to decode QR code from uploaded image"""
    if request.method == 'POST':
        try:
            # Get uploaded file
            if 'image' not in request.FILES:
                return JsonResponse({
                    'success': False,
                    'error': 'No image file provided'
                })
            
            image_file = request.FILES['image']
            
            # Validate file type
            if not image_file.content_type.startswith('image/'):
                return JsonResponse({
                    'success': False,
                    'error': 'Invalid file type. Please upload an image file.'
                })
            
            # Validate file size (max 5MB)
            if image_file.size > 5 * 1024 * 1024:
                return JsonResponse({
                    'success': False,
                    'error': 'File size must be less than 5MB'
                })
            
            # Try to decode QR code using pyzbar (if available)
            try:
                from PIL import Image
                try:
                    from pyzbar.pyzbar import decode
                except ImportError:
                    # Fallback if pyzbar is not installed
                    decode = None

                if decode:
                    # Read and process image
                    image = Image.open(io.BytesIO(image_file.read()))
                    
                    # Try to decode QR code using pyzbar
                    decoded_objects = decode(image)
                else:
                    decoded_objects = None
                
                if decoded_objects:
                    # Get the first decoded QR code
                    qr_data = decoded_objects[0].data.decode('utf-8')
                    
                    # Try to parse as JSON
                    try:
                        qr_json = json.loads(qr_data)
                        order_id = (qr_json.get('id') or qr_json.get('orderId') or 
                                   qr_json.get('order_id') or qr_json.get('order_identifier'))
                        
                        return JsonResponse({
                            'success': True,
                            'qr_data': qr_json,
                            'order_id': order_id,
                            'raw_data': qr_data
                        })
                    except json.JSONDecodeError:
                        # If not JSON, treat as plain order ID
                        return JsonResponse({
                            'success': True,
                            'order_id': qr_data,
                            'raw_data': qr_data
                        })
                else:
                    return JsonResponse({
                        'success': False,
                        'error': 'No QR code found in the image. Please ensure the image contains a valid QR code.'
                    })
            except ImportError:
                # pyzbar not available, return error suggesting frontend decoding
                return JsonResponse({
                    'success': False,
                    'error': 'Server-side QR decoding not available. Please use the frontend QR scanner.'
                })
            except Exception as decode_error:
                return JsonResponse({
                    'success': False,
                    'error': f'Failed to decode QR code: {str(decode_error)}'
                })
                
        except Exception as e:
            import traceback
            traceback.print_exc()
            return JsonResponse({
                'success': False,
                'error': f'Error processing image: {str(e)}'
            })
    
    return JsonResponse({
        'success': False,
        'error': 'Invalid request method'
    })

@login_required
def activity_log(request):
    """View to display system activity log"""
    from datetime import timedelta

    from django.core.paginator import Paginator
    from django.utils import timezone

    # Start with all activities ordered by most recent first
    activities = ActivityLog.objects.all().order_by('-created_at')
    
    # Filter by activity type if provided
    activity_type = request.GET.get('type')
    if activity_type:
        activities = activities.filter(activity_type=activity_type)
    
    # Filter by date range if provided
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    
    if date_from:
        activities = activities.filter(created_at__date__gte=date_from)
    if date_to:
        activities = activities.filter(created_at__date__lte=date_to)
    
    # Calculate statistics (before pagination)
    total_activities = ActivityLog.objects.count()
    
    # Recent activities (last 7 days)
    seven_days_ago = timezone.now() - timedelta(days=7)
    recent_activities = ActivityLog.objects.filter(created_at__gte=seven_days_ago).count()
    
    # Today's activities
    today_start = timezone.now().replace(hour=0, minute=0, second=0, microsecond=0)
    today_activities = ActivityLog.objects.filter(created_at__gte=today_start).count()
    
    # Paginate with 5 items per page
    paginator = Paginator(activities, 5)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    context = {
        'activities': page_obj,
        'activity_types': ActivityLog.ACTIVITY_TYPES,
        'selected_type': activity_type,
        'date_from': date_from,
        'date_to': date_to,
        'total_activities': total_activities,
        'recent_activities': recent_activities,
        'today_activities': today_activities,
    }
    
    return render(request, 'business/activity_log.html', context)


# ==================== DATABASE CONNECTION FUNCTIONS ====================

def validate_products_exist(product_names, product_type=None, order_type=None):
    """
    Validate that all products exist in the database before use.
    Prevents auto-creation of static/hardcoded data.
    
    Args:
        product_names: List of product names or dicts with 'name' key
        product_type: Optional filter by product type
        order_type: Optional order type for context
    
    Returns:
        dict with 'valid': bool, 'missing_products': list, 'existing_products': list
    """
    missing_products = []
    existing_products = []
    
    # Normalize product_names to list of strings
    if isinstance(product_names, str):
        product_names = [product_names]
    elif isinstance(product_names, dict):
        product_names = [p.get('name', '') if isinstance(p, dict) else str(p) for p in product_names]
    
    for product_name in product_names:
        if not product_name or not product_name.strip():
            continue
            
        # Build query
        query = Product.objects.filter(name=product_name.strip(), is_archived=False)
        
        if product_type:
            query = query.filter(product_type=product_type)
        
        # Check if product exists
        if query.exists():
            existing_products.append(product_name.strip())
        else:
            missing_products.append({
                'name': product_name.strip(),
                'type': product_type or 'any',
                'message': f'Product "{product_name.strip()}" does not exist in database. Please add it to materials/inventory first.'
            })
    
    return {
        'valid': len(missing_products) == 0,
        'missing_products': missing_products,
        'existing_products': existing_products
    }


def check_inventory_availability(products_data, order_type):
    """Check if all required products are available in inventory"""
    unavailable_items = []
    
    for product_data in products_data:
        product_name = product_data.get('name', '')
        product_type = product_data.get('type', 'material')
        quantity_needed = product_data.get('quantity', 1)
        
        # First validate that product exists (prevent static data)
        validation = validate_products_exist(
            [product_name],
            product_type=product_type,
            order_type=order_type
        )
        
        if not validation['valid']:
            unavailable_items.append({
                'name': product_name,
                'reason': validation['missing_products'][0]['message'],
                'type': product_type
            })
            continue
        
        try:
            product = Product.objects.get(name=product_name, product_type=product_type, is_archived=False)
            
            if order_type in ['rent', 'rental']:
                # For rental orders, check if product is available for rental
                if not product.is_available or product.rental_status != 'available':
                    unavailable_items.append({
                        'name': product_name,
                        'reason': 'Not available for rental',
                        'current_status': product.rental_status
                    })
            else:
                # For repair/customize orders, check quantity
                if product.quantity < quantity_needed:
                    unavailable_items.append({
                        'name': product_name,
                        'reason': f'Insufficient quantity (needed: {quantity_needed}, available: {product.quantity})',
                        'available': product.quantity,
                        'needed': quantity_needed
                    })
                    
        except Product.DoesNotExist:
            # Product doesn't exist - should have been caught by validation
            unavailable_items.append({
                'name': product_name,
                'reason': f'Product "{product_name}" does not exist in database. Please add it to materials/inventory first.',
                'type': product_type
            })
    
    return {
        'available': len(unavailable_items) == 0,
        'unavailable_items': unavailable_items
    }


def calculate_thread_for_zipper(zipper_inches):
    """
    Calculate thread needed for zipper repair.
    Thread needed = 2.5 × zipper length in meters
    """
    from decimal import Decimal

    # Convert inches to meters (1 inch = 0.0254 meters)
    zipper_meters = Decimal(str(zipper_inches)) * Decimal('0.0254')
    # Thread needed = 2.5 × zipper length
    thread_meters = zipper_meters * Decimal('2.5')
    return float(thread_meters)

def find_thread_by_color(thread_color):
    """
    Find thread material matching the specified color.
    Returns Product object or None.
    """
    from .models import MaterialType

    # Try to find thread material type
    thread_type = MaterialType.objects.filter(
        Q(name__icontains='thread')
    ).first()
    
    if not thread_type:
        return None
    
    # Find thread products
    thread_products = Product.objects.filter(
        product_type='material',
        material_type=thread_type,
        is_archived=False,
        is_active=True
    )
    
    if not thread_products.exists():
        return None
    
    # If color specified, try to match
    if thread_color and thread_color.lower() not in ['', 'n/a', 'none']:
        # Normalize color name
        color_lower = thread_color.lower().strip()
        
        # Try exact match in name or description
        for thread in thread_products:
            name_lower = thread.name.lower() if thread.name else ''
            desc_lower = thread.description.lower() if thread.description else ''
            
            if color_lower in name_lower or color_lower in desc_lower:
                return thread
        
        # Try partial match
        for thread in thread_products:
            name_lower = thread.name.lower() if thread.name else ''
            desc_lower = thread.description.lower() if thread.description else ''
            
            # Check if color appears in name/description
            if any(c in name_lower or c in desc_lower for c in [color_lower[:3], color_lower]):
                return thread
    
    # Return first available thread if no color match
    return thread_products.first()

def check_thread_availability(thread_color):
    """
    Check if thread with specified color is available in stock.
    Returns dict with 'available' (bool) and 'quantity' (float) or None if not found.
    """
    thread = find_thread_by_color(thread_color)
    if not thread:
        return {'available': False, 'quantity': 0, 'found': False}
    
    # Check quantity based on unit of measurement
    if thread.unit_of_measurement == 'meters':
        quantity = float(thread.quantity) if thread.quantity else 0
    elif thread.unit_of_measurement in ['cm', 'centimeters']:
        quantity = float(thread.quantity) / 100.0 if thread.quantity else 0  # Convert cm to meters
    else:
        quantity = float(thread.quantity) if thread.quantity else 0
    
    return {
        'available': quantity > 0,
        'quantity': quantity,
        'found': True,
        'unit': thread.unit_of_measurement or 'meters'
    }

def deduct_thread_from_inventory(thread_meters, thread_color, order, repair_type):
    """
    Deduct thread from inventory matching the color.
    Returns True if successful, False otherwise.
    """
    from decimal import Decimal
    
    thread = find_thread_by_color(thread_color)
    if not thread:
        return False
    
    try:
        meters_used = Decimal(str(thread_meters))
        
        # Check if thread is stored in meters or other unit
        if thread.unit_of_measurement == 'meters':
            if thread.quantity >= meters_used:
                thread.quantity -= meters_used
                thread.current_quantity_in_stock = thread.quantity
                thread.save()
                
                InventoryTransaction.objects.create(
                    product=thread,
                    transaction_type='out',
                    quantity=-float(meters_used),
                    reference_order=order,
                    notes=f'Used {thread_meters:.2f}m {thread_color} thread for {repair_type} - Order {order.order_identifier}'
                )
                return True
        else:
            # Convert meters to the unit stored (assuming cm if not meters)
            if thread.unit_of_measurement in ['cm', 'centimeters']:
                cm_used = float(meters_used * 100)
                if thread.quantity >= cm_used:
                    thread.quantity -= int(cm_used)
                    thread.current_quantity_in_stock = thread.quantity
                    thread.save()
                    
                    InventoryTransaction.objects.create(
                        product=thread,
                        transaction_type='out',
                        quantity=-int(cm_used),
                        reference_order=order,
                        notes=f'Used {thread_meters:.2f}m ({cm_used}cm) {thread_color} thread for {repair_type} - Order {order.order_identifier}'
                    )
                    return True
            else:
                # Try direct deduction
                if thread.quantity >= float(meters_used):
                    thread.quantity -= float(meters_used)
                    thread.current_quantity_in_stock = thread.quantity
                    thread.save()
                    
                    InventoryTransaction.objects.create(
                        product=thread,
                        transaction_type='out',
                        quantity=-float(meters_used),
                        reference_order=order,
                        notes=f'Used {thread_meters:.2f}m {thread_color} thread for {repair_type} - Order {order.order_identifier}'
                    )
                    return True
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f'Error deducting thread: {str(e)}')
    
    return False

def deduct_repair_materials(order, order_data):
    """
    Automatically deduct materials used in repair orders from inventory.
    Implements comprehensive material deduction based on repair type with thread calculations.
    """
    if order.order_type != 'repair':
        return
    
    repair_type = order_data.get('repair_type')
    if not repair_type:
        return
    
    materials_deducted = []
    from decimal import Decimal
    
    try:
        # ========== ZIPPER REPLACEMENT ==========
        if repair_type == 'zipper_replacement':
            zipper_provided = order_data.get('zipper_provided', 'no')
            if zipper_provided != 'yes':
                selected_zipper_id = order_data.get('selected_zipper_id')
                selected_zipper_inches = order_data.get('selected_zipper_inches_used')
                # Get thread color from zipper modal or repair thread details
                thread_color = order_data.get('thread_color') or order_data.get('selected_thread_color') or order_data.get('threadColor', '')
                
                if selected_zipper_id and selected_zipper_inches:
                    try:
                        zipper = Product.objects.get(
                            id=selected_zipper_id,
                            product_type='material',
                            is_archived=False
                        )
                        inches_used = float(selected_zipper_inches)
                        
                        # Deduct zipper (stored in inches)
                        if zipper.quantity >= inches_used:
                            zipper.quantity -= Decimal(str(inches_used))
                            zipper.current_quantity_in_stock = zipper.quantity
                            zipper.save()
                            
                            InventoryTransaction.objects.create(
                                product=zipper,
                                transaction_type='out',
                                quantity=-float(inches_used),
                                reference_order=order,
                                notes=f'Used {inches_used} inches zipper for repair - Order {order.order_identifier}'
                            )
                            materials_deducted.append(f'Zipper: {inches_used} inches')
                            
                            # Use thread meters from form (auto-calculated) or calculate if not provided
                            thread_meters_input = order_data.get('thread_meters')
                            if thread_meters_input:
                                try:
                                    thread_meters = float(thread_meters_input)
                                except (ValueError, TypeError):
                                    thread_meters = calculate_thread_for_zipper(inches_used)
                            else:
                                thread_meters = calculate_thread_for_zipper(inches_used)
                            
                            if thread_meters > 0 and thread_color:
                                if deduct_thread_from_inventory(thread_meters, thread_color, order, repair_type):
                                    materials_deducted.append(f'Thread: {thread_meters:.2f}m ({thread_color})')
                    except Product.DoesNotExist:
                        pass
        
        # ========== LOCK/KAWIT REPAIR ==========
        elif repair_type == 'lock_repair':
            selected_lock_id = order_data.get('selected_lock_id')
            selected_lock_groups = order_data.get('selected_lock_groups_used')
            # Get thread color from lock modal or thread details section
            thread_color = order_data.get('thread_color') or order_data.get('threadColor') or order_data.get('selected_thread_color_locks', '')
            
            if selected_lock_id and selected_lock_groups:
                try:
                    lock = Product.objects.get(
                        id=selected_lock_id,
                        product_type='material',
                        is_archived=False
                    )
                    groups_used = int(selected_lock_groups)
                    
                    # Deduct locks/kawit (stored in groups)
                    if lock.quantity >= groups_used:
                        lock.quantity -= groups_used
                        lock.current_quantity_in_stock = lock.quantity
                        lock.save()
                        
                        InventoryTransaction.objects.create(
                            product=lock,
                            transaction_type='out',
                            quantity=-groups_used,
                            reference_order=order,
                            notes=f'Used {groups_used} groups locks/kawit for repair - Order {order.order_identifier}'
                        )
                        materials_deducted.append(f'Lock/Kawit: {groups_used} groups')
                        
                        # Use thread meters from form (auto-calculated) or calculate if not provided
                        thread_meters_input = order_data.get('thread_meters')
                        if thread_meters_input:
                            try:
                                thread_meters_total = float(thread_meters_input)
                            except (ValueError, TypeError):
                                # Calculate: 0.25 m × 8 = 2 meters per group
                                thread_meters_per_group = 2.0
                                thread_meters_total = groups_used * thread_meters_per_group
                        else:
                            # Calculate: 0.25 m × 8 = 2 meters per group
                            thread_meters_per_group = 2.0
                            thread_meters_total = groups_used * thread_meters_per_group
                        
                        if thread_meters_total > 0 and thread_color:
                            if deduct_thread_from_inventory(thread_meters_total, thread_color, order, repair_type):
                                materials_deducted.append(f'Thread: {thread_meters_total:.2f}m ({thread_color})')
                except Product.DoesNotExist:
                    pass
        
        # ========== BUTTON REPAIR ==========
        elif repair_type == 'buttons':
            buttons_provided = order_data.get('buttons_provided', 'no')
            if buttons_provided != 'yes':
                selected_button_id = order_data.get('selected_button_id')
                selected_button_quantity = order_data.get('selected_button_quantity_used')
                # Get thread color from button modal or thread details section
                thread_color = order_data.get('thread_color') or order_data.get('threadColor') or order_data.get('selected_thread_color_buttons', '')
                
                if selected_button_id and selected_button_quantity:
                    try:
                        button = Product.objects.get(
                            id=selected_button_id,
                            product_type='material',
                            is_archived=False
                        )
                        # Button quantity is in pieces, 1 group = 8 pieces
                        pieces_used = int(selected_button_quantity)
                        groups_used = pieces_used / 8.0
                        
                        # Deduct buttons
                        if button.quantity >= pieces_used:
                            button.quantity -= pieces_used
                            button.current_quantity_in_stock = button.quantity
                            button.save()
                            
                            InventoryTransaction.objects.create(
                                product=button,
                                transaction_type='out',
                                quantity=-pieces_used,
                                reference_order=order,
                                notes=f'Used {pieces_used} pieces ({groups_used:.1f} groups) buttons for repair - Order {order.order_identifier}'
                            )
                            materials_deducted.append(f'Buttons: {pieces_used} pieces ({groups_used:.1f} groups)')
                            
                            # Use thread meters from form (auto-calculated) or calculate if not provided
                            thread_meters_input = order_data.get('thread_meters')
                            if thread_meters_input:
                                try:
                                    thread_meters_total = float(thread_meters_input)
                                except (ValueError, TypeError):
                                    # Calculate: 2–2.4 meters per 8 pieces (1 group), use average: 2.2 meters per group
                                    thread_meters_per_group = 2.2
                                    thread_meters_total = groups_used * thread_meters_per_group
                            else:
                                # Calculate: 2–2.4 meters per 8 pieces (1 group), use average: 2.2 meters per group
                                thread_meters_per_group = 2.2
                                thread_meters_total = groups_used * thread_meters_per_group
                            
                            if thread_meters_total > 0 and thread_color:
                                if deduct_thread_from_inventory(thread_meters_total, thread_color, order, repair_type):
                                    materials_deducted.append(f'Thread: {thread_meters_total:.2f}m ({thread_color})')
                    except Product.DoesNotExist:
                        pass
        
        # ========== GARTER REPAIR (Bewang) ==========
        elif repair_type == 'bewang':
            selected_garter_id = order_data.get('selected_garter_id')
            selected_garter_inches = order_data.get('selected_garter_inches_used')
            thread_color = order_data.get('thread_color') or order_data.get('threadColor', '')
            
            if selected_garter_id and selected_garter_inches:
                try:
                    garter = Product.objects.get(
                        id=selected_garter_id,
                        product_type='material',
                        is_archived=False
                    )
                    inches_used = float(selected_garter_inches)
                    
                    # Deduct garter (stored in inches)
                    if garter.quantity >= inches_used:
                        garter.quantity -= Decimal(str(inches_used))
                        garter.current_quantity_in_stock = garter.quantity
                        garter.save()
                        
                        InventoryTransaction.objects.create(
                            product=garter,
                            transaction_type='out',
                            quantity=-float(inches_used),
                            reference_order=order,
                            notes=f'Used {inches_used} inches garter for bewang repair - Order {order.order_identifier}'
                        )
                        materials_deducted.append(f'Garter: {inches_used} inches')
                        
                        # Use thread meters from form (auto-calculated) or calculate if not provided
                        thread_meters_input = order_data.get('thread_meters')
                        if thread_meters_input:
                            try:
                                thread_meters = float(thread_meters_input)
                            except (ValueError, TypeError):
                                thread_meters = 5.0  # Default average
                        else:
                            thread_meters = 5.0  # Default average
                        
                        if thread_meters > 0 and thread_color:
                            if deduct_thread_from_inventory(thread_meters, thread_color, order, repair_type):
                                materials_deducted.append(f'Thread: {thread_meters:.2f}m ({thread_color})')
                except Product.DoesNotExist:
                    pass
        
        # ========== ELASTIC REPAIR ==========
        elif repair_type == 'elastic':
            selected_garter_id = order_data.get('selected_garter_id') or order_data.get('selected_garter_id_elastic')
            selected_garter_inches = order_data.get('selected_garter_inches_used') or order_data.get('selected_garter_inches_used_elastic')
            thread_color = order_data.get('thread_color') or order_data.get('threadColor', '')
            
            if selected_garter_id and selected_garter_inches:
                try:
                    garter = Product.objects.get(
                        id=selected_garter_id,
                        product_type='material',
                        is_archived=False
                    )
                    inches_used = float(selected_garter_inches)
                    
                    # Deduct garter/elastic (stored in inches)
                    if garter.quantity >= inches_used:
                        garter.quantity -= Decimal(str(inches_used))
                        garter.current_quantity_in_stock = garter.quantity
                        garter.save()
                        
                        InventoryTransaction.objects.create(
                            product=garter,
                            transaction_type='out',
                            quantity=-float(inches_used),
                            reference_order=order,
                            notes=f'Used {inches_used} inches garter/elastic for elastic repair - Order {order.order_identifier}'
                        )
                        materials_deducted.append(f'Garter/Elastic: {inches_used} inches')
                        
                        # Use thread meters from form (auto-calculated) or calculate if not provided
                        thread_meters_input = order_data.get('thread_meters')
                        if thread_meters_input:
                            try:
                                thread_meters = float(thread_meters_input)
                            except (ValueError, TypeError):
                                thread_meters = 5.0  # Default average
                        else:
                            thread_meters = 5.0  # Default average
                        
                        if thread_meters > 0 and thread_color:
                            if deduct_thread_from_inventory(thread_meters, thread_color, order, repair_type):
                                materials_deducted.append(f'Thread: {thread_meters:.2f}m ({thread_color})')
                except Product.DoesNotExist:
                    pass
        
        # ========== PATCH REPAIR ==========
        elif repair_type == 'patches':
            patch_provided = order_data.get('patch_provided', 'no')
            if patch_provided != 'yes':
                selected_patch_id = order_data.get('selected_patch_id')
                selected_patch_quantity = order_data.get('selected_patch_quantity_used')
                # Get thread color and length from patch section
                thread_color = order_data.get('thread_color') or order_data.get('selected_thread_color_patches', '')
                thread_color_other = order_data.get('thread_color_other') or order_data.get('selected_thread_color_other_patches', '')
                final_thread_color = thread_color_other if thread_color == 'other' else thread_color
                thread_length = float(order_data.get('thread_length') or order_data.get('selected_patch_thread_length') or 0)
                
                if selected_patch_id and selected_patch_quantity:
                    try:
                        patch = Product.objects.get(
                            id=selected_patch_id,
                            product_type='material',
                            is_archived=False
                        )
                        qty_used = int(selected_patch_quantity)
                        
                        # Deduct patches (stored in pieces)
                        if patch.quantity >= qty_used:
                            patch.quantity -= qty_used
                            patch.current_quantity_in_stock = patch.quantity
                            patch.save()
                            
                            InventoryTransaction.objects.create(
                                product=patch,
                                transaction_type='out',
                                quantity=-qty_used,
                                reference_order=order,
                                notes=f'Used {qty_used} pieces patches for repair - Order {order.order_identifier}'
                            )
                            materials_deducted.append(f'Patches: {qty_used} pieces ({patch.name})')
                            
                            # Deduct thread using the provided thread length (from auto-calculated field)
                            # thread_length comes from selectedPatchThreadLength which is auto-calculated
                            if thread_length > 0 and final_thread_color:
                                if deduct_thread_from_inventory(thread_length, final_thread_color, order, repair_type):
                                    materials_deducted.append(f'Thread: {thread_length:.2f}m ({final_thread_color})')
                    except Product.DoesNotExist:
                        pass
                    except Exception as e:
                        import logging
                        logger = logging.getLogger(__name__)
                        logger.error(f'Error deducting patches: {str(e)} - Order {order.order_identifier}')
        
        # ========== OTHER REPAIR TYPES (Thread only) ==========
        thread_required_types = [
            'ambel', 'baston', 'baston_suklot', 'baston_putol', 
            'putol', 'suklot', 'pasada', 'bewang', 'elastic',
            'general_repair', 'general_tshirt_repair'
        ]
        
        if repair_type in thread_required_types:
            # Prioritize thread_meters from form (auto-calculated)
            thread_meters_input = order_data.get('thread_meters') or order_data.get('threadMeters')
            thread_color = order_data.get('thread_color') or order_data.get('threadColor', '')
            thread_color_other = order_data.get('thread_color_other') or order_data.get('threadColorOther', '')
            final_thread_color = thread_color_other if thread_color == 'other' else thread_color
            
            # If thread meters not provided, calculate based on repair type
            if not thread_meters_input or thread_meters_input == '' or thread_meters_input == 'N/A':
                thread_meters_ranges = {
                    'ambel': (1.5, 3.0),  # Average: 2.25
                    'baston': (2.0, 4.0),  # Average: 3.0
                    'baston_suklot': (3.0, 5.0),  # Average: 4.0
                    'baston_putol': (4.0, 6.0),  # Average: 5.0
                    'putol': (0.5, 3.0),  # Average: 1.75
                    'bewang': (3.0, 6.0),  # Average: 4.5
                    'general_repair': (1.0, 3.0),  # Average: 2.0
                    'general_tshirt_repair': (1.0, 2.0),  # Average: 1.5
                    'suklot': (2.0, 4.0),  # Average: 3.0 (similar to baston)
                    'pasada': (2.0, 4.0),  # Average: 3.0 (similar to baston)
                    'elastic': (4.0, 6.0),  # Average: 5.0 (same as bewang)
                }
                
                if repair_type in thread_meters_ranges:
                    min_meters, max_meters = thread_meters_ranges[repair_type]
                    thread_meters = (min_meters + max_meters) / 2.0  # Use average
                else:
                    thread_meters = 2.0  # Default
            else:
                try:
                    thread_meters = float(thread_meters_input)
                except (ValueError, TypeError):
                    thread_meters = 2.0  # Default
            
            # Deduct thread using the calculated or provided thread meters
            if thread_meters > 0 and final_thread_color:
                if deduct_thread_from_inventory(thread_meters, final_thread_color, order, repair_type):
                    materials_deducted.append(f'Thread: {thread_meters:.2f}m ({final_thread_color})')
        
        # Log materials deducted
        if materials_deducted:
            import logging
            logger = logging.getLogger(__name__)
            logger.info(f'Materials deducted for order {order.order_identifier}: {", ".join(materials_deducted)}')
            
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f'Error deducting materials for order {order.order_identifier}: {str(e)}')
        import traceback
        traceback.print_exc()


def _find_material_by_type(type_name):
    from .models import MaterialType
    mt = MaterialType.objects.filter(Q(name__icontains=type_name)).first()
    if not mt:
        return None
    qs = Product.objects.filter(product_type='material', material_type=mt, is_archived=False, is_active=True).order_by('-quantity')
    return qs.first() if qs.exists() else None

def _find_fabric_by_type_and_color(fabric_type, fabric_color):
    """
    Find fabric material by type and color.
    Returns Product object or None.
    """
    from .models import MaterialType

    # Find fabric material type
    fabric_material_type = MaterialType.objects.filter(
        Q(name__icontains='fabric')
    ).first()
    
    if not fabric_material_type:
        return None
    
    # Find fabric products
    fabric_products = Product.objects.filter(
        product_type='material',
        material_type=fabric_material_type,
        is_archived=False,
        is_active=True
    )
    
    if not fabric_products.exists():
        return None
    
    # Normalize search terms
    fabric_type_lower = (fabric_type or '').lower().strip()
    fabric_color_lower = (fabric_color or '').lower().strip()
    
    # Try to match by description (contains "Type of Fabric: {type}" and "Color: {color}")
    for fabric in fabric_products:
        desc_lower = (fabric.description or '').lower()
        name_lower = (fabric.name or '').lower()
        
        # Check if fabric type matches
        type_match = False
        if fabric_type_lower:
            # Check in description: "Type of Fabric: {type}"
            if f'type of fabric: {fabric_type_lower}' in desc_lower or \
               f'type of fabric:{fabric_type_lower}' in desc_lower or \
               fabric_type_lower in desc_lower or \
               fabric_type_lower in name_lower:
                type_match = True
        else:
            type_match = True  # If no type specified, match any
        
        # Check if color matches
        color_match = False
        if fabric_color_lower:
            # Check in description: "Color: {color}"
            if f'color: {fabric_color_lower}' in desc_lower or \
               f'color:{fabric_color_lower}' in desc_lower or \
               fabric_color_lower in desc_lower or \
               fabric_color_lower in name_lower:
                color_match = True
        else:
            color_match = True  # If no color specified, match any
        
        if type_match and color_match:
            return fabric
    
    # If no exact match, return first available fabric
    return fabric_products.first()

def _find_thread_by_brand_and_color(thread_brand, thread_color):
    """
    Find thread material by brand (Type Of Thread) and color.
    Returns Product object or None.
    """
    from .models import MaterialType

    # Find thread material type
    thread_material_type = MaterialType.objects.filter(
        Q(name__icontains='thread')
    ).first()
    
    if not thread_material_type:
        return None
    
    # Find thread products
    thread_products = Product.objects.filter(
        product_type='material',
        material_type=thread_material_type,
        is_archived=False,
        is_active=True
    )
    
    if not thread_products.exists():
        return None
    
    # Normalize search terms
    thread_brand_lower = (thread_brand or '').lower().strip()
    thread_color_lower = (thread_color or '').lower().strip()
    
    # Try to match by description or name
    for thread in thread_products:
        desc_lower = (thread.description or '').lower()
        name_lower = (thread.name or '').lower()
        
        # Check if brand (Type Of Thread) matches
        brand_match = False
        if thread_brand_lower:
            # Check in description or name for brand
            if thread_brand_lower in desc_lower or thread_brand_lower in name_lower:
                brand_match = True
        else:
            brand_match = True  # If no brand specified, match any
        
        # Check if color matches
        color_match = False
        if thread_color_lower:
            # Check in description: "Color: {color}" or in name
            if f'color: {thread_color_lower}' in desc_lower or \
               f'color:{thread_color_lower}' in desc_lower or \
               thread_color_lower in desc_lower or \
               thread_color_lower in name_lower:
                color_match = True
        else:
            color_match = True  # If no color specified, match any
        
        if brand_match and color_match:
            return thread
    
    # If no exact match, return first available thread
    return thread_products.first()

def _find_button_by_type_and_color(button_type, button_color):
    """
    Find button material by type and color.
    Returns Product object or None.
    """
    from .models import MaterialType

    # Find button material type
    button_material_type = MaterialType.objects.filter(
        Q(name__icontains='button')
    ).first()
    
    if not button_material_type:
        return None
    
    # Find button products
    button_products = Product.objects.filter(
        product_type='material',
        material_type=button_material_type,
        is_archived=False,
        is_active=True
    )
    
    if not button_products.exists():
        return None
    
    # Normalize search terms
    button_type_lower = (button_type or '').lower().strip()
    button_color_lower = (button_color or '').lower().strip()
    
    # Try to match by description or name
    for button in button_products:
        desc_lower = (button.description or '').lower()
        name_lower = (button.name or '').lower()
        
        # Check if button type matches
        type_match = False
        if button_type_lower:
            # Check in description or name for type
            if button_type_lower in desc_lower or button_type_lower in name_lower:
                type_match = True
        else:
            type_match = True  # If no type specified, match any
        
        # Check if color matches
        color_match = False
        if button_color_lower:
            # Check in description: "Color: {color}" or in name
            if f'color: {button_color_lower}' in desc_lower or \
               f'color:{button_color_lower}' in desc_lower or \
               button_color_lower in desc_lower or \
               button_color_lower in name_lower:
                color_match = True
        else:
            color_match = True  # If no color specified, match any
        
        if type_match and color_match:
            return button
    
    # If no exact match, return first available button
    return button_products.first()

def _find_garter():
    """
    Find garter material (no attributes needed).
    Returns Product object or None.
    """
    from .models import MaterialType

    # Find garter material type
    garter_material_type = MaterialType.objects.filter(
        Q(name__icontains='garter') | Q(name__icontains='elastic')
    ).first()
    
    if not garter_material_type:
        return None
    
    # Find garter products
    garter_products = Product.objects.filter(
        product_type='material',
        material_type=garter_material_type,
        is_archived=False,
        is_active=True
    ).order_by('-quantity')
    
    return garter_products.first() if garter_products.exists() else None

def _find_locks():
    """
    Find locks/kawit material (no attributes needed).
    Returns Product object or None.
    """
    from .models import MaterialType

    # Find locks material type
    locks_material_type = MaterialType.objects.filter(
        Q(name__icontains='lock') | Q(name__icontains='kawit')
    ).first()
    
    if not locks_material_type:
        return None
    
    # Find locks products
    locks_products = Product.objects.filter(
        product_type='material',
        material_type=locks_material_type,
        is_archived=False,
        is_active=True
    ).order_by('-quantity')
    
    return locks_products.first() if locks_products.exists() else None

def _deduct_material_amount(product, amount, unit, order, note):
    from decimal import Decimal
    if not product or amount <= 0:
        return False
    amt = Decimal(str(amount))
    
    # Handle unit conversions
    if unit == 'meters' and product.unit_of_measurement and 'yard' in product.unit_of_measurement.lower():
        amt = amt * Decimal('1.09361')
    elif unit == 'yards' and product.unit_of_measurement and 'meter' in product.unit_of_measurement.lower():
        amt = amt * Decimal('0.9144')
    elif unit == 'inches' and product.unit_of_measurement and 'cm' in product.unit_of_measurement.lower():
        # Convert inches to cm (1 inch = 2.54 cm)
        amt = amt * Decimal('2.54')
    elif unit == 'cm' and product.unit_of_measurement and 'inch' in product.unit_of_measurement.lower():
        # Convert cm to inches (1 cm = 0.393701 inches)
        amt = amt * Decimal('0.393701')
    elif unit == 'groups':
        # Groups are handled directly - no conversion needed
        # The product's unit_of_measurement should already be 'group' or 'groups'
        pass
    
    # Check if product has enough quantity
    if product.quantity >= amt:
        product.quantity -= amt
        product.current_quantity_in_stock = product.quantity
        product.save()
        InventoryTransaction.objects.create(product=product, transaction_type='out', quantity=-float(amt), reference_order=order, notes=note)
        return True
    else:
        import logging
        logger = logging.getLogger(__name__)
        logger.warning(f'Insufficient stock for {product.name}: Need {amt}, Have {product.quantity} - Order {order.order_identifier}')
    return False

def _parse_measurements_from_product(product):
    m = {}
    desc = product.description or ''
    if 'Measurements:' in desc:
        s = desc.find('Measurements:')
        js = desc[s + len('Measurements:'):].strip()
        try:
            m = json.loads(js)
        except Exception:
            import re
            mm = re.search(r'\{[\s\S]*\}', js, re.MULTILINE | re.DOTALL)
            if mm:
                try:
                    m = json.loads(mm.group())
                except Exception:
                    m = {}
    if not isinstance(m, dict):
        m = {}
    return {k.strip().lower().replace(' ', '_'): (str(v).strip() if v is not None else '') for k, v in m.items()}

def deduct_customize_materials(order, order_data=None):
    """
    Deduct materials from inventory for customization orders.
    Matches materials by their attributes (type, color, brand) and deducts the correct amounts.
    """
    if order.order_type != 'customize':
        return
    
    if not order_data:
        order_data = {}
    
    # Get customization type from order_data or product
    customize_type = order_data.get('customize_type') or order_data.get('type_of_customize', '').lower()
    
    # If not in order_data, try to extract from order items
    if not customize_type:
        items = list(order.items.select_related('product'))
        for it in items:
            p = it.product
            if p.product_type != 'service':
                continue
            name = (p.name or '').lower()
            cat = p.category.name.lower() if p.category and p.category.name else ''
            
            if any(x in cat for x in ['polo', 'polo_shirt']) or any(x in name for x in ['polo', 'polo shirt']):
                customize_type = 'polo' if 'polo_shirt' not in cat and 'polo shirt' not in name else 'polo_shirt'
            elif 'blouse' in cat or 'blouse' in name:
                customize_type = 'blouse'
            elif any(x in cat for x in ['pants', 'pe_pants']) or 'pants' in name:
                customize_type = 'pants'
            elif 'short' in cat or 'short' in name:
                customize_type = 'shorts'
            elif 'skirt' in cat or 'palda' in cat or 'skirt' in name or 'palda' in name:
                customize_type = 'skirt_palda'
            break
    
    if not customize_type:
        customize_type = 'polo'  # Default fallback
    
    # Get quantity
    qty = max(1, int(order_data.get('customize_quantity', 1) or 1))
    
    # Extract material information from order_data
    fabric_type = order_data.get('customize_fabric', '').strip()
    fabric_color = order_data.get('customize_fabric_color', '').strip()  # May not be in form, try to extract
    fabric_yards = order_data.get('customize_fabric_yards', 0)
    
    thread_brand = order_data.get('customize_thread_brand', '').strip() or order_data.get('thread_brand', '').strip()  # Type Of Thread
    thread_color = order_data.get('customize_thread_color', '').strip() or order_data.get('thread_color', '').strip()
    thread_meters = order_data.get('customize_thread_meters', 0) or order_data.get('thread_meters', 0)
    
    button_type = order_data.get('uniform_button_type', '').strip()
    button_color = order_data.get('uniform_button_color', '').strip()  # May not be in form
    button_quantity = order_data.get('uniform_buttons_needed', 0) or 0
    
    garter_inches = order_data.get('selected_garter_inches_used', 0) or order_data.get('garter_cm', 0) or 0
    # Convert cm to inches if needed (1 cm = 0.393701 inches)
    try:
        if garter_inches and float(garter_inches) > 100:  # Likely in cm if > 100
            garter_inches = float(garter_inches) * 0.393701
        else:
            garter_inches = float(garter_inches) if garter_inches else 0
    except (ValueError, TypeError):
        garter_inches = 0
    
    lock_groups = order_data.get('selected_lock_groups_used', 0) or order_data.get('lock_groups', 0) or 0
    try:
        lock_groups = int(lock_groups) if lock_groups else 0
    except (ValueError, TypeError):
        lock_groups = 0
    
    materials_deducted = []
    
    # ========== DEDUCT FABRIC ==========
    # For all customization types: Polo, Polo Shirt, Blouse, Pants, Shorts, Skirt/Palda
    if fabric_type and fabric_yards:
        try:
            fabric_yards_float = float(fabric_yards) * qty
            fabric = _find_fabric_by_type_and_color(fabric_type, fabric_color)
            if fabric:
                if _deduct_material_amount(fabric, fabric_yards_float, 'yards', order, 
                    f'Fabric ({fabric_type} {fabric_color or ""}) used for customize {customize_type} - {fabric_yards_float} yards - Order {order.order_identifier}'):
                    materials_deducted.append(f'Fabric: {fabric_yards_float} yards ({fabric_type} {fabric_color or ""})')
            else:
                import logging
                logger = logging.getLogger(__name__)
                logger.warning(f'Fabric not found: Type={fabric_type}, Color={fabric_color} - Order {order.order_identifier}')
        except (ValueError, TypeError) as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f'Error deducting fabric: {e} - Order {order.order_identifier}')
    
    # ========== DEDUCT THREAD ==========
    # For all customization types: Polo, Polo Shirt, Blouse, Pants, Shorts, Skirt/Palda
    if thread_color and thread_meters:
        try:
            thread_meters_float = float(thread_meters) * qty
            thread = _find_thread_by_brand_and_color(thread_brand, thread_color)
            if thread:
                # Use the existing deduct_thread_from_inventory function which handles meters/cm conversion
                if deduct_thread_from_inventory(thread_meters_float, thread_color, order, 'customize'):
                    materials_deducted.append(f'Thread: {thread_meters_float}m ({thread_brand or ""} {thread_color})')
            else:
                import logging
                logger = logging.getLogger(__name__)
                logger.warning(f'Thread not found: Brand={thread_brand}, Color={thread_color} - Order {order.order_identifier}')
        except (ValueError, TypeError) as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f'Error deducting thread: {e} - Order {order.order_identifier}')
    
    # ========== DEDUCT BUTTONS ==========
    # Only for Polo, Polo Shirt, Blouse
    if customize_type in ['polo', 'polo_shirt', 'blouse'] and button_type and button_quantity:
        try:
            button_qty_int = int(button_quantity) * qty
            # Button quantity is in pieces, 1 group = 8 pieces
            groups_used = button_qty_int / 8.0
            
            button = _find_button_by_type_and_color(button_type, button_color)
            if button:
                # Check button unit_of_measurement to determine if we deduct by groups or pieces
                if button.unit_of_measurement and 'group' in button.unit_of_measurement.lower():
                    # Deduct by groups
                    if _deduct_material_amount(button, groups_used, 'groups', order,
                        f'Buttons ({button_type} {button_color or ""}) used for customize {customize_type} - {groups_used} groups ({button_qty_int} pieces) - Order {order.order_identifier}'):
                        materials_deducted.append(f'Buttons: {groups_used} groups ({button_qty_int} pieces) ({button_type} {button_color or ""})')
                else:
                    # Deduct by pieces
                    if _deduct_material_amount(button, button_qty_int, 'pieces', order,
                        f'Buttons ({button_type} {button_color or ""}) used for customize {customize_type} - {button_qty_int} pieces - Order {order.order_identifier}'):
                        materials_deducted.append(f'Buttons: {button_qty_int} pieces ({button_type} {button_color or ""})')
            else:
                import logging
                logger = logging.getLogger(__name__)
                logger.warning(f'Button not found: Type={button_type}, Color={button_color} - Order {order.order_identifier}')
        except (ValueError, TypeError) as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f'Error deducting buttons: {e} - Order {order.order_identifier}')
    
    # ========== DEDUCT GARTER ==========
    # Only for Pants, Shorts, Skirt/Palda
    if customize_type in ['pants', 'shorts', 'skirt_palda'] and garter_inches > 0:
        try:
            garter = _find_garter()
            if garter:
                # Garter is measured in inches or cm
                # Check unit_of_measurement
                if garter.unit_of_measurement and 'cm' in garter.unit_of_measurement.lower():
                    # Convert inches to cm (1 inch = 2.54 cm)
                    garter_amount = garter_inches * 2.54
                    unit = 'cm'
                else:
                    garter_amount = garter_inches
                    unit = 'inches'
                
                if _deduct_material_amount(garter, garter_amount, unit, order,
                    f'Garter used for customize {customize_type} - {garter_amount} {unit} - Order {order.order_identifier}'):
                    materials_deducted.append(f'Garter: {garter_amount} {unit}')
            else:
                import logging
                logger = logging.getLogger(__name__)
                logger.warning(f'Garter not found - Order {order.order_identifier}')
        except (ValueError, TypeError) as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f'Error deducting garter: {e} - Order {order.order_identifier}')
    
    # ========== DEDUCT LOCKS/KAWIT ==========
    # Only for Skirt/Palda
    if customize_type == 'skirt_palda' and lock_groups > 0:
        try:
            locks = _find_locks()
            if locks:
                # Locks are measured in groups (1 group = 5 locks)
                if _deduct_material_amount(locks, lock_groups, 'groups', order,
                    f'Locks/Kawit used for customize {customize_type} - {lock_groups} groups ({lock_groups * 5} pieces) - Order {order.order_identifier}'):
                    materials_deducted.append(f'Locks/Kawit: {lock_groups} groups ({lock_groups * 5} pieces)')
            else:
                import logging
                logger = logging.getLogger(__name__)
                logger.warning(f'Locks/Kawit not found - Order {order.order_identifier}')
        except (ValueError, TypeError) as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f'Error deducting locks: {e} - Order {order.order_identifier}')
    
    # Log materials deducted
    if materials_deducted:
        import logging
        logger = logging.getLogger(__name__)
        logger.info(f'Materials deducted for customize order {order.order_identifier} ({customize_type}): {", ".join(materials_deducted)}')

def ensure_materials_recorded_for_order(order, order_data=None):
    """
    Ensure all materials used in an order are recorded as InventoryTransaction records.
    This function is called after order creation to verify completeness.
    Note: This is a verification function - actual material deduction happens in deduct_repair_materials/deduct_customize_materials
    """
    try:
        # Check if order already has material transactions
        existing_transactions = InventoryTransaction.objects.filter(
            reference_order=order,
            transaction_type='out',
            product__product_type='material'
        ).count()
        
        # Log if no transactions found for repair/customize orders (for debugging)
        if existing_transactions == 0 and order.order_type in ['repair', 'customize']:
            import logging
            logger = logging.getLogger(__name__)
            logger.info(f'Order {order.order_identifier} ({order.order_type}) has no material transactions recorded. This may be normal if materials were provided by customer or not available.')
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f'Error checking materials recorded for order {order.order_identifier}: {e}')

def update_inventory_for_order(product, quantity, order, order_type):
    """Update inventory based on order type and create transaction record"""
    from datetime import timedelta

    # Ensure quantity is an integer
    quantity = int(quantity) if quantity else 0
    
    if order_type in ['rent', 'rental']:
        # For rental orders, mark as rented
        product.rental_status = 'rented'
        product.current_rental_order = order
        product.rental_start_date = timezone.now()
        product.rental_due_date = timezone.now() + timedelta(days=3)
        
        # Create inventory transaction
        InventoryTransaction.objects.create(
            product=product,
            transaction_type='rental_out',
            quantity=1,  # Rental is always 1 item
            reference_order=order,
            notes=f'Rented out for order {order.order_identifier}'
        )
        
    else:
        # For repair/customize orders, deduct from inventory
        if product.quantity >= quantity:
            product.quantity -= quantity
            
            # Create inventory transaction
            InventoryTransaction.objects.create(
                product=product,
                transaction_type='out',
                quantity=-quantity,  # Negative for stock out
                reference_order=order,
                notes=f'Used for {order_type} order {order.order_identifier}'
            )
    
    product.save()


def update_dashboard_statistics():
    """Update dashboard statistics in real-time"""
    # This function can be called to refresh dashboard data
    # Statistics are calculated in real-time in the dashboard view
    pass


def calculate_actual_sales():
    """Calculate actual sales from completed orders"""
    from decimal import Decimal

    # Get all completed orders
    completed_orders = Order.objects.filter(status='completed')
    
    total_sales = Decimal('0')
    sales_count = 0
    
    for order in completed_orders:
        # Check if sales record exists
        sales, created = Sales.objects.get_or_create(
            order=order,
            defaults={
                'amount': order.total_amount,
                'payment_method': 'cash'  # Default payment method
            }
        )
        
        if created:
            total_sales += order.total_amount
            sales_count += 1
    
    return {
        'total_sales': total_sales,
        'sales_count': sales_count,
        'new_sales_created': sales_count
    }


def get_inventory_status():
    """Get real-time inventory status for all products - EXCLUDE service/customize products"""
    # Exclude service products from inventory status calculations
    products = Product.objects.filter(
        is_active=True, 
        is_archived=False
    ).exclude(product_type='service')
    
    inventory_status = {
        'total_products': products.count(),
        'low_stock': 0,
        'out_of_stock': 0,
        'rental_available': 0,
        'rental_rented': 0,
        'rental_overdue': 0,
        'products': []
    }
    
    for product in products:
        product_status = {
            'id': product.id,
            'name': product.name,
            'type': product.product_type,
            'quantity': product.quantity,
            'price': float(product.price),
            'is_available': product.is_available,
            'is_low_stock': product.is_low_stock,
            'rental_status': product.rental_status if product.product_type == 'rental' else None,
            'is_overdue': product.is_overdue if product.product_type == 'rental' else False
        }
        
        inventory_status['products'].append(product_status)
        
        # Update counters
        if product.is_low_stock:
            inventory_status['low_stock'] += 1
        if product.quantity == 0:
            inventory_status['out_of_stock'] += 1
        if product.product_type == 'rental':
            if product.rental_status == 'available':
                inventory_status['rental_available'] += 1
            elif product.rental_status == 'rented':
                inventory_status['rental_rented'] += 1
                if product.is_overdue:
                    inventory_status['rental_overdue'] += 1
    
    return inventory_status


def get_order_tracking_data(order_identifier):
    """Get comprehensive order tracking data"""
    try:
        order = Order.objects.get(order_identifier=order_identifier)
        
        tracking_data = {
            'order': {
                'id': order.id,
                'identifier': order.order_identifier,
                'type': order.get_order_type_display(),
                'status': order.get_status_display(),
                'total_amount': float(order.total_amount),
                'paid_amount': float(order.paid_amount),
                'balance': float(order.balance),
                'created_at': order.created_at,
                'due_date': order.due_date,
                'notes': order.notes
            },
            'customer': {
                'name': order.customer.name,
                'phone': order.customer.phone,
                'email': order.customer.email,
                'address': order.customer.address
            },
            'items': [],
            'inventory_status': [],
            'sales_info': None
        }
        
        # Get order items
        for item in order.items.all():
            tracking_data['items'].append({
                'product_name': item.product.name,
                'quantity': item.quantity,
                'unit_price': float(item.unit_price),
                'total_price': float(item.total_price)
            })
            
            # Get inventory status for each product
            product_status = {
                'product_name': item.product.name,
                'current_quantity': item.product.quantity,
                'is_available': item.product.is_available,
                'rental_status': item.product.rental_status if item.product.product_type == 'rental' else None,
                'is_overdue': item.product.is_overdue if item.product.product_type == 'rental' else False
            }
            tracking_data['inventory_status'].append(product_status)
        
        # Get sales information if exists
        try:
            sales = Sales.objects.get(order=order)
            tracking_data['sales_info'] = {
                'sales_identifier': sales.sales_identifier,
                'amount': float(sales.amount),
                'payment_method': sales.payment_method,
                'created_at': sales.created_at
            }
        except Sales.DoesNotExist:
            pass
        
        return tracking_data
        
    except Order.DoesNotExist:
        return None


@login_required
def api_inventory_status(request):
    """API endpoint to get real-time inventory status"""
    if request.method == 'GET':
        inventory_status = get_inventory_status()
        return JsonResponse(inventory_status)
    
    return JsonResponse({'error': 'Invalid request method'})


@login_required
def api_sales_calculation(request):
    """API endpoint to calculate and update sales"""
    if request.method == 'POST':
        sales_data = calculate_actual_sales()
        return JsonResponse(sales_data)
    
    return JsonResponse({'error': 'Invalid request method'})


def safe_json_response(data):
    """Global helper function to safely serialize JSON with Decimal support"""
    from decimal import Decimal
    
    def convert_decimals(obj):
        """Recursively convert Decimal values to float in nested structures"""
        if isinstance(obj, Decimal):
            return float(obj)
        elif isinstance(obj, dict):
            return {key: convert_decimals(value) for key, value in obj.items()}
        elif isinstance(obj, (list, tuple)):
            return [convert_decimals(item) for item in obj]
        elif hasattr(obj, '__dict__'):
            return convert_decimals(obj.__dict__)
        return obj
    
    def decimal_default(obj):
        if isinstance(obj, Decimal):
            return float(obj)
        raise TypeError(f"Object of type {type(obj)} is not JSON serializable")
    
    # Convert all Decimals in the data structure first
    converted_data = convert_decimals(data)
    return JsonResponse(converted_data, json_dumps_params={'default': decimal_default})

@login_required
def api_check_repair_materials_availability(request):
    """API endpoint to check repair materials availability"""
    if request.method == 'POST':
        try:
            import json
            data = json.loads(request.body)
            
            order_type = data.get('orderType') or data.get('service_type')
            repair_type = data.get('repair_type')
            
            unavailable_materials = []
            
            # For repair orders, check specific materials based on repair type
            if order_type == 'repair' and repair_type:
                # Check zipper availability for zipper_replacement
                if repair_type == 'zipper_replacement':
                    zipper_provided = data.get('zipper_provided', 'no')
                    # Only check if zipper is NOT provided by customer
                    if zipper_provided != 'yes':
                        selected_zipper_id = data.get('selected_zipper_id')
                        selected_zipper_inches = data.get('selected_zipper_inches_used')
                        
                        if selected_zipper_id and selected_zipper_inches:
                            try:
                                selected_zipper = Product.objects.get(
                                    id=selected_zipper_id,
                                    product_type='material',
                                    is_archived=False,
                                    is_active=True
                                )
                                available_inches = float(selected_zipper.quantity) if selected_zipper.quantity else 0
                                inches_needed = float(selected_zipper_inches)
                                
                                if inches_needed > available_inches:
                                    unavailable_materials.append({
                                        'material': f'Zipper - {selected_zipper.name}',
                                        'needed': inches_needed,
                                        'available': available_inches,
                                        'message': f'Insufficient zipper length. Need {inches_needed} inches, but only {available_inches} inches available.'
                                    })
                            except Product.DoesNotExist:
                                unavailable_materials.append({
                                    'material': 'Selected Zipper',
                                    'needed': selected_zipper_inches,
                                    'available': 0,
                                    'message': 'Selected zipper not found in inventory.'
                                })
                        elif not selected_zipper_id:
                            # If no zipper is selected, try to find any available zipper
                            zippers = Product.objects.filter(
                                product_type='material',
                                is_archived=False,
                                is_active=True
                            )
                            # Check if product name contains "zipper" (case-insensitive)
                            zippers = [z for z in zippers if 'zipper' in z.name.lower() or (z.material_type and 'zipper' in z.material_type.name.lower())]
                            if not zippers or all(z.quantity == 0 for z in zippers):
                                unavailable_materials.append({
                                    'material': 'Zipper',
                                    'needed': 1,
                                    'available': 0,
                                    'message': 'No zippers available in inventory. Please select a zipper or mark as provided.'
                                })
                
                # Check button availability for buttons repair
                elif repair_type == 'buttons':
                    buttons_provided = data.get('buttons_provided', 'no')
                    # Only check if buttons are NOT provided by customer
                    if buttons_provided != 'yes':
                        selected_button_id = data.get('selected_button_id')
                        selected_button_quantity = data.get('selected_button_quantity_used')
                        
                        if selected_button_id and selected_button_quantity:
                            try:
                                selected_button = Product.objects.get(
                                    id=selected_button_id,
                                    product_type='material',
                                    is_archived=False,
                                    is_active=True
                                )
                                available_qty = int(selected_button.quantity) if selected_button.quantity else 0
                                qty_needed = int(selected_button_quantity)
                                
                                if qty_needed > available_qty:
                                    unavailable_materials.append({
                                        'material': f'Button - {selected_button.name}',
                                        'needed': qty_needed,
                                        'available': available_qty,
                                        'message': f'Insufficient buttons. Need {qty_needed} pieces, but only {available_qty} pieces available.'
                                    })
                            except Product.DoesNotExist:
                                unavailable_materials.append({
                                    'material': 'Selected Button',
                                    'needed': selected_button_quantity,
                                    'available': 0,
                                    'message': 'Selected button not found in inventory.'
                                })
                        elif not selected_button_id:
                            # If no button is selected, try to find any available button
                            buttons = Product.objects.filter(
                                product_type='material',
                                is_archived=False,
                                is_active=True
                            )
                            # Check if product name or material type contains "button" (case-insensitive)
                            buttons = [b for b in buttons if 'button' in b.name.lower() or (b.material_type and 'button' in b.material_type.name.lower())]
                            if not buttons or all(b.quantity == 0 for b in buttons):
                                unavailable_materials.append({
                                    'material': 'Button',
                                    'needed': 4,  # Default: 4 pieces for repair
                                    'available': 0,
                                    'message': 'No buttons available in inventory. Please select a button or mark as provided.'
                                })
                
                # Check lock availability for lock_repair
                elif repair_type == 'lock_repair':
                    selected_lock_id = data.get('selected_lock_id')
                    selected_lock_groups = data.get('selected_lock_groups_used')
                    
                    if selected_lock_id and selected_lock_groups:
                        try:
                            selected_lock = Product.objects.get(
                                id=selected_lock_id,
                                product_type='material',
                                is_archived=False,
                                is_active=True
                            )
                            available_groups = int(selected_lock.quantity) if selected_lock.quantity else 0
                            groups_needed = int(selected_lock_groups)
                            
                            if groups_needed > available_groups:
                                unavailable_materials.append({
                                    'material': f'Lock/Kawit - {selected_lock.name}',
                                    'needed': groups_needed,
                                    'available': available_groups,
                                    'message': f'Insufficient lock/kawit groups. Need {groups_needed} groups, but only {available_groups} groups available.'
                                })
                        except Product.DoesNotExist:
                            unavailable_materials.append({
                                'material': 'Selected Lock/Kawit',
                                'needed': selected_lock_groups,
                                'available': 0,
                                'message': 'Selected lock/kawit not found in inventory.'
                            })
                    elif not selected_lock_id:
                        # If no lock is selected, try to find any available lock
                        locks = Product.objects.filter(
                            product_type='material',
                            is_archived=False,
                            is_active=True
                        )
                        # Check if product name or material type contains "lock" or "kawit" (case-insensitive)
                        locks = [l for l in locks if 'lock' in l.name.lower() or 'kawit' in l.name.lower() or (l.material_type and ('lock' in l.material_type.name.lower() or 'kawit' in l.material_type.name.lower()))]
                        if not locks or all(l.quantity == 0 for l in locks):
                            unavailable_materials.append({
                                'material': 'Lock/Kawit',
                                'needed': 1,  # Default: 1 group
                                'available': 0,
                                'message': 'No locks/kawit available in inventory. Please select a lock/kawit.'
                            })
                
                # Check garter availability for bewang
                elif repair_type == 'bewang':
                    selected_garter_id = data.get('selected_garter_id')
                    selected_garter_inches = data.get('selected_garter_inches_used')
                    
                    if selected_garter_id and selected_garter_inches:
                        try:
                            selected_garter = Product.objects.get(
                                id=selected_garter_id,
                                product_type='material',
                                is_archived=False,
                                is_active=True
                            )
                            available_inches = float(selected_garter.quantity) if selected_garter.quantity else 0
                            inches_needed = float(selected_garter_inches)
                            
                            if inches_needed > available_inches:
                                unavailable_materials.append({
                                    'material': f'Garter - {selected_garter.name}',
                                    'needed': inches_needed,
                                    'available': available_inches,
                                    'message': f'Insufficient garter length. Need {inches_needed} inches, but only {available_inches} inches available.'
                                })
                        except Product.DoesNotExist:
                            unavailable_materials.append({
                                'material': 'Selected Garter',
                                'needed': selected_garter_inches,
                                'available': 0,
                                'message': 'Selected garter not found in inventory.'
                            })
                    elif not selected_garter_id:
                        # If no garter is selected, try to find any available garter
                        garters = Product.objects.filter(
                            product_type='material',
                            is_archived=False,
                            is_active=True
                        )
                        # Check if product name or material type contains "garter" (case-insensitive)
                        garters = [g for g in garters if 'garter' in g.name.lower() or (g.material_type and 'garter' in g.material_type.name.lower())]
                        if not garters or all(g.quantity == 0 for g in garters):
                            unavailable_materials.append({
                                'material': 'Garter',
                                'needed': 0.85,  # Default: approximately 0.85 meters (converted to inches in actual use)
                                'available': 0,
                                'message': 'No garters available in inventory. Please select a garter.'
                            })
            
            # Return availability status
            if unavailable_materials:
                return safe_json_response({
                    'available': False,
                    'unavailable_materials': unavailable_materials,
                    'message': 'Some materials are unavailable. Please check the materials page and select available materials.'
                })
            else:
                return safe_json_response({
                    'available': True,
                    'message': 'All materials are available.'
                })
                
        except Exception as e:
            import traceback
            traceback.print_exc()
            return safe_json_response({
                'available': False,
                'error': str(e),
                'message': f'Error checking material availability: {str(e)}'
            })
    
    return safe_json_response({'available': False, 'error': 'Invalid request method'})


@login_required
def api_revenue_details(request):
    """API endpoint to get revenue details by period"""
    try:
        period = request.GET.get('period', 'total')  # daily, weekly, monthly, yearly, total
        
        # Base queryset - only include completed orders, including archived ones for reports
        orders = Order.objects.filter(status='completed').select_related('customer')
        
        # Calculate date range based on period
        from datetime import datetime, timedelta
        now = timezone.now()
        
        if period == 'daily':
            start_date = now.replace(hour=0, minute=0, second=0, microsecond=0)
            orders = orders.filter(created_at__gte=start_date)
        elif period == 'weekly':
            today = now.date()
            days_since_monday = today.weekday()  # Monday is 0, Sunday is 6
            start_of_week = today - timedelta(days=days_since_monday)
            start_of_week_datetime = datetime.combine(start_of_week, datetime.min.time())
            start_of_week_datetime = timezone.make_aware(start_of_week_datetime)
            orders = orders.filter(created_at__gte=start_of_week_datetime)
        elif period == 'monthly':
            current_month = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            orders = orders.filter(created_at__gte=current_month)
        elif period == 'yearly':
            current_year = now.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
            orders = orders.filter(created_at__gte=current_year)
        # else: total - no filter
        
        # Calculate total revenue and orders
        from django.db.models import Count, Sum
        total_revenue = orders.aggregate(total=Sum('total_amount'))['total'] or 0
        total_orders = orders.count()
        
        # Revenue by order type
        revenue_by_type = []
        for order_type in ['rental', 'repair', 'customize']:
            type_orders = orders.filter(order_type=order_type)
            type_revenue = type_orders.aggregate(total=Sum('total_amount'))['total'] or 0
            type_count = type_orders.count()
            
            if type_count > 0:
                revenue_by_type.append({
                    'order_type': order_type,
                    'revenue': float(type_revenue),
                    'count': type_count
                })
        
        # Monthly breakdown (for yearly and total periods)
        monthly_breakdown = []
        if period in ['yearly', 'total']:
            # Get all months in the current year or all months if total
            start_date = now.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0) if period == 'yearly' else orders.order_by('created_at').first().created_at if orders.exists() else now
            end_date = now
            
            current = start_date
            while current <= end_date:
                month_start = current.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
                if month_start.month == 12:
                    month_end = month_start.replace(year=month_start.year + 1, month=1)
                else:
                    month_end = month_start.replace(month=month_start.month + 1)
                
                month_orders = orders.filter(created_at__gte=month_start, created_at__lt=month_end)
                month_revenue = month_orders.aggregate(total=Sum('total_amount'))['total'] or 0
                month_count = month_orders.count()
                
                if month_count > 0:
                    monthly_breakdown.append({
                        'month': month_start.strftime('%B %Y'),
                        'revenue': float(month_revenue),
                        'count': month_count
                    })
                
                current = month_end
        
        return JsonResponse({
            'success': True,
            'period': period,
            'total_revenue': float(total_revenue),
            'total_orders': total_orders,
            'revenue_by_type': revenue_by_type,
            'monthly_breakdown': monthly_breakdown
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'error': str(e)
        })


@login_required
def api_order_tracking(request):
    """API endpoint for order tracking"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            order_identifier = data.get('order_identifier')
            
            if order_identifier:
                tracking_data = get_order_tracking_data(order_identifier)
                if tracking_data:
                    return JsonResponse({
                        'success': True,
                        'data': tracking_data
                    })
                else:
                    return JsonResponse({
                        'success': False,
                        'error': 'Order not found'
                    })
            else:
                return JsonResponse({
                    'success': False,
                    'error': 'Order identifier required'
                })
                
        except json.JSONDecodeError:
            return JsonResponse({
                'success': False,
                'error': 'Invalid JSON data'
            })
    
    return JsonResponse({'error': 'Invalid request method'})


@login_required
def order_receipt_new(request):
    return render(request, 'business/order_receipt_new.html')


@login_required
def create_order_from_session(request):
    def safe_json(data):
        """Helper function to safely serialize JSON with Decimal support"""
        import json as json_module
        from decimal import Decimal
        
        def convert_decimals(obj):
            """Recursively convert Decimal values to float in nested structures"""
            if isinstance(obj, Decimal):
                return float(obj)
            elif isinstance(obj, dict):
                return {str(key): convert_decimals(value) for key, value in obj.items()}
            elif isinstance(obj, (list, tuple)):
                return [convert_decimals(item) for item in obj]
            elif isinstance(obj, (str, int, float, bool, type(None))):
                return obj
            elif hasattr(obj, '__dict__'):
                try:
                    return convert_decimals(obj.__dict__)
                except:
                    return str(obj)
            else:
                try:
                    return str(obj)
                except:
                    return None
        
        def decimal_default(obj):
            if isinstance(obj, Decimal):
                return float(obj)
            try:
                return str(obj)
            except:
                return None
        
        try:
            # Convert all Decimals in the data structure first
            converted_data = convert_decimals(data)
            # Test serialization
            json_module.dumps(converted_data, default=decimal_default)
            return JsonResponse(converted_data, json_dumps_params={'default': decimal_default})
        except Exception as e:
            # Fallback: use string conversion for everything
            return JsonResponse({
                'success': False,
                'error': f'Serialization error: {str(e)}'
            }, json_dumps_params={'default': str})
    if request.method == 'POST':
        try:
            import json
            from decimal import Decimal

            from django.utils import timezone

            # Get order data from request
            try:
                order_data = json.loads(request.body)
            except json.JSONDecodeError as e:
                return safe_json({
                    'success': False,
                    'error': f'Invalid JSON data: {str(e)}'
                })
            
            # Validate required fields
            if not order_data.get('customerName'):
                return safe_json({
                    'success': False,
                    'error': 'Customer name is required.'
                })
            
            if not order_data.get('mobileNumber'):
                return safe_json({
                    'success': False,
                    'error': 'Customer mobile number is required.'
                })
            
            if not order_data.get('orderType'):
                return safe_json({
                    'success': False,
                    'error': 'Order type is required.'
                })
            
            if not order_data.get('items') or len(order_data.get('items', [])) == 0:
                return safe_json({
                    'success': False,
                    'error': 'At least one item is required to create an order.'
                })
            
            # Create or get customer - handle duplicates by using filter().first()
            # If multiple customers exist with same phone, get the first one
            customer = Customer.objects.filter(phone=order_data['mobileNumber']).first()
            if not customer:
                # No customer found, create new one
                customer = Customer.objects.create(
                    phone=order_data['mobileNumber'],
                    name=order_data['customerName'],
                    email=f"{order_data['customerName'].lower().replace(' ', '')}@example.com"
                )
            else:
                # Customer exists, update name if provided and different
                if order_data.get('customerName') and customer.name != order_data['customerName']:
                    customer.name = order_data['customerName']
                    customer.save()
            
            # Build order notes for repair orders
            notes_parts = []
            if order_data['orderType'] == 'repair':
                if order_data.get('repair_urgency'):
                    urgency_display = 'Urgent' if order_data.get('repair_urgency') == 'urgent' else 'Not Urgent'
                    notes_parts.append(f'Urgency: {urgency_display}')
                if order_data.get('sewing_style'):
                    sewing_style_display = 'Straight Stitch' if order_data.get('sewing_style') == 'straight_stitch' else 'Zigzag Stitch'
                    notes_parts.append(f'Sewing Style: {sewing_style_display}')
            
            order_notes = '; '.join(notes_parts) if notes_parts else ''
            
            # Create order
            # Get payment method from order_data, default to 'cash'
            payment_method = order_data.get('payment_method', order_data.get('paymentMethod', 'cash'))
            order = Order.objects.create(
                customer=customer,
                order_type=order_data['orderType'],
                status='pending',
                total_amount=Decimal(str(order_data['totalCost'])),
                paid_amount=Decimal(str(order_data['totalCost'])),
                balance=Decimal('0'),
                payment_method=payment_method,
                notes=order_notes
            )
            
            # Generate unique identifier
            order.generate_order_identifier()
            order.save()
            
            # Create order items
            for item in order_data['items']:
                # Get the product ID, name and class from the item
                product_id = item.get('id') or item.get('product_id')
                product_name = item.get('name', f"{order_data['orderType']} Service")
                item_class = item.get('class', 'standard')
                
                # For rental orders, find the exact product that was selected
                if order_data['orderType'] in ['rent', 'rental']:
                    # First try to find by product ID (most reliable)
                    if product_id:
                        try:
                            product = Product.objects.get(
                                id=int(product_id),
                                product_type='rental',
                                is_active=True
                            )
                        except (Product.DoesNotExist, ValueError, TypeError):
                            product = None
                    else:
                        product = None
                    
                    # If not found by ID, try to find by exact name
                    if not product:
                        try:
                            product = Product.objects.get(
                                name=product_name,
                                product_type='rental',
                                is_active=True
                            )
                        except Product.DoesNotExist:
                            # Try partial match as last resort
                            product = Product.objects.filter(
                                product_type='rental',
                                name__icontains=product_name,
                                rental_status='available',
                                is_active=True
                            ).first()
                    
                    # If still not found, validate that product must exist (do not auto-create)
                    if not product:
                        # Try alternative name format
                        alternative_name = f"Rental - {product_name} (Class {item_class})"
                        validation = validate_products_exist(
                            [product_name, alternative_name],
                            product_type='rental',
                            order_type='rent'
                        )
                        
                        if not validation['valid']:
                            return safe_json({
                                'success': False,
                                'error': f'Rental product "{product_name}" not found in database. Please add it to inventory first.',
                                'missing_products': validation['missing_products']
                            })
                        
                        # Try to get by alternative name if original didn't work
                        try:
                            product = Product.objects.get(
                                name=alternative_name,
                                product_type='rental',
                                is_archived=False,
                                is_active=True
                            )
                        except Product.DoesNotExist:
                            return safe_json({
                                'success': False,
                                'error': f'Rental product "{product_name}" not found in database. Please add it to inventory first.'
                            })
                    
                    # Check if product is available for rental
                    if product.rental_status != 'available':
                        return safe_json({
                            'success': False, 
                            'error': f'Product {product.name} is not available for rental (status: {product.rental_status})'
                        })
                    
                    # Update product rental status IMMEDIATELY
                    product.rental_status = 'rented'
                    product.current_rental_order = order
                    product.rental_start_date = timezone.now()
                    # Set due date (3 days from now for rental)
                    product.rental_due_date = timezone.now() + timezone.timedelta(days=3)
                    product.save()
                    
                    # Also use update_inventory_for_order for consistency
                    update_inventory_for_order(product, item.get('quantity', 1), order, 'rent')
                    
                else:
                    # For repair/customize orders, products must exist in inventory
                    # NO AUTO-CREATION - prevents static data
                    
                    # First, try to use customize_product_id if provided (for customize orders)
                    product = None
                    if order_data.get('orderType') == 'customize' and order_data.get('customize_product_id'):
                        try:
                            product = Product.objects.get(
                                id=order_data.get('customize_product_id'),
                                product_type='service',
                                is_archived=False,
                                is_active=True
                            )
                        except (Product.DoesNotExist, ValueError):
                            product = None
                    
                    # If no product found by ID, try to find by name
                    if not product:
                        # For repair orders, try repair-specific formats first
                        if order_data.get('orderType') == 'repair':
                            repair_type = product_name.lower().strip()
                            # Try multiple formats for repair orders
                            repair_name_formats = [
                                f"Repair - {repair_type.title()} (Class {item_class})",
                                f"repair - {repair_type.title()} (Class {item_class})",
                                f"Repair - {repair_type.title()}",
                                f"repair - {repair_type.title()}",
                                f"Repair - {repair_type} (Class {item_class})",
                                f"repair - {repair_type} (Class {item_class})",
                                f"{order_data['orderType']} - {product_name} (Class {item_class})",
                                product_name,  # Try exact name as fallback
                            ]
                            
                            for repair_name_format in repair_name_formats:
                                try:
                                    product = Product.objects.get(
                                        name=repair_name_format,
                                        product_type='service',
                                        is_archived=False,
                                        is_active=True
                                    )
                                    break
                                except Product.DoesNotExist:
                                    continue
                            
                            # If still not found, try case-insensitive search
                            if not product:
                                try:
                                    product = Product.objects.get(
                                        name__icontains=f"Repair - {repair_type}",
                                        product_type='service',
                                        is_archived=False,
                                        is_active=True
                                    )
                                except (Product.DoesNotExist, Product.MultipleObjectsReturned):
                                    # Try just the repair type in product name
                                    try:
                                        product = Product.objects.filter(
                                            name__icontains=repair_type,
                                            product_type='service',
                                            is_archived=False,
                                            is_active=True
                                        ).first()
                                    except:
                                        pass
                            
                            # If still not found, try to use a generic "Repair Service" product as fallback
                            if not product:
                                try:
                                    # Try to find a generic "Repair Service" product
                                    generic_repair_names = [
                                        "Repair Service",
                                        "repair service",
                                        "Repair - General",
                                        "repair - general",
                                        "General Repair Service",
                                        "general repair service"
                                    ]
                                    for generic_name in generic_repair_names:
                                        try:
                                            product = Product.objects.get(
                                                name=generic_name,
                                                product_type='service',
                                                is_archived=False,
                                                is_active=True
                                            )
                                            break
                                        except Product.DoesNotExist:
                                            continue
                                    
                                    # If generic name not found, try case-insensitive search
                                    if not product:
                                        try:
                                            product = Product.objects.filter(
                                                name__icontains="repair service",
                                                product_type='service',
                                                is_archived=False,
                                                is_active=True
                                            ).first()
                                        except:
                                            pass
                                    
                                    # Last resort: find ANY active repair service product
                                    if not product:
                                        try:
                                            product = Product.objects.filter(
                                                product_type='service',
                                                is_archived=False,
                                                is_active=True,
                                                name__icontains="repair"
                                            ).first()
                                        except:
                                            pass
                                    
                                    # Auto-create service product if it doesn't exist
                                    if not product:
                                        try:
                                            # Get or create a category for repair services
                                            from .models import Category
                                            repair_category, _ = Category.objects.get_or_create(
                                                name='Repair Services',
                                                defaults={'description': 'Repair service products'}
                                            )
                                            
                                            # Create the service product
                                            product = Product.objects.create(
                                                name=f"Repair - {product_name.title()}",
                                                product_type='service',
                                                category=repair_category,
                                                description=f"Repair service: {product_name}",
                                                price=Decimal('0'),
                                                cost=Decimal('0'),
                                                quantity=0,
                                                is_active=True,
                                                is_archived=False
                                            )
                                        except Exception as create_error:
                                            import logging
                                            logger = logging.getLogger(__name__)
                                            logger.error(f"Error auto-creating repair service product: {create_error}")
                                            return safe_json({
                                                'success': False,
                                                'error': f'Repair service product "{product_name}" not found in inventory and could not be auto-created. Please add a service product for this repair type to inventory first before creating orders.',
                                                'missing_products': [product_name]
                                            })
                                except Exception as e:
                                    # Log but don't fail - we'll handle error below
                                    import logging
                                    logger = logging.getLogger(__name__)
                                    logger.warning(f"Error finding generic repair service: {e}")
                        else:
                            # For non-repair orders, use standard format
                            product_name_to_find = f"{order_data['orderType']} - {product_name} (Class {item_class})"
                            
                            # Try to find existing product
                            try:
                                product = Product.objects.get(
                                    name=product_name_to_find,
                                    product_type='service',
                                    is_archived=False,
                                    is_active=True
                                )
                            except Product.DoesNotExist:
                                # Try alternative name formats for customize orders
                                if order_data.get('orderType') == 'customize':
                                    # Try multiple formats for customize orders
                                    customize_name_formats = [
                                        f"Customize - {product_name.title()} (Class {item_class})",
                                        f"customize - {product_name.title()} (Class {item_class})",
                                        f"Customize - {product_name.title()}",
                                        f"customize - {product_name.title()}",
                                        f"{order_data['orderType']} - {product_name} (Class {item_class})",
                                        product_name,  # Try exact name as fallback
                                    ]
                                    
                                    for customize_name_format in customize_name_formats:
                                        try:
                                            product = Product.objects.get(
                                                name=customize_name_format,
                                                product_type='service',
                                                is_archived=False,
                                                is_active=True
                                            )
                                            break
                                        except Product.DoesNotExist:
                                            continue
                                    
                                    # Try case-insensitive search
                                    if not product:
                                        try:
                                            product = Product.objects.filter(
                                                name__icontains=f"Customize - {product_name}",
                                                product_type='service',
                                                is_archived=False,
                                                is_active=True
                                            ).first()
                                        except:
                                            pass
                                    
                                    # For PE orders, try to find by PE type
                                    if not product and order_data.get('type_of_customize') == 'pe' and order_data.get('pe_type'):
                                        pe_type = order_data.get('pe_type', '').title()
                                        try:
                                            # Try "PE - {pe_type}" format
                                            product = Product.objects.filter(
                                                name__icontains=f"PE - {pe_type}",
                                                product_type='service',
                                                is_archived=False,
                                                is_active=True
                                            ).first()
                                        except:
                                            # Try just the pe_type
                                            try:
                                                product = Product.objects.filter(
                                                    name__icontains=pe_type,
                                                    product_type='service',
                                                    is_archived=False,
                                                    is_active=True
                                                ).first()
                                            except:
                                                pass
                                    
                                    # Try to find any customize service product as fallback
                                    if not product:
                                        try:
                                            product = Product.objects.filter(
                                                name__icontains="customize",
                                                product_type='service',
                                                is_archived=False,
                                                is_active=True
                                            ).first()
                                        except:
                                            pass
                                else:
                                    # For non-customize orders, try alternative name format
                                    try:
                                        product = Product.objects.get(
                                            name=product_name,
                                            product_type='service',
                                            is_archived=False,
                                            is_active=True
                                        )
                                    except Product.DoesNotExist:
                                        pass
                    
                    # Final check - ensure product was found before creating OrderItem
                    if not product:
                        # Auto-create customize or repair products if they don't exist
                        if order_data.get('orderType') == 'customize':
                            try:
                                from .models import Category
                                service_category, _ = Category.objects.get_or_create(
                                    name='Customize Services',
                                    defaults={'description': 'Customize service products'}
                                )
                                
                                product = Product.objects.create(
                                    name=product_name,
                                    product_type='service',
                                    category=service_category,
                                    description=f"Customize Service: {product_name}",
                                    price=Decimal('0'),
                                    cost=Decimal('0'),
                                    quantity=0,
                                    is_active=True,
                                    is_archived=False
                                )
                            except Exception as create_error:
                                import logging
                                logger = logging.getLogger(__name__)
                                logger.error(f"Error auto-creating customize product: {create_error}")
                                return safe_json({
                                    'success': False,
                                    'error': f'Error creating customize service product "{product_name}": {str(create_error)}'
                                })
                        elif order_data.get('orderType') == 'repair':
                            # Auto-create repair product
                            try:
                                from .models import Category
                                service_category, _ = Category.objects.get_or_create(
                                    name='Repair Services',
                                    defaults={'description': 'Repair service products'}
                                )
                                
                                suggested_name = f"Repair - {product_name.title()}"
                                product = Product.objects.create(
                                    name=suggested_name,
                                    product_type='service',
                                    category=service_category,
                                    description=f"Repair Service: {product_name}",
                                    price=Decimal('0'),
                                    cost=Decimal('0'),
                                    quantity=0,
                                    is_active=True,
                                    is_archived=False
                                )
                            except Exception as create_error:
                                import logging
                                logger = logging.getLogger(__name__)
                                logger.error(f"Error auto-creating repair product: {create_error}")
                                return safe_json({
                                    'success': False,
                                    'error': f'Error creating repair service product "{product_name}": {str(create_error)}'
                                })
                        else:
                            # For other order types, try to auto-create as well
                            try:
                                from .models import Category
                                service_category, _ = Category.objects.get_or_create(
                                    name='Services',
                                    defaults={'description': 'Service products'}
                                )
                                
                                product = Product.objects.create(
                                    name=product_name,
                                    product_type='service',
                                    category=service_category,
                                    description=f"Service: {product_name}",
                                    price=Decimal('0'),
                                    cost=Decimal('0'),
                                    quantity=0,
                                    is_active=True,
                                    is_archived=False
                                )
                            except Exception as create_error:
                                import logging
                                logger = logging.getLogger(__name__)
                                logger.error(f"Error auto-creating service product: {create_error}")
                                return safe_json({
                                    'success': False,
                                    'error': f'Service product "{product_name}" not found in inventory and could not be auto-created. Please add it to inventory first before creating orders.'
                                })
                
                # Only create OrderItem if product is valid
                if product:
                    # Always fetch the actual product price from database
                    # Refresh product from database to ensure we have latest price
                    product.refresh_from_db()
                    
                    # Get actual unit price from product database
                    # Use product.price if available, otherwise fall back to frontend cost
                    actual_unit_price = product.price if product.price and product.price > 0 else Decimal(str(item.get('cost', 0)))
                    
                    # Calculate total price based on actual unit price and quantity
                    quantity = int(item.get('quantity', 1))
                    actual_total_price = actual_unit_price * quantity
                    
                    # For customize orders, add any additional costs from frontend if product base price is being used
                    if order_data['orderType'] == 'customize' and product.price and product.price > 0:
                        # If frontend cost is higher than base product price, it might include additional costs
                        frontend_cost = Decimal(str(item.get('cost', 0)))
                        if frontend_cost > actual_total_price:
                            # Add the difference as additional cost
                            additional_cost = frontend_cost - actual_total_price
                            actual_total_price += additional_cost
                    
                    OrderItem.objects.create(
                        order=order,
                        product=product,
                        quantity=quantity,
                        unit_price=actual_unit_price,
                        total_price=actual_total_price
                    )
                else:
                    # This should not happen due to the check above, but add as safety
                    return safe_json({
                        'success': False,
                        'error': f'Failed to find product for order item. Product name: "{product_name}"'
                    })
            
            # Deduct materials and ensure all are recorded
            if order_data['orderType'] == 'repair':
                deduct_repair_materials(order, order_data)
            elif order_data['orderType'] == 'customize':
                deduct_customize_materials(order, order_data)
            
            # Ensure all materials used are recorded in transactions
            # This is a safety check to ensure completeness
            ensure_materials_recorded_for_order(order, order_data)
            
            # Recalculate order total based on actual item prices from database
            from django.db.models import Sum
            actual_total = order.items.aggregate(total=Sum('total_price'))['total'] or Decimal('0')
            
            # Update order with actual total amount
            order.total_amount = actual_total
            order.paid_amount = actual_total  # For now, assume full payment
            order.balance = Decimal('0')
            order.save()
            
            # Note: Sales record will be created automatically when order status changes to 'completed'
            
            # Generate QR code
            qr_data = {
                'order_id': order.id,
                'customer_name': customer.name,
                'total_amount': str(order.total_amount),
                'order_type': order.order_type,
                'date': order.created_at.isoformat()
            }
            
            return safe_json({
                'success': True,
                'order_id': order.id,
                'order_identifier': order.order_identifier,
                'message': 'Order created successfully'
            })
            
        except Exception as e:
            import traceback

            # Get error message, ensuring no Decimal values
            try:
                error_msg = str(e)
                # Test if error message can be serialized
                test_dict = {'error': error_msg}
                json.dumps(test_dict, default=str)
            except Exception as serialization_error:
                # If serialization fails, use a safe error message
                error_msg = f"Error occurred: {type(e).__name__}"
            return safe_json({
                'success': False,
                'error': error_msg
            })
    
    return safe_json({'success': False, 'error': 'Invalid request method'})


@login_required
@csrf_protect
def order_items(request, order_id):
    order = get_object_or_404(Order, id=order_id)
    
    if request.method == 'POST':
        product_id = request.POST.get('product_id')
        quantity = int(request.POST.get('quantity', 1))
        
        product = get_object_or_404(Product, id=product_id)
        
        # Check availability
        if product.product_type == 'rental':
            if product.quantity < quantity:
                messages.error(request, f'Only {product.quantity} {product.name} available for rent.')
                return redirect('order_items', order_id=order_id)
        else:
            if product.quantity < quantity:
                messages.error(request, f'Only {product.quantity} {product.name} available in stock.')
                return redirect('order_items', order_id=order_id)
        
        # Create order item
        OrderItem.objects.create(
            order=order,
            product=product,
            quantity=quantity,
            unit_price=product.price
        )
        
        # Update inventory
        if product.product_type == 'rental':
            InventoryTransaction.objects.create(
                product=product,
                transaction_type='rental_out',
                quantity=-quantity,
                reference_order=order,
                notes=f'Rental order {order.order_id}'
            )
        else:
            if order.order_type == 'customize' and product.product_type == 'service':
                deduct_customize_materials(order)
            else:
                InventoryTransaction.objects.create(
                    product=product,
                    transaction_type='out',
                    quantity=-quantity,
                    reference_order=order,
                    notes=f'Order {order.order_id}'
                )
                product.quantity -= quantity
                product.save()
        
        # Update order total
        from decimal import Decimal
        total = order.items.aggregate(total=Sum('total_price'))['total'] or 0
        order.total_amount = Decimal(str(total))
        order.save()
        
        messages.success(request, f'{product.name} added to order.')
        return redirect('order_items', order_id=order_id)
    
    products = Product.objects.filter(is_active=True)
    order_items = order.items.all()
    
    context = {
        'order': order,
        'products': products,
        'order_items': order_items,
    }
    
    return render(request, 'business/order_items.html', context)


@login_required
def update_order_item(request, order_id, item_id):
    """Update order item quantity"""
    order = get_object_or_404(Order, id=order_id)
    order_item = get_object_or_404(OrderItem, id=item_id, order=order)
    
    if request.method == 'POST':
        try:
            new_quantity = int(request.POST.get('quantity', 1))
            if new_quantity < 1:
                messages.error(request, 'Quantity must be at least 1.')
                return redirect('order_items', order_id=order_id)
            
            product = order_item.product
            old_quantity = order_item.quantity
            quantity_diff = new_quantity - old_quantity
            
            # Check availability if increasing quantity
            if quantity_diff > 0:
                if product.product_type == 'rental':
                    if product.quantity < quantity_diff:
                        messages.error(request, f'Only {product.quantity} {product.name} available for rent.')
                        return redirect('order_items', order_id=order_id)
                else:
                    if product.quantity < quantity_diff:
                        messages.error(request, f'Only {product.quantity} {product.name} available in stock.')
                        return redirect('order_items', order_id=order_id)
            
            # Update order item
            order_item.quantity = new_quantity
            order_item.save()
            
            # Update inventory
            if quantity_diff != 0:
                if product.product_type == 'rental':
                    # For rentals, adjust the rental status
                    if quantity_diff > 0:
                        InventoryTransaction.objects.create(
                            product=product,
                            transaction_type='rental_out',
                            quantity=-quantity_diff,
                            reference_order=order,
                            notes=f'Updated rental order {order.order_id} - increased quantity'
                        )
                        product.quantity -= quantity_diff
                    else:
                        InventoryTransaction.objects.create(
                            product=product,
                            transaction_type='rental_in',
                            quantity=abs(quantity_diff),
                            reference_order=order,
                            notes=f'Updated rental order {order.order_id} - decreased quantity'
                        )
                        product.quantity += abs(quantity_diff)
                else:
                    # For non-rental items
                    if quantity_diff > 0:
                        InventoryTransaction.objects.create(
                            product=product,
                            transaction_type='out',
                            quantity=-quantity_diff,
                            reference_order=order,
                            notes=f'Updated order {order.order_id} - increased quantity'
                        )
                        product.quantity -= quantity_diff
                    else:
                        InventoryTransaction.objects.create(
                            product=product,
                            transaction_type='in',
                            quantity=abs(quantity_diff),
                            reference_order=order,
                            notes=f'Updated order {order.order_id} - decreased quantity'
                        )
                        product.quantity += abs(quantity_diff)
                
                product.save()
            
            # Update order total
            from decimal import Decimal
            total = order.items.aggregate(total=Sum('total_price'))['total'] or 0
            order.total_amount = Decimal(str(total))
            order.save()
            
            messages.success(request, f'{product.name} quantity updated to {new_quantity}.')
        except ValueError:
            messages.error(request, 'Invalid quantity value.')
        except Exception as e:
            messages.error(request, f'Error updating item: {str(e)}')
    
    return redirect('order_items', order_id=order_id)


@login_required
def delete_order_item(request, order_id, item_id):
    """Delete order item and restore inventory"""
    order = get_object_or_404(Order, id=order_id)
    order_item = get_object_or_404(OrderItem, id=item_id, order=order)
    
    if request.method == 'POST':
        try:
            product = order_item.product
            quantity = order_item.quantity
            
            # Restore inventory
            if product.product_type == 'rental':
                InventoryTransaction.objects.create(
                    product=product,
                    transaction_type='rental_in',
                    quantity=quantity,
                    reference_order=order,
                    notes=f'Removed from rental order {order.order_id}'
                )
                # Restore rental status if this was the only rental
                other_rental_items = OrderItem.objects.filter(
                    order__status__in=['pending', 'confirmed', 'in_progress'],
                    product=product,
                    order__order_type__in=['rent', 'rental']
                ).exclude(id=item_id)
                
                if not other_rental_items.exists():
                    product.rental_status = 'available'
                    product.current_rental_order = None
                    product.rental_start_date = None
                    product.rental_due_date = None
                
                product.quantity += quantity
            else:
                InventoryTransaction.objects.create(
                    product=product,
                    transaction_type='in',
                    quantity=quantity,
                    reference_order=order,
                    notes=f'Removed from order {order.order_id}'
                )
                product.quantity += quantity
            
            product.save()
            
            # Delete order item
            product_name = order_item.product.name
            order_item.delete()
            
            # Update order total
            from decimal import Decimal
            total = order.items.aggregate(total=Sum('total_price'))['total'] or 0
            order.total_amount = Decimal(str(total))
            order.save()
            
            messages.success(request, f'{product_name} removed from order.')
        except Exception as e:
            messages.error(request, f'Error deleting item: {str(e)}')
    
    return redirect('order_items', order_id=order_id)


@login_required
def order_payment(request, order_id):
    order = get_object_or_404(Order, id=order_id)
    
    if request.method == 'POST':
        from decimal import Decimal
        paid_amount = Decimal(str(request.POST.get('paid_amount', 0)))
        payment_method = request.POST.get('payment_method', 'cash')
        order.paid_amount = paid_amount
        order.payment_method = payment_method  # Save payment method to order
        order.status = 'confirmed'
        order.save()
        
        # Create or update sales record
        sales_record, created = Sales.objects.get_or_create(
            order=order,
            defaults={
                'amount': paid_amount,
                'payment_method': payment_method
            }
        )
        
        # If sales record already exists, update it
        if not created:
            sales_record.amount = paid_amount
            sales_record.payment_method = payment_method
            sales_record.save()
        
        # Generate QR code
        generate_qr_code(order)
        
        # Send SMS notification
        send_sms_notification(order)
        
        messages.success(request, 'Order confirmed and payment recorded.')
        return redirect('order_receipt', order_id=order.id)
    
    return render(request, 'business/order_payment.html', {'order': order})


@login_required
def order_receipt(request, order_id):
    order = get_object_or_404(Order, id=order_id)
    qr_code = QRCode.objects.filter(order=order).first()
    
    return render(request, 'business/order_receipt.html', {
        'order': order,
        'qr_code': qr_code
    })

@login_required
@csrf_protect
@require_http_methods(["GET", "POST"])
def complete_order(request, order_id):
    order = get_object_or_404(Order, id=order_id)
    
    # Ensure CSRF token is set for GET requests
    if request.method == 'GET':
        get_token(request)
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'complete':
            # Update order status to completed
            order.status = 'completed'
            order.save()
            messages.success(request, f'Order {order.order_identifier} has been marked as completed.')
        elif action == 'return':
            # Mark rental order as returned
            from .rental_manager import RentalStatusManager
            updated_count = RentalStatusManager.mark_products_as_available(order, request.user)
            order.status = 'returned'
            order.save()
            
            # Double-check: Ensure all products in this order are marked as available
            # This is a safety check in case mark_products_as_available missed any
            for item in order.items.filter(product__product_type='rental'):
                if item.product.rental_status != 'available':
                    item.product.rental_status = 'available'
                    item.product.current_rental_order = None
                    item.product.rental_start_date = None
                    item.product.rental_due_date = None
                    item.product.save()
            
            messages.success(request, f'Order {order.order_identifier} has been marked as returned. {updated_count} item(s) updated.')
        elif action == 'repair_done':
            order.status = 'repair_done'
            order.save()
            messages.success(request, f'Order {order.order_identifier} repair is marked as done.')
        elif action == 'customize_done':
            order.status = 'ready_to_pick_up'
            order.save()
            messages.success(request, f'Order {order.order_identifier} customization is marked as done.')
        
        # Preserve the filter type from the referrer or request
        referer = request.META.get('HTTP_REFERER', '')
        filter_type = request.GET.get('filter_type', None)
        
        # Try to get filter type from referer URL
        if not filter_type and referer:
            if 'type=repair' in referer:
                filter_type = 'repair'
            elif 'type=rental' in referer or 'type=rent' in referer:
                filter_type = 'rental'
            elif 'type=customize' in referer:
                filter_type = 'customize'
            elif 'type=all' in referer:
                filter_type = 'all'
        
        # Default to 'all' if no filter found
        if not filter_type:
            filter_type = 'all'
        
        # Redirect back to orders with the same filter
        if filter_type == 'all':
            return redirect('orders')
        else:
            from django.urls import reverse
            return redirect(f'{reverse("orders")}?type={filter_type}')
    
    # Get all order items with product details
    order_items = order.items.select_related('product', 'product__category').all()
    
    # Separate rental items from other items
    rental_items = [item for item in order_items if item.product.product_type == 'rental']
    all_items = list(order_items)
    
    # Get staff members for assignment (only active staff users)
    from django.contrib.auth.models import User
    staff_members = User.objects.filter(is_staff=True, is_active=True).order_by('first_name', 'username')
    
    context = {
        'order': order,
        'order_items': order_items,
        'rental_items': rental_items,
        'all_items': all_items,
        'staff_members': staff_members,
    }
    
    return render(request, 'business/complete_order.html', context)


@login_required
@csrf_protect
def archive_completed_orders(request):
    """Archive all completed orders - hide them from orders list but keep in reports"""
    if request.method == 'POST':
        # Get all completed orders that are not yet archived
        completed_orders = Order.objects.filter(status='completed', is_archived=False)
        count = completed_orders.count()
        
        # Archive all completed orders
        completed_orders.update(is_archived=True)
        
        if count > 0:
            messages.success(request, f'Successfully archived {count} completed order(s). They will be hidden from the orders list but remain available in reports.')
        else:
            messages.info(request, 'No completed orders to archive.')
        
        return redirect('orders')
    
    # If not POST, redirect to orders page
    return redirect('orders')


@login_required
def inventory_list(request):
    """Inventory list with real-time database connections and status tracking"""
    # Get filter parameter
    product_type_filter = request.GET.get('type', 'all')
    
    # Get real-time inventory status
    inventory_status = get_inventory_status()
    
    # Base queryset with real-time status - EXCLUDE service/customize products from main inventory
    # Service products should only show in their own dedicated section or when explicitly filtered
    products = Product.objects.filter(
        is_archived=False
    ).exclude(product_type='service').order_by('name')
    
    # Apply filter if specified
    if product_type_filter == 'rental':
        products = products.filter(product_type='rental')
    elif product_type_filter == 'material':
        products = products.filter(product_type='material')
    # else: 'all' - already filtered to exclude service products above
    # Service/customize products are shown in a separate table below, not in these tabs
    
    # FIX: Check active rental orders to ensure rental products show correct status
    # Get all products that are in active rental orders (source of truth)
    from .models import Order, OrderItem
    active_rental_orders = Order.objects.filter(
        order_type__in=['rent', 'rental'],
        status__in=['rented', 'pending', 'almost_due', 'due', 'overdue', 'in_progress']
    ).prefetch_related('items__product')
    
    # Build a set of product IDs that are actually rented (from active orders)
    actually_rented_product_ids = set()
    for order in active_rental_orders:
        for item in order.items.filter(product__product_type='rental'):
            actually_rented_product_ids.add(item.product.id)
    
    # Update rental_status for products that are in active orders but have wrong status
    # AND fix products that have rental_status='rented' but are NOT in any active order
    for product in products.filter(product_type='rental'):
        if product.id in actually_rented_product_ids:
            # Product is in an active order, ensure it's marked as rented
            if product.rental_status != 'rented':
                product.rental_status = 'rented'
                # Get the order for this product
                order_item = OrderItem.objects.filter(
                    product=product,
                    order__order_type__in=['rent', 'rental'],
                    order__status__in=['rented', 'pending', 'almost_due', 'due', 'overdue', 'in_progress']
                ).select_related('order').first()
                if order_item and order_item.order:
                    product.current_rental_order = order_item.order
                    if not product.rental_start_date:
                        product.rental_start_date = order_item.order.created_at
                    if not product.rental_due_date and order_item.order.due_date:
                        product.rental_due_date = order_item.order.due_date
                product.save()
        else:
            # Product is NOT in any active order, but might still have rental_status='rented'
            # This happens when orders are returned/completed but product status wasn't updated
            if product.rental_status == 'rented':
                # Check if there's any active order for this product
                has_active_order = OrderItem.objects.filter(
                    product=product,
                    order__order_type__in=['rent', 'rental'],
                    order__status__in=['rented', 'pending', 'almost_due', 'due', 'overdue', 'in_progress']
                ).exists()
                
                if not has_active_order:
                    # No active order found, mark as available
                    product.rental_status = 'available'
                    product.current_rental_order = None
                    product.rental_start_date = None
                    product.rental_due_date = None
                    product.save()
    
    # Calculate product type statistics - EXCLUDE service products from counts
    rental_products_count = Product.objects.filter(
        product_type='rental',
        is_archived=False
    ).count()
    material_products_count = Product.objects.filter(
        product_type='material',
        is_archived=False
    ).count()
    service_products_count = Product.objects.filter(
        product_type='service',
        is_archived=False
    ).count()
    total_products = rental_products_count + material_products_count  # Don't include service in total
    
    # Get low stock and out of stock products - EXCLUDE service products
    low_stock_products = Product.objects.filter(
        is_archived=False,
        quantity__lte=models.F('min_quantity')
    ).exclude(product_type='service').order_by('quantity')
    
    out_of_stock_products = Product.objects.filter(
        is_archived=False,
        quantity=0
    ).exclude(product_type='service').order_by('name')
    
    # Get overdue rental products
    overdue_rentals = Product.objects.filter(
        is_archived=False,
        product_type='rental',
        rental_status='rented',
        rental_due_date__lt=timezone.now()
    ).order_by('rental_due_date')
    
    # Calculate available items count
    # Available items = rental items with status 'available' + material items with quantity > 0
    available_rental_items = Product.objects.filter(
        is_archived=False,
        product_type='rental',
        rental_status='available'
    ).count()
    
    available_material_items = Product.objects.filter(
        is_archived=False,
        product_type='material',
        quantity__gt=0
    ).count()
    
    available_items = available_rental_items + available_material_items
    
    # Get customize products for the separate section below the inventory table
    # Use the customize_product_manager to ensure no duplicates are shown
    from business.customize_product_manager import \
        get_unique_customize_products
    customize_products = get_unique_customize_products()
    
    # Pagination for products
    paginator = Paginator(products, 12)  # Show 12 products per page
    page = request.GET.get('page')
    try:
        products = paginator.page(page)
    except PageNotAnInteger:
        products = paginator.page(1)
    except EmptyPage:
        products = paginator.page(paginator.num_pages)
    
    context = {
        'products': products,
        'customize_products': customize_products,  # For the separate customize product section
        'rental_products': rental_products_count,
        'material_products': material_products_count,
        'service_products': service_products_count,
        'total_products': total_products,
        'available_items': available_items,  # Add available items count
        'current_filter': product_type_filter,
        'inventory_status': inventory_status,
        'low_stock_products': low_stock_products,
        'out_of_stock_products': out_of_stock_products,
        'overdue_rentals': overdue_rentals,
        'low_stock_count': inventory_status['low_stock'],
        'out_of_stock_count': inventory_status['out_of_stock'],
        'rental_available': inventory_status['rental_available'],
        'rental_rented': inventory_status['rental_rented'],
        'rental_overdue': inventory_status['rental_overdue'],
    }
    
    return render(request, 'business/inventory.html', context)


@login_required
@csrf_protect
def archive_product(request, product_id):
    product = get_object_or_404(Product, id=product_id)
    if request.method == 'POST':
        product.is_archived = True
        product.is_active = False  # Also deactivate the product
        product.save()
        messages.success(request, f'{product.name} has been archived successfully.')
    return redirect('inventory')


@login_required
@csrf_protect
def return_rental_product(request, product_id):
    product = get_object_or_404(Product, id=product_id)
    if request.method == 'POST':
        if product.rental_status == 'rented':
            # Return the product
            product.rental_status = 'available'
            product.current_rental_order = None
            product.rental_start_date = None
            product.rental_due_date = None
            product.save()
            
            messages.success(request, f'{product.name} has been returned successfully.')
        else:
            messages.error(request, f'{product.name} is not currently rented.')
    return redirect('inventory')


@login_required
@csrf_protect
def send_overdue_notification(request, product_id):
    product = get_object_or_404(Product, id=product_id)
    if request.method == 'POST':
        if product.is_overdue and product.current_rental_order:
            # Send SMS notification
            customer = product.current_rental_order.customer
            message = f"Hello {customer.name}, your rental item '{product.name}' is overdue. Please return it as soon as possible. Due date: {product.rental_due_date.strftime('%Y-%m-%d')}"
            
            # Here you would integrate with your SMS service
            # For now, we'll just show a success message
            messages.success(request, f'Overdue notification sent to {customer.name} for {product.name}')
        else:
            messages.error(request, f'{product.name} is not overdue.')
    return redirect('inventory')


@login_required
def rental_management(request):
    """View to manage all rental items and their status"""
    # Get all rental products
    rental_products = Product.objects.filter(
        product_type='rental',
        is_archived=False
    ).order_by('rental_status', 'name')
    
    # Calculate statistics from unfiltered queryset
    available_rentals = rental_products.filter(rental_status='available').count()
    rented_items = rental_products.filter(rental_status='rented').count()
    overdue_items = rental_products.filter(
        rental_status='rented',
        rental_due_date__lt=timezone.now()
    ).count()
    maintenance_items = rental_products.filter(rental_status='maintenance').count()
    
    # Pagination for rental products
    paginator = Paginator(rental_products, 12)  # Show 12 products per page
    page = request.GET.get('page')
    try:
        rental_products = paginator.page(page)
    except PageNotAnInteger:
        rental_products = paginator.page(1)
    except EmptyPage:
        rental_products = paginator.page(paginator.num_pages)
    
    context = {
        'products': rental_products,  # Changed to 'products' to match template
        'rental_products': rental_products,
        'available_rentals': available_rentals,
        'rented_items': rented_items,
        'overdue_items': overdue_items,
        'maintenance_items': maintenance_items,
    }
    
    return render(request, 'business/rental_management.html', context)


@login_required
def add_product_from_estimator(request):
    """API endpoint to add product to inventory from estimator"""
    if request.method == 'POST':
        try:
            import json
            from decimal import Decimal
            
            data = json.loads(request.body)
            
            # Get or create category
            category, created = Category.objects.get_or_create(
                name=data.get('category', 'General'),
                defaults={'description': f'Category for {data.get("category", "General")} products'}
            )
            
            # Create product
            product = Product.objects.create(
                name=data['name'],
                description=data.get('description', ''),
                category=category,
                product_type=data['product_type'],
                price=Decimal(str(data['price'])),
                cost=Decimal(str(data.get('cost', 0))),
                quantity=data.get('quantity', 1),
                min_quantity=data.get('min_quantity', 0),
                is_active=True,
                is_archived=False
            )
            
            return JsonResponse({
                'success': True,
                'product_id': product.id,
                'message': f'Product "{product.name}" added to inventory successfully'
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            })
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})


@login_required
def create_order_from_estimator(request):
    """API endpoint to create order from estimator"""
    def safe_json(data):
        """Helper function to safely serialize JSON with Decimal support"""
        import json as json_module
        from decimal import Decimal
        
        def convert_decimals(obj):
            """Recursively convert Decimal values to float in nested structures"""
            if isinstance(obj, Decimal):
                return float(obj)
            elif isinstance(obj, dict):
                return {str(key): convert_decimals(value) for key, value in obj.items()}
            elif isinstance(obj, (list, tuple)):
                return [convert_decimals(item) for item in obj]
            elif isinstance(obj, (str, int, float, bool, type(None))):
                return obj
            elif hasattr(obj, '__dict__'):
                try:
                    return convert_decimals(obj.__dict__)
                except:
                    return str(obj)
            else:
                try:
                    return str(obj)
                except:
                    return None
        
        def decimal_default(obj):
            if isinstance(obj, Decimal):
                return float(obj)
            try:
                return str(obj)
            except:
                return None
        
        try:
            # Convert all Decimals in the data structure first
            converted_data = convert_decimals(data)
            # Test serialization
            json_module.dumps(converted_data, default=decimal_default)
            return JsonResponse(converted_data, json_dumps_params={'default': decimal_default})
        except Exception as e:
            # Fallback: use string conversion for everything
            return JsonResponse({
                'success': False,
                'error': f'Serialization error: {str(e)}'
            }, json_dumps_params={'default': str})
    
    if request.method == 'POST':
        try:
            import json
            from decimal import Decimal

            from django.utils import timezone
            
            data = json.loads(request.body)
            
            # Create or get customer - handle duplicates by using filter().first()
            customer = Customer.objects.filter(phone=data.get('mobile_number', '00000000000')).first()
            if not customer:
                # No customer found, create new one
                customer = Customer.objects.create(
                    phone=data.get('mobile_number', '00000000000'),
                    name=data['customer_name'],
                    email=f"{data['customer_name'].lower().replace(' ', '')}@example.com"
                )
            else:
                # Customer exists, update name if provided and different
                if data.get('customer_name') and customer.name != data['customer_name']:
                    customer.name = data['customer_name']
                    customer.save()
            
            # Create order
            # Get payment method from data, default to 'cash'
            payment_method = data.get('payment_method', data.get('paymentMethod', 'cash'))
            order = Order.objects.create(
                customer=customer,
                order_type=data['service_type'],
                status='pending',
                total_amount=Decimal(str(data['total_cost'])),
                paid_amount=Decimal('0'),
                balance=Decimal(str(data['total_cost'])),
                payment_method=payment_method
            )
            
            # Generate unique identifier
            order.generate_order_identifier()
            order.save()
            
            # Create order item based on service type
            # NO AUTO-CREATION - products must exist in inventory
            if data['service_type'] == 'rent':
                # For rental, product must exist
                rent_type = data.get('rent_type', '').title()
                try:
                    # Try exact match first
                    product = Product.objects.get(
                        name__iexact=f"Rental - {rent_type}",
                        product_type='rental',
                        is_archived=False,
                        is_active=True
                    )
                except Product.DoesNotExist:
                    # Try without "Rental -" prefix
                    try:
                        product = Product.objects.get(
                            name__iexact=rent_type,
                            product_type='rental',
                            is_archived=False,
                            is_active=True
                        )
                    except Product.DoesNotExist:
                        return safe_json({
                            'success': False,
                            'error': f'Rental product "{rent_type}" not found in inventory. Please add it to inventory first.'
                        })
                
                # Always fetch the actual product price from database
                product.refresh_from_db()
                
                # Get actual unit price from product database
                # Use product.price if available, otherwise fall back to frontend cost
                quantity = int(data.get('rent_quantity', 1))
                actual_unit_price = product.price if product.price and product.price > 0 else Decimal(str(data.get('total_cost', 0)))
                actual_total_price = actual_unit_price * quantity
                
                OrderItem.objects.create(
                    order=order,
                    product=product,
                    quantity=quantity,
                    unit_price=actual_unit_price,
                    total_price=actual_total_price
                )
                
            elif data['service_type'] == 'repair':
                # For repair, product must exist - try multiple lookup strategies
                repair_type = data.get('repair_type', '').replace('_', ' ').title()
                repair_type_lower = repair_type.lower().strip()
                product = None
                
                # Try multiple formats for repair orders
                repair_name_formats = [
                    f"Repair - {repair_type}",
                    f"repair - {repair_type}",
                    f"Repair - {repair_type_lower.title()}",
                    repair_type,
                ]
                
                for repair_name_format in repair_name_formats:
                    try:
                        product = Product.objects.get(
                            name=repair_name_format,
                            product_type='service',
                            is_archived=False,
                            is_active=True
                        )
                        break
                    except Product.DoesNotExist:
                        continue
                
                # If still not found, try case-insensitive search
                if not product:
                    try:
                        product = Product.objects.get(
                            name__icontains=f"Repair - {repair_type_lower}",
                            product_type='service',
                            is_archived=False,
                            is_active=True
                        )
                    except (Product.DoesNotExist, Product.MultipleObjectsReturned):
                        # Try just the repair type in product name
                        try:
                            product = Product.objects.filter(
                                name__icontains=repair_type_lower,
                                product_type='service',
                                is_archived=False,
                                is_active=True
                            ).first()
                        except:
                            pass
                
                # If still not found, try to use a generic "Repair Service" product as fallback
                if not product:
                    try:
                        # Try to find a generic "Repair Service" product
                        generic_repair_names = [
                            "Repair Service",
                            "repair service",
                            "Repair - General",
                            "repair - general",
                            "General Repair Service",
                            "general repair service"
                        ]
                        for generic_name in generic_repair_names:
                            try:
                                product = Product.objects.get(
                                    name=generic_name,
                                    product_type='service',
                                    is_archived=False,
                                    is_active=True
                                )
                                break
                            except Product.DoesNotExist:
                                continue
                        
                        # If generic name not found, try case-insensitive search
                        if not product:
                            try:
                                product = Product.objects.filter(
                                    name__icontains="repair service",
                                    product_type='service',
                                    is_archived=False,
                                    is_active=True
                                ).first()
                            except:
                                pass
                        
                        # Last resort: find ANY active repair service product
                        if not product:
                            try:
                                product = Product.objects.filter(
                                    product_type='service',
                                    is_archived=False,
                                    is_active=True,
                                    name__icontains="repair"
                                ).first()
                            except:
                                pass
                        
                        # Auto-create service product if it doesn't exist
                        if not product:
                            try:
                                # Get or create a category for repair services
                                from .models import Category
                                repair_category, _ = Category.objects.get_or_create(
                                    name='Repair Services',
                                    defaults={'description': 'Repair service products'}
                                )
                                
                                # Get the repair type or product name from data
                                repair_type = data.get('repair_type', data.get('repairType', 'repair service'))
                                product_name_to_create = f"Repair - {repair_type.title()}"
                                
                                # Create the service product
                                product = Product.objects.create(
                                    name=product_name_to_create,
                                    product_type='service',
                                    category=repair_category,
                                    description=f"Repair service: {repair_type}",
                                    price=Decimal('0'),
                                    cost=Decimal('0'),
                                    quantity=0,
                                    is_active=True,
                                    is_archived=False
                                )
                            except Exception as create_error:
                                import logging
                                logger = logging.getLogger(__name__)
                                logger.error(f"Error auto-creating repair service product: {create_error}")
                                repair_type = data.get('repair_type', data.get('repairType', 'repair service'))
                                return safe_json({
                                    'success': False,
                                    'error': f'Repair service product for "{repair_type}" not found in inventory and could not be auto-created. Please add a service product for this repair type to inventory first before creating orders.',
                                    'missing_products': [repair_type]
                                })
                    except Exception as e:
                        # Log but don't fail - we'll handle error below
                        import logging
                        logger = logging.getLogger(__name__)
                        logger.warning(f"Error finding generic repair service: {e}")
                
                # If still no product found, return error
                if not product:
                    return safe_json({
                        'success': False,
                        'error': f'Repair service product for "{repair_type}" not found in inventory. Please add it to inventory first before creating orders.'
                    })
                
                # Always fetch the actual product price from database
                product.refresh_from_db()
                
                # Get actual unit price from product database
                # Use product.price if available, otherwise fall back to frontend cost
                quantity = 1
                actual_unit_price = product.price if product.price and product.price > 0 else Decimal(str(data.get('total_cost', 0)))
                actual_total_price = actual_unit_price * quantity
                
                # For repair orders, if frontend cost is higher, it might include additional costs
                frontend_cost = Decimal(str(data.get('total_cost', 0)))
                if frontend_cost > actual_total_price:
                    additional_cost = frontend_cost - actual_total_price
                    actual_total_price += additional_cost
                
                OrderItem.objects.create(
                    order=order,
                    product=product,
                    quantity=quantity,
                    unit_price=actual_unit_price,
                    total_price=actual_total_price
                )
                
            elif data['service_type'] == 'customize':
                # For customize, try to find product or auto-create if not found
                customize_type = data.get('customize_type', '').replace('_', ' ').title()
                product = None
                try:
                    product = Product.objects.get(
                        name__icontains=customize_type,
                        product_type='service',
                        is_archived=False,
                        is_active=True
                    )
                except (Product.DoesNotExist, Product.MultipleObjectsReturned):
                    # Try alternative searches
                    try:
                        product = Product.objects.filter(
                            name__icontains="Customize",
                            product_type='service',
                            is_archived=False,
                            is_active=True
                        ).first()
                    except:
                        pass
                
                # Auto-create customize product if not found
                if not product:
                    try:
                        from .models import Category
                        service_category, _ = Category.objects.get_or_create(
                            name='Customize Services',
                            defaults={'description': 'Customize service products'}
                        )
                        
                        product = Product.objects.create(
                            name=customize_type,
                            product_type='service',
                            category=service_category,
                            description=f"Customize Service: {customize_type}",
                            price=Decimal('0'),
                            cost=Decimal('0'),
                            quantity=0,
                            is_active=True,
                            is_archived=False
                        )
                    except Exception as create_error:
                        import logging
                        logger = logging.getLogger(__name__)
                        logger.error(f"Error auto-creating customize product: {create_error}")
                        return safe_json({
                            'success': False,
                            'error': f'Error creating customize service product for "{customize_type}": {str(create_error)}'
                        })
                
                # Always fetch the actual product price from database
                product.refresh_from_db()
                
                # Get actual unit price from product database
                # Use product.price if available, otherwise fall back to frontend cost
                quantity = 1
                actual_unit_price = product.price if product.price and product.price > 0 else Decimal(str(data.get('total_cost', 0)))
                actual_total_price = actual_unit_price * quantity
                
                # For customize orders, add any additional costs from frontend if product base price is being used
                frontend_cost = Decimal(str(data.get('total_cost', 0)))
                if frontend_cost > actual_total_price:
                    # Add the difference as additional cost (e.g., fabric costs)
                    additional_cost = frontend_cost - actual_total_price
                    actual_total_price += additional_cost
                
                OrderItem.objects.create(
                    order=order,
                    product=product,
                    quantity=quantity,
                    unit_price=actual_unit_price,
                    total_price=actual_total_price
                )
            
            # Recalculate order total based on actual item prices from database
            from django.db.models import Sum
            actual_total = order.items.aggregate(total=Sum('total_price'))['total'] or Decimal('0')
            
            # Update order with actual total amount
            order.total_amount = actual_total
            order.paid_amount = actual_total  # For now, assume full payment (can be adjusted later)
            order.balance = Decimal('0')
            order.save()
            
            return safe_json({
                'success': True,
                'order_id': order.id,
                'order_identifier': order.order_identifier,
                'message': f'Order {order.order_identifier} created successfully'
            })
            
        except Exception as e:
            import traceback

            # Get error message, ensuring no Decimal values
            try:
                error_msg = str(e)
                # Test if error message can be serialized
                test_dict = {'error': error_msg}
                json.dumps(test_dict, default=str)
            except Exception as serialization_error:
                # If serialization fails, use a safe error message
                error_msg = f"Error occurred: {type(e).__name__}"
            return safe_json({
                'success': False,
                'error': error_msg
            })
    
    return safe_json({'success': False, 'error': 'Invalid request method'})


@login_required
def add_product(request):
    if request.method == 'POST':
        form = ProductForm(request.POST, request.FILES)
        if form.is_valid():
            product = form.save()
            messages.success(request, f'{product.name} added to inventory.')
            return redirect('inventory')
    else:
        form = ProductForm()
    
    return render(request, 'business/add_product.html', {'form': form})


@login_required
def sales_page(request):
    """Sales page with real-time database connections and automatic calculations"""
    from decimal import Decimal

    # Use database connection function to calculate actual sales
    sales_calculation = calculate_actual_sales()
    
    # Only show sales from completed orders
    completed_sales = Sales.objects.filter(order__status='completed')
    
    # Sales statistics with real-time data
    total_sales = completed_sales.aggregate(total=Sum('amount'))['total'] or Decimal('0')
    monthly_sales = completed_sales.filter(
        created_at__gte=timezone.now().replace(day=1)
    ).aggregate(total=Sum('amount'))['total'] or Decimal('0')
    
    # Calculate yesterday's sales
    yesterday = timezone.now() - timedelta(days=1)
    yesterday_start = yesterday.replace(hour=0, minute=0, second=0, microsecond=0)
    yesterday_end = yesterday.replace(hour=23, minute=59, second=59, microsecond=999999)
    yesterday_sales = completed_sales.filter(
        created_at__gte=yesterday_start,
        created_at__lte=yesterday_end
    ).aggregate(total=Sum('amount'))['total'] or Decimal('0')
    
    # Calculate this week's sales (last 7 days)
    seven_days_ago = timezone.now() - timedelta(days=7)
    week_sales = completed_sales.filter(
        created_at__gte=seven_days_ago
    ).aggregate(total=Sum('amount'))['total'] or Decimal('0')
    
    # Calculate projected month (based on current day of month)
    today = timezone.now()
    day_of_month = today.day
    days_in_month = 30  # Approximate
    if day_of_month > 0:
        projected_month = (monthly_sales / day_of_month) * days_in_month
    else:
        projected_month = Decimal('0')
    
    projected_percentage = (monthly_sales / projected_month * 100) if projected_month > 0 else Decimal('0')
    
    # Total transactions
    total_transactions = completed_sales.count()
    completed_transactions = completed_sales.count()
    
    # Sales by payment method
    sales_by_method = completed_sales.values('payment_method').annotate(
        total=Sum('amount'),
        count=Count('id')
    ).order_by('-total')
    
    total_by_method = sum(method['total'] for method in sales_by_method) or Decimal('0')
    
    # Recent sales with pagination
    recent_sales_qs = completed_sales.select_related('order__customer').order_by('-created_at')
    
    # Pagination for recent sales
    page = request.GET.get('page', 1)
    paginator = Paginator(recent_sales_qs, 10)
    try:
        recent_sales = paginator.page(page)
    except PageNotAnInteger:
        recent_sales = paginator.page(1)
    except EmptyPage:
        recent_sales = paginator.page(paginator.num_pages)
    
    # Calculate average sale
    average_sale = Decimal('0')
    if total_transactions > 0:
        average_sale = total_sales / total_transactions
    
    # Get sales trends (last 7 days)
    daily_sales = completed_sales.filter(
        created_at__gte=seven_days_ago
    ).annotate(
        day=TruncDate('created_at')
    ).values('day').annotate(
        daily_total=Sum('amount')
    ).order_by('day')
    
    # Sales by service type
    sales_by_service_type = completed_sales.values('order__order_type').annotate(
        total_revenue=Sum('amount'),
        count=Count('id')
    ).order_by('-total_revenue')
    
    # Calculate average order value per service type
    for service in sales_by_service_type:
        if service['count'] > 0:
            service['avg_order_value'] = service['total_revenue'] / service['count']
        else:
            service['avg_order_value'] = Decimal('0')
        service['type'] = service['order__order_type']
    
    # Service type data for chart
    service_type_data = [
        {
            'type': service['type'].title(),
            'revenue': float(service['total_revenue'])
        }
        for service in sales_by_service_type
    ]
    
    # Monthly costs (estimate based on orders)
    monthly_orders = Order.objects.filter(
        status='completed',
        created_at__gte=timezone.now().replace(day=1)
    )
    
    # Estimate material costs (30% of revenue as estimate)
    estimated_material_cost = monthly_sales * Decimal('0.3')
    monthly_costs = estimated_material_cost  # Simplified - can be enhanced with actual cost tracking
    monthly_profit = monthly_sales - monthly_costs
    
    # Accounting metrics
    # Cash on hand = total from completed sales (since Sales represents payments received)
    cash_on_hand = completed_sales.aggregate(total=Sum('amount'))['total'] or Decimal('0')
    
    # Accounts receivable = completed orders total - sales (paid amounts)
    completed_orders_total = Order.objects.filter(status='completed').aggregate(
        total=Sum('total_amount')
    )['total'] or Decimal('0')
    accounts_receivable = completed_orders_total - cash_on_hand
    if accounts_receivable < 0:
        accounts_receivable = Decimal('0')
    
    # Total profit = total sales - total costs (estimated)
    total_costs = total_sales * Decimal('0.3')  # 30% cost estimate
    total_profit = total_sales - total_costs
    profit_margin = (total_profit / total_sales * 100) if total_sales > 0 else Decimal('0')
    
    # Top 5 most profitable orders
    top_profitable_orders = []
    for order in Order.objects.filter(status='completed').select_related('customer').order_by('-total_amount')[:5]:
        order_revenue = order.total_amount
        order_cost = order_revenue * Decimal('0.3')  # Estimate
        order_profit = order_revenue - order_cost
        order_margin = (order_profit / order_revenue * 100) if order_revenue > 0 else Decimal('0')
        
        top_profitable_orders.append({
            'order': order,
            'revenue': order_revenue,
            'cost': order_cost,
            'profit': order_profit,
            'margin': order_margin
        })
    
    # Get top selling products/services
    top_products = OrderItem.objects.filter(
        order__status='completed'
    ).values('product__name').annotate(
        total_sold=Sum('quantity'),
        total_revenue=Sum('total_price')
    ).order_by('-total_revenue')[:5]
    
    context = {
        'total_sales': total_sales,
        'monthly_sales': monthly_sales,
        'yesterday_sales': yesterday_sales,
        'week_sales': week_sales,
        'projected_month': projected_month,
        'projected_percentage': projected_percentage,
        'total_transactions': total_transactions,
        'completed_transactions': completed_transactions,
        'sales_by_method': sales_by_method,
        'total_by_method': total_by_method,
        'recent_sales': recent_sales,
        'average_sale': average_sale,
        'daily_sales': json.dumps([{'day': str(item['day']), 'daily_total': float(item['daily_total'])} for item in daily_sales]),
        'top_products': top_products,
        'sales_calculation': sales_calculation,
        'new_sales_created': sales_calculation['new_sales_created'],
        'cash_on_hand': cash_on_hand,
        'accounts_receivable': accounts_receivable,
        'total_profit': total_profit,
        'profit_margin': profit_margin,
        'total_costs': total_costs,
        'sales_by_service_type': sales_by_service_type,
        'service_type_data': json.dumps(service_type_data),
        'monthly_costs': monthly_costs,
        'monthly_profit': monthly_profit,
        'estimated_material_cost': estimated_material_cost,
        'top_profitable_orders': top_profitable_orders,
    }
    
    return render(request, 'business/sales.html', context)


@login_required
@csrf_protect
def track_order(request):
    """Enhanced track order view with QR code support and better error handling"""
    if request.method == 'POST':
        order_id = request.POST.get('order_id', '').strip()
        
        if not order_id:
            messages.error(request, 'Please enter an Order ID')
            return render(request, 'business/track_order.html')
        
        # Try multiple search methods (case-insensitive, includes archived orders)
        order = None
        
        # Normalize the search input - handle variations in format
        # Convert "TS01REP-037" to also try "TS01REP-O37" and "TS01REP-O037"
        search_variants = [order_id]
        
        # If the format is like "TS01REP-037" (without O), also try "TS01REP-O37" and "TS01REP-O037"
        if '-' in order_id:
            parts = order_id.split('-')
            if len(parts) == 2:
                prefix = parts[0]
                number_part = parts[1]
                
                # Remove leading zeros from number part for matching
                number_without_leading_zeros = str(int(number_part)) if number_part.isdigit() else number_part
                
                # If the format doesn't have "O" (like "TS01REP-037"), try with "O"
                if not 'O' in parts[1].upper() and number_part.isdigit():
                    # Try with O prefix and original number: TS01REP-O037
                    search_variants.append(f"{prefix}-O{number_part}")
                    # Try with O prefix and without leading zeros: TS01REP-O37
                    search_variants.append(f"{prefix}-O{number_without_leading_zeros}")
                    # Try with O prefix and 2-digit format: TS01REP-O37 (padded)
                    if len(number_without_leading_zeros) == 1:
                        search_variants.append(f"{prefix}-O0{number_without_leading_zeros}")
                    elif len(number_without_leading_zeros) == 2:
                        search_variants.append(f"{prefix}-O{number_without_leading_zeros}")
                
                # If the format has "O" (like "TS01REP-O37"), also try without "O"
                elif 'O' in parts[1].upper() and number_part.replace('O', '').replace('o', '').isdigit():
                    number_only = number_part.replace('O', '').replace('o', '')
                    # Try without O: TS01REP-37
                    search_variants.append(f"{prefix}-{number_only}")
                    # Try without O with leading zeros: TS01REP-037
                    if len(number_only) == 2:
                        search_variants.append(f"{prefix}-0{number_only}")
                    elif len(number_only) == 1:
                        search_variants.append(f"{prefix}-00{number_only}")
        
        # Method 1: Try exact match on order_identifier (case-sensitive first for speed)
        for variant in search_variants:
            try:
                order = Order.objects.get(order_identifier=variant)
                break
            except Order.DoesNotExist:
                continue
        
        if not order:
            # Method 2: Try case-insensitive match on order_identifier
            for variant in search_variants:
                try:
                    order = Order.objects.get(order_identifier__iexact=variant)
                    break
                except Order.DoesNotExist:
                    continue
        
        if not order:
            # Method 3: Try order_id field (UUID) - exact match
            try:
                order = Order.objects.get(order_id=order_id)
            except (Order.DoesNotExist, ValidationError):
                # Method 4: Try by numeric ID if it's a number
                try:
                    if order_id.isdigit():
                        order = Order.objects.get(id=int(order_id))
                    else:
                        # Method 5: Try case-insensitive contains search on order_identifier
                        # This handles partial matches like "TS01REP" matching "TS01REP-O37"
                        orders = Order.objects.filter(order_identifier__icontains=order_id)
                        if orders.exists():
                            # If multiple matches, prefer exact or closest match
                            # Try all variants for exact match first
                            for variant in search_variants:
                                exact_match = orders.filter(order_identifier__iexact=variant).first()
                                if exact_match:
                                    order = exact_match
                                    break
                            
                            # If no exact match, try starts with
                            if not order:
                                for variant in search_variants:
                                    starts_match = orders.filter(order_identifier__istartswith=variant).first()
                                    if starts_match:
                                        order = starts_match
                                        break
                            
                            # If still no match, get the first match (most recent)
                            if not order:
                                order = orders.order_by('-created_at').first()
                        else:
                            raise Order.DoesNotExist
                except (Order.DoesNotExist, ValueError):
                    messages.error(request, f'Order "{order_id}" not found. Please check your Order ID and try again.')
                    return render(request, 'business/track_order.html', {'order_id': order_id})
        
        if not order:
            messages.error(request, f'Order "{order_id}" not found. Please check your Order ID and try again.')
            return render(request, 'business/track_order.html', {'order_id': order_id})
        
        # Check if this is from QR scan (via POST means it's from QR/form submission)
        from_qr = request.POST.get('from_qr', False) or request.GET.get('from_qr', False)
        
        # Load order with all related data for detailed display
        order = Order.objects.select_related('customer').prefetch_related('items__product', 'items__product__category').get(id=order.id)
        
        # Add to recent searches (you could implement this with a model)
        return render(request, 'business/track_result.html', {
            'order': order, 
            'found': True,
            'order_id': order_id,
            'from_qr': from_qr
        })
    
    return render(request, 'business/track_order.html')

@login_required
def track_result(request):
    """Display track order result"""
    order_id = request.GET.get('order_id')
    if order_id:
        try:
            # Try to find by ID (primary key) first
            try:
                order = Order.objects.get(id=order_id)
            except (Order.DoesNotExist, ValueError):
                # Try to find by order_identifier
                try:
                    order = Order.objects.get(order_identifier=order_id)
                except Order.DoesNotExist:
                    # Try to find by order_id (UUID)
                    try:
                        order = Order.objects.get(order_id=order_id)
                    except Order.DoesNotExist:
                        raise Order.DoesNotExist
            
            # Check if this is from QR scan
            from_qr = request.GET.get('from_qr', False)
            
            # Load order with all related data for detailed display
            order = Order.objects.select_related('customer').prefetch_related('items__product', 'items__product__category').get(id=order.id)
            
            return render(request, 'business/track_result.html', {
                'order': order, 
                'found': True,
                'order_id': order_id,
                'from_qr': from_qr
            })
        except Order.DoesNotExist:
            return render(request, 'business/track_result.html', {
                'found': False, 
                'order_id': order_id
            })
    
    return render(request, 'business/track_result.html', {'found': False})


def generate_qr_code(order):
    """Generate QR code with complete order details"""
    import json

    from django.utils import timezone

    # Ensure balance is calculated
    if not hasattr(order, 'balance') or order.balance is None:
        order.balance = order.total_amount - order.paid_amount
    
    # Prepare order details for QR code (simplified for easier scanning)
    qr_data = {
        'id': str(order.order_identifier) if order.order_identifier else str(order.order_id),  # Shortened key
        'type': order.get_order_type_display() if hasattr(order, 'get_order_type_display') else str(order.order_type),  # Shortened key
        'customer': order.customer.name if order.customer else 'N/A',  # Shortened key
        'date': order.created_at.strftime('%Y-%m-%d') if order.created_at else timezone.now().strftime('%Y-%m-%d'),  # Shortened key and format
        'balance': float(order.balance)  # Keep balance as requested
    }
    
    # Convert to JSON string
    qr_json = json.dumps(qr_data)
    
    # Generate QR code with order details
    # Use auto version to handle any amount of data
    # Use higher error correction for better scanning reliability
    qr = qrcode.QRCode(
        version=None,  # Auto-select version based on data size
        box_size=12,  # Slightly larger for better scanning
        border=4,  # Standard border
        error_correction=qrcode.constants.ERROR_CORRECT_H  # High error correction for reliability
    )
    qr.add_data(qr_json)
    qr.make(fit=True)
    
    # Create image with higher quality
    img = qr.make_image(fill_color="black", back_color="white")
    
    # Save to file
    buffer = io.BytesIO()
    img.save(buffer, format='PNG')
    buffer.seek(0)
    
    qr_code_obj, created = QRCode.objects.get_or_create(order=order)
    qr_code_obj.qr_code_image.save(
        f'qr_{order.order_id}.png',
        ContentFile(buffer.getvalue()),
        save=True
    )
    
    return qr_json  # Return the JSON data for debugging/verification


def send_sms_notification(order):
    # This is a placeholder for SMS functionality
    # In a real implementation, you would integrate with an SMS service
    SMSNotification.objects.create(
        order=order,
        phone_number=order.customer.phone,
        message=f"Your order {order.order_id} has been confirmed. Total: ₱{order.total_amount}",
        status='sent'
    )


@login_required
def generate_pdf_report(request, report_type):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{report_type}_report.pdf"'
    
    p = canvas.Canvas(response, pagesize=letter)
    p.drawString(100, 750, f"{report_type.title()} Report")
    p.drawString(100, 730, f"Generated on: {timezone.now().strftime('%Y-%m-%d %H:%M')}")
    
    if report_type == 'sales':
        sales = Sales.objects.all()
        y = 700
        for sale in sales:
            p.drawString(100, y, f"Order: {sale.order.order_id} - Amount: ₱{sale.amount}")
            y -= 20
    
    p.showPage()
    p.save()
    return response


@login_required
def generate_excel_report(request, report_type):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{report_type}_report.xlsx"'
    
    wb = Workbook()
    ws = wb.active
    ws.title = f"{report_type.title()} Report"
    
    if report_type == 'sales':
        ws['A1'] = 'Order ID'
        ws['B1'] = 'Customer'
        ws['C1'] = 'Amount'
        ws['D1'] = 'Date'
        
        row = 2
        for sale in Sales.objects.all():
            ws[f'A{row}'] = str(sale.order.order_id)
            ws[f'B{row}'] = sale.order.customer.name
            ws[f'C{row}'] = float(sale.amount)
            ws[f'D{row}'] = sale.created_at.strftime('%Y-%m-%d')
            row += 1
    
    wb.save(response)
    return response


def normalize_phone_number(raw_phone: str) -> str:
    """Convert common phone formats to E.164 (+countrycode) when possible."""
    if not raw_phone:
        return ''
    
    cleaned = ''.join(ch for ch in raw_phone.strip() if ch.isdigit() or ch == '+')
    if not cleaned:
        return ''
    
    if cleaned.startswith('+'):
        return cleaned
    
    if cleaned.startswith('00'):
        return '+' + cleaned[2:]
    
    if cleaned.startswith('63'):
        return '+' + cleaned
    
    if cleaned.startswith('0') and len(cleaned) in (10, 11):
        # Assume PH local format (09XXXXXXXXX)
        return '+63' + cleaned.lstrip('0')
    
    if not cleaned.startswith('+'):
        return '+' + cleaned
    
    return cleaned


# ==================== CUSTOMER MANAGEMENT VIEWS ====================

@login_required
def customer_list(request):
    """Customer list with search and filtering"""
    search_query = request.GET.get('search', '')
    customers = Customer.objects.all()
    
    if search_query:
        customers = customers.filter(
            Q(name__icontains=search_query) |
            Q(phone__icontains=search_query) |
            Q(email__icontains=search_query)
        )
    
    # Get customer statistics
    total_customers = customers.count()
    customers_with_orders = customers.filter(order__isnull=False).distinct().count()
    customers_without_orders = total_customers - customers_with_orders
    
    # Recent customers (last 30 days)
    thirty_days_ago = timezone.now() - timedelta(days=30)
    recent_customers = customers.filter(created_at__gte=thirty_days_ago).count()
    
    # Order customers for display
    customers = customers.order_by('-created_at')
    
    # Pagination for customers
    paginator = Paginator(customers, 10)  # Show 10 customers per page
    page = request.GET.get('page')
    try:
        customers = paginator.page(page)
    except PageNotAnInteger:
        customers = paginator.page(1)
    except EmptyPage:
        customers = paginator.page(paginator.num_pages)
    
    context = {
        'customers': customers,
        'search_query': search_query,
        'total_customers': total_customers,
        'customers_with_orders': customers_with_orders,
        'customers_without_orders': customers_without_orders,
        'recent_customers': recent_customers,
    }
    
    return render(request, 'business/customer_list.html', context)


@login_required
def add_customer(request):
    """Add new customer"""
    if request.method == 'POST':
        form = CustomerForm(request.POST)
        if form.is_valid():
            customer = form.save()
            messages.success(request, f'Customer "{customer.name}" added successfully.')
            return redirect('customer_list')
    else:
        form = CustomerForm()
    
    return render(request, 'business/add_customer.html', {'form': form})


@login_required
def edit_customer(request, customer_id):
    """Edit existing customer"""
    customer = get_object_or_404(Customer, id=customer_id)
    
    if request.method == 'POST':
        form = CustomerForm(request.POST, instance=customer)
        if form.is_valid():
            customer = form.save()
            messages.success(request, f'Customer "{customer.name}" updated successfully.')
            return redirect('customer_detail', customer_id=customer.id)
    else:
        form = CustomerForm(instance=customer)
    
    return render(request, 'business/edit_customer.html', {'form': form, 'customer': customer})


@login_required
def delete_customer(request, customer_id):
    """Delete customer"""
    customer = get_object_or_404(Customer, id=customer_id)
    
    if request.method == 'POST':
        customer_name = customer.name
        customer.delete()
        messages.success(request, f'Customer "{customer_name}" deleted successfully.')
        return redirect('customer_list')
    
    return render(request, 'business/delete_customer.html', {'customer': customer})


@login_required
def customer_detail(request, customer_id):
    """Customer detail view with order history and loyalty status"""
    customer = get_object_or_404(Customer, id=customer_id)
    
    # Get customer's orders
    orders = Order.objects.filter(customer=customer).order_by('-created_at')
    
    # Calculate customer statistics
    total_orders = orders.count()
    total_spent = orders.aggregate(total=Sum('total_amount'))['total'] or 0
    completed_orders = orders.filter(status='completed').count()
    pending_orders = orders.filter(status='pending').count()
    
    # Check for loyalty customer status (10+ orders in last 2 months)
    two_months_ago = timezone.now() - timedelta(days=60)
    recent_orders_count = orders.filter(
        created_at__gte=two_months_ago,
        status='completed'
    ).count()
    
    is_loyal_customer = recent_orders_count >= 10
    
    # Get detailed purchase history with order items
    purchase_history = []
    for order in orders:
        order_items = OrderItem.objects.filter(order=order)
        purchase_history.append({
            'order': order,
            'items': order_items,
            'total': order.total_amount,
            'date': order.created_at,
            'status': order.status,
            'order_type': order.order_type
        })
    
    # Get order type breakdown
    order_types = orders.values('order_type').annotate(count=Count('id')).order_by('-count')
    
    # Get monthly spending breakdown for last 6 months
    six_months_ago = timezone.now() - timedelta(days=180)
    monthly_spending = orders.filter(
        created_at__gte=six_months_ago,
        status='completed'
    ).annotate(
        month=TruncDate('created_at')
    ).values('month').annotate(
        total=Sum('total_amount'),
        count=Count('id')
    ).order_by('month')
    
    context = {
        'customer': customer,
        'orders': orders,
        'purchase_history': purchase_history,
        'total_orders': total_orders,
        'total_spent': total_spent,
        'completed_orders': completed_orders,
        'pending_orders': pending_orders,
        'order_types': order_types,
        'is_loyal_customer': is_loyal_customer,
        'recent_orders_count': recent_orders_count,
        'monthly_spending': monthly_spending,
    }
    
    return render(request, 'business/customer_detail.html', context)


# ==================== CUSTOMER API VIEWS ====================

@login_required
def api_customers_list(request):
    """API endpoint to get list of customers"""
    if request.method == 'GET':
        try:
            customers = Customer.objects.all().order_by('name')
            customers_data = []
            
            for customer in customers:
                try:
                    created_at = customer.created_at.isoformat() if customer.created_at else None
                    order_count = customer.order_set.count() if hasattr(customer, 'order_set') else 0
                    
                    customers_data.append({
                        'id': customer.id,
                        'name': customer.name or '',
                        'phone': customer.phone or '',
                        'email': customer.email or '',
                        'address': customer.address or '',
                        'created_at': created_at,
                        'order_count': order_count
                    })
                except Exception as e:
                    # Skip problematic customer and continue
                    import logging
                    logger = logging.getLogger(__name__)
                    logger.warning(f"Error processing customer {customer.id}: {str(e)}")
                    continue
            
            return JsonResponse({
                'success': True,
                'customers': customers_data
            })
        except Exception as e:
            import traceback
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Error in api_customers_list: {str(e)}")
            traceback.print_exc()
            return JsonResponse({
                'success': False,
                'error': f'Error loading customers: {str(e)}',
                'customers': []
            }, status=500)
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})


@login_required
def api_customer_detail(request, customer_id):
    """API endpoint to get customer details"""
    if request.method == 'GET':
        try:
            customer = Customer.objects.get(id=customer_id)
            customer_data = {
                'id': customer.id,
                'name': customer.name,
                'phone': customer.phone,
                'email': customer.email,
                'address': customer.address,
                'created_at': customer.created_at.isoformat(),
                'order_count': customer.order_set.count()
            }
            
            return JsonResponse({
                'success': True,
                'customer': customer_data
            })
            
        except Customer.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Customer not found'
            })
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})


# ==================== PRODUCT API VIEWS ====================

@login_required
def api_products_list(request):
    """API endpoint to get list of rental products (excludes customize/service products)"""
    if request.method == 'GET':
        try:
            from .models import OrderItem
            from .rental_manager import RentalStatusManager

            # Only return rental products, exclude service/customize products
            products = Product.objects.filter(
                product_type='rental',
                is_active=True,
                is_archived=False
            ).select_related('category').order_by('name')
            
            # DIRECT APPROACH: Get all products that are in active rental orders
            # This is the source of truth - check OrderItems directly
            from .models import Order
            active_rental_orders = Order.objects.filter(
                order_type__in=['rent', 'rental'],
                status__in=['rented', 'pending', 'almost_due', 'due', 'overdue', 'in_progress']
            ).prefetch_related('items__product')
            
            # Build a set of product IDs that are actually rented (from active orders)
            actually_rented_product_ids = set()
            for order in active_rental_orders:
                for item in order.items.filter(product__product_type='rental'):
                    actually_rented_product_ids.add(item.product.id)
            
            # Get accurate rental status for all products (for additional info)
            rental_status_data = RentalStatusManager.get_rental_status_for_all_products()
            
            products_data = []
            
            for product in products:
                # Check if product is in active rental orders (DIRECT CHECK - most reliable)
                is_actually_rented = product.id in actually_rented_product_ids
                
                # FIX: If product has rental_status='rented' but is NOT in any active order,
                # it means the order was returned but product status wasn't updated
                # Mark it as available and update the database
                if product.rental_status == 'rented' and not is_actually_rented:
                    # Double-check: verify there's no active order
                    has_active_order = OrderItem.objects.filter(
                        product=product,
                        order__order_type__in=['rent', 'rental'],
                        order__status__in=['rented', 'pending', 'almost_due', 'due', 'overdue', 'in_progress']
                    ).exists()
                    
                    if not has_active_order:
                        # No active order found, fix the product status
                        product.rental_status = 'available'
                        product.current_rental_order = None
                        product.rental_start_date = None
                        product.rental_due_date = None
                        product.save()
                
                # Get additional status info from RentalStatusManager
                product_status = rental_status_data.get(product.id, {})
                accurate_rental_status = product_status.get('rental_status', product.rental_status or 'available')
                
                # Use the direct check result - if it's in an active order, it's rented
                if is_actually_rented:
                    accurate_rental_status = 'rented'
                    is_rented = True
                else:
                    # Fall back to RentalStatusManager result
                    is_rented = product_status.get('is_rented', False) or product.rental_status == 'rented'
                    if not is_rented:
                        accurate_rental_status = product.rental_status or 'available'
                
                # Check for overdue or almost due status
                is_overdue = product_status.get('is_overdue', False) or product.is_overdue
                is_almost_due = product.is_almost_due if hasattr(product, 'is_almost_due') else False
                
                # If almost due or overdue, update status
                if is_overdue or is_almost_due:
                    accurate_rental_status = 'overdue'
                
                # Determine if product is available
                # For rental items, quantity doesn't matter - they're unique items
                # Availability is based on rental_status only
                if product.product_type == 'rental':
                    is_available = accurate_rental_status == 'available'
                else:
                    is_available = product.quantity > 0 and accurate_rental_status == 'available'
                
                products_data.append({
                    'id': product.id,
                    'name': product.name,
                    'price': float(product.price) if product.price else 0.0,
                    'image': product.image.url if product.image else None,
                    'image_url': product.image.url if product.image else None,
                    'product_type': product.product_type,
                    'quantity': product.quantity,
                    'is_available': is_available,
                    'rental_status': accurate_rental_status,
                    'is_rented': is_rented,
                    'is_overdue': is_overdue or is_almost_due,
                    'is_almost_due': is_almost_due,
                    'category': product.category.name if product.category else 'Uncategorized'
                })
            
            return JsonResponse({
                'success': True,
                'products': products_data
            })
        except Exception as e:
            import traceback
            traceback.print_exc()
            # Fallback to basic product data if RentalStatusManager fails
            products = Product.objects.filter(
                product_type='rental',
                is_active=True,
                is_archived=False
            ).select_related('category').order_by('name')
            
            products_data = []
            for product in products:
                # FIX: For rental items, check if they're actually in active orders
                is_actually_rented = False
                if product.rental_status == 'rented':
                    # Check if product is in an active order
                    from .models import Order, OrderItem
                    has_active_order = OrderItem.objects.filter(
                        product=product,
                        order__order_type__in=['rent', 'rental'],
                        order__status__in=['rented', 'pending', 'almost_due', 'due', 'overdue', 'in_progress']
                    ).exists()
                    
                    if not has_active_order:
                        # No active order, fix the status
                        product.rental_status = 'available'
                        product.current_rental_order = None
                        product.rental_start_date = None
                        product.rental_due_date = None
                        product.save()
                    else:
                        is_actually_rented = True
                
                # For rental items, availability is based on rental_status, not quantity
                if product.product_type == 'rental':
                    is_available = product.rental_status == 'available' and not is_actually_rented
                else:
                    is_available = product.quantity > 0 and product.rental_status == 'available'
                
                products_data.append({
                    'id': product.id,
                    'name': product.name,
                    'price': float(product.price) if product.price else 0.0,
                    'image': product.image.url if product.image else None,
                    'image_url': product.image.url if product.image else None,
                    'product_type': product.product_type,
                    'quantity': product.quantity,
                    'is_available': is_available,
                    'rental_status': product.rental_status or 'available',
                    'is_rented': is_actually_rented,
                    'category': product.category.name if product.category else 'Uncategorized'
                })
            
            return JsonResponse({
                'success': True,
                'products': products_data
            })
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})


@login_required
def api_zippers_list(request):
    """API endpoint to get all zippers from materials for repair brochure"""
    try:
        from .models import MaterialType

        # Get zipper material type
        zipper_type = MaterialType.objects.filter(name__icontains='zipper').first()
        
        # Get all active zipper materials
        if zipper_type:
            zippers = Product.objects.filter(
                product_type='material',
                material_type=zipper_type,
                is_archived=False,
                is_active=True
            ).select_related('material_type')
        else:
            # Fallback: search by keywords
            zippers = Product.objects.filter(
                product_type='material',
                is_archived=False,
                is_active=True
            ).select_related('material_type')
            
            zipper_list = []
            for product in zippers:
                searchable_text = ' '.join([
                    product.name.lower() if product.name else '',
                    product.description.lower() if product.description else '',
                    product.material_type.name.lower() if product.material_type else ''
                ])
                if 'zipper' in searchable_text or 'zippers' in searchable_text:
                    zipper_list.append(product.id)
            
            zippers = Product.objects.filter(id__in=zipper_list)
        
        zippers_list = []
        
        for zipper in zippers:
            try:
                # Convert quantity to inches if needed (zippers are stored in inches)
                quantity = zipper.quantity if zipper.unit_of_measurement == 'inches' else zipper.quantity
                
                # Safely get image URL - handle cases where image field exists but file is missing
                image_url = None
                if zipper.image:
                    try:
                        image_url = zipper.image.url
                    except Exception:
                        # Image field exists but file is missing or inaccessible
                        image_url = None
                
                zippers_list.append({
                    'id': zipper.id,
                    'name': zipper.name or '',
                    'description': zipper.description or '',
                    'price': float(zipper.price) if zipper.price else 0.0,
                    'quantity': quantity,  # Available inches
                    'unit_of_measurement': zipper.unit_of_measurement or 'inches',
                    'image': image_url,
                    'is_available': quantity > 0,
                    'cost': float(zipper.cost) if zipper.cost else 0.0
                })
            except Exception as item_error:
                # Skip problematic zipper and continue
                import logging
                logger = logging.getLogger(__name__)
                logger.warning(f"Error processing zipper {zipper.id}: {str(item_error)}")
                continue
        
        return JsonResponse({
            'success': True,
            'zippers': zippers_list,
            'count': len(zippers_list)
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'error': str(e),
            'zippers': [],
            'count': 0
        })


@login_required
def api_buttons_list(request):
    """API endpoint to get all buttons from materials for repair brochure"""
    try:
        from .models import MaterialType

        # Get buttons material type
        buttons_type = MaterialType.objects.filter(name__icontains='button').first()
        
        # Get all active button materials
        if buttons_type:
            buttons = Product.objects.filter(
                product_type='material',
                material_type=buttons_type,
                is_archived=False,
                is_active=True
            ).select_related('material_type')
        else:
            # Fallback: search by keywords
            buttons = Product.objects.filter(
                product_type='material',
                is_archived=False,
                is_active=True
            ).select_related('material_type')
            
            button_list = []
            for product in buttons:
                searchable_text = ' '.join([
                    product.name.lower() if product.name else '',
                    product.description.lower() if product.description else '',
                    product.material_type.name.lower() if product.material_type else ''
                ])
                if 'button' in searchable_text or 'buttons' in searchable_text:
                    button_list.append(product.id)
            
            buttons = Product.objects.filter(id__in=button_list)
        
        buttons_list = []
        
        for button in buttons:
            quantity = button.quantity
            unit = button.unit_of_measurement or 'pieces'
            
            # Safely get image URL - handle cases where image field exists but file is missing
            image_url = None
            if button.image:
                try:
                    image_url = button.image.url
                except Exception:
                    # Image field exists but file is missing or inaccessible
                    image_url = None
            
            buttons_list.append({
                'id': button.id,
                'name': button.name,
                'description': button.description or '',
                'price': float(button.price) if button.price else 0.0,
                'quantity': quantity,
                'unit_of_measurement': unit,
                'image': image_url,
                'is_available': quantity > 0,
                'cost': float(button.cost) if button.cost else 0.0
            })
        
        return JsonResponse({
            'success': True,
            'buttons': buttons_list,
            'count': len(buttons_list)
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'error': str(e),
            'buttons': [],
            'count': 0
        })


@login_required
def api_patches_list(request):
    """API endpoint to get all patches from materials for repair brochure"""
    try:
        from .models import MaterialType

        # Get patches material type
        patches_type = MaterialType.objects.filter(name__icontains='patch').first()
        
        # Get all active patch materials
        if patches_type:
            patches = Product.objects.filter(
                product_type='material',
                material_type=patches_type,
                is_archived=False,
                is_active=True
            ).select_related('material_type')
        else:
            # Fallback: search by keywords
            patches = Product.objects.filter(
                product_type='material',
                is_archived=False,
                is_active=True
            ).select_related('material_type')
            
            patch_list = []
            for product in patches:
                searchable_text = ' '.join([
                    product.name.lower() if product.name else '',
                    product.description.lower() if product.description else '',
                    product.material_type.name.lower() if product.material_type else ''
                ])
                if 'patch' in searchable_text or 'patches' in searchable_text:
                    patch_list.append(product.id)
            
            patches = Product.objects.filter(id__in=patch_list)
        
        patches_list = []
        
        for patch in patches:
            quantity = patch.quantity
            unit = patch.unit_of_measurement or 'pieces'
            
            # Safely get image URL - handle cases where image field exists but file is missing
            image_url = None
            if patch.image:
                try:
                    image_url = patch.image.url
                except Exception:
                    # Image field exists but file is missing or inaccessible
                    image_url = None
            
            patches_list.append({
                'id': patch.id,
                'name': patch.name,
                'description': patch.description or '',
                'price': float(patch.price) if patch.price else 0.0,
                'quantity': quantity,
                'unit_of_measurement': unit,
                'image': image_url,
                'is_available': quantity > 0,
                'cost': float(patch.cost) if patch.cost else 0.0
            })
        
        return JsonResponse({
            'success': True,
            'patches': patches_list,
            'count': len(patches_list)
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'error': str(e),
            'patches': [],
            'count': 0
        })


@login_required
def api_locks_list(request):
    """API endpoint to get all locks/kawit from materials for repair brochure"""
    try:
        from .models import MaterialType

        # Get locks/kawit material type - prioritize exact matches and variations
        # Look for "Locks/Kawit", "locks/kawit", or names containing both "lock" and "kawit"
        locks_type = MaterialType.objects.filter(
            Q(name__iexact='Locks/Kawit') |
            Q(name__iexact='locks/kawit') |
            Q(name__icontains='locks/kawit') |
            (Q(name__icontains='lock') & Q(name__icontains='kawit'))
        ).first()
        
        # If not found, try just containing lock or kawit
        if not locks_type:
            locks_type = MaterialType.objects.filter(
                Q(name__icontains='lock') | Q(name__icontains='kawit')
            ).exclude(name__iexact='locks').exclude(name__iexact='elastic').first()
        
        # Get all active lock materials
        if locks_type:
            locks = Product.objects.filter(
                product_type='material',
                material_type=locks_type,
                is_archived=False,
                is_active=True
            ).select_related('material_type').order_by('name')
        else:
            # Fallback: search by keywords in product name/description/material type
            locks = Product.objects.filter(
                product_type='material',
                is_archived=False,
                is_active=True
            ).select_related('material_type')
            
            lock_list = []
            for product in locks:
                # Check if material type contains lock/kawit (excluding "Locks" and "Elastic")
                material_type_name = product.material_type.name.lower() if product.material_type else ''
                
                # Skip "Locks" and "Elastic" material types (removed from dropdown)
                if material_type_name in ['locks', 'elastic']:
                    continue
                
                searchable_text = ' '.join([
                    product.name.lower() if product.name else '',
                    product.description.lower() if product.description else '',
                    material_type_name
                ])
                
                # Check if it's a locks/kawit product
                # Must contain both "lock" and "kawit" in material type name, OR have "locks/kawit" in name
                is_locks_kawit = False
                if material_type_name:
                    # Check if material type is "Locks/Kawit" or contains both lock and kawit
                    if 'locks/kawit' in material_type_name:
                        is_locks_kawit = True
                    elif 'lock' in material_type_name and 'kawit' in material_type_name:
                        is_locks_kawit = True
                
                # Also check product name and description
                if not is_locks_kawit:
                    if 'lock' in searchable_text and 'kawit' in searchable_text:
                        is_locks_kawit = True
                
                if is_locks_kawit:
                    lock_list.append(product.id)
            
            locks = Product.objects.filter(id__in=lock_list).order_by('name') if lock_list else Product.objects.none()
        
        locks_list = []
        
        for lock in locks:
            quantity = lock.quantity
            unit = lock.unit_of_measurement or 'group'
            
            # Safely get image URL - handle cases where image field exists but file is missing
            image_url = None
            if lock.image:
                try:
                    image_url = lock.image.url
                except Exception:
                    # Image field exists but file is missing or inaccessible
                    image_url = None
            
            locks_list.append({
                'id': lock.id,
                'name': lock.name,
                'description': lock.description or '',
                'price': float(lock.price) if lock.price else 0.0,
                'quantity': quantity,  # Available groups
                'unit_of_measurement': unit,
                'image': image_url,
                'is_available': quantity > 0,
                'cost': float(lock.cost) if lock.cost else 0.0
            })
        
        return JsonResponse({
            'success': True,
            'locks': locks_list,
            'count': len(locks_list)
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'error': str(e),
            'locks': [],
            'count': 0
        })


@login_required
def api_garters_list(request):
    """API endpoint to get all garters from materials for repair brochure"""
    try:
        from .models import MaterialType

        # Get garter material type
        garter_type = MaterialType.objects.filter(name__icontains='garter').first()
        
        # Get all active garter materials
        if garter_type:
            garters = Product.objects.filter(
                product_type='material',
                material_type=garter_type,
                is_archived=False,
                is_active=True
            ).select_related('material_type')
        else:
            # Fallback: search by keywords
            garters = Product.objects.filter(
                product_type='material',
                is_archived=False,
                is_active=True
            ).select_related('material_type')
            
            garter_list = []
            for product in garters:
                searchable_text = ' '.join([
                    product.name.lower() if product.name else '',
                    product.description.lower() if product.description else '',
                    product.material_type.name.lower() if product.material_type else ''
                ])
                if 'garter' in searchable_text or 'elastic' in searchable_text:
                    garter_list.append(product.id)
            
            garters = Product.objects.filter(id__in=garter_list)
        
        garters_list = []
        
        for garter in garters:
            # Get quantity in inches (garters use inches as unit)
            quantity = garter.quantity if garter.unit_of_measurement == 'inches' else garter.quantity
            
            # Safely get image URL - handle cases where image field exists but file is missing
            image_url = None
            if garter.image:
                try:
                    image_url = garter.image.url
                except Exception:
                    # Image field exists but file is missing or inaccessible
                    image_url = None
            
            garters_list.append({
                'id': garter.id,
                'name': garter.name,
                'description': garter.description or '',
                'price': float(garter.price) if garter.price else 0.0,
                'quantity': quantity,  # Available inches
                'unit_of_measurement': garter.unit_of_measurement or 'inches',
                'image': image_url,
                'is_available': quantity > 0,
                'cost': float(garter.cost) if garter.cost else 0.0
            })
        
        return JsonResponse({
            'success': True,
            'garters': garters_list,
            'count': len(garters_list)
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'error': str(e),
            'garters': [],
            'count': 0
        })


@login_required
def api_thread_availability(request):
    """API endpoint to check thread availability by color"""
    if request.method == 'GET':
        thread_color = request.GET.get('color', '').strip()
        
        if not thread_color:
            return JsonResponse({
                'success': False,
                'error': 'Thread color is required'
            })
        
        availability = check_thread_availability(thread_color)
        
        return JsonResponse({
            'success': True,
            'available': availability['available'],
            'quantity': availability['quantity'],
            'found': availability['found'],
            'unit': availability.get('unit', 'meters'),
            'color': thread_color
        })
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})

@login_required
def api_rental_status(request):
    """API endpoint to get rental status of all products"""
    try:
        from .rental_manager import RentalStatusManager

        # Get rental status for all products using RentalStatusManager
        rental_status_data = RentalStatusManager.get_rental_status_for_all_products()
        
        # Format response for API
        rented_product_ids = []
        rental_status_dict = {}
        
        for product_id, status_info in rental_status_data.items():
            rental_status_dict[str(product_id)] = {
                'is_rented': status_info.get('rental_status') == 'rented',
                'rental_status': status_info.get('rental_status', 'available'),
                'has_current_order': status_info.get('has_current_order', False),
                'current_order_id': status_info.get('current_order_id'),
                'rental_start_date': status_info.get('rental_start_date'),
                'rental_due_date': status_info.get('rental_due_date'),
                'is_overdue': status_info.get('is_overdue', False)
            }
            
            if status_info.get('rental_status') == 'rented':
                rented_product_ids.append(product_id)
        
        return JsonResponse({
            'success': True,
            'rental_status': rental_status_dict,
            'rented_products': rented_product_ids,
            'count': len(rented_product_ids)
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        # Fallback: return basic rental status from Product model
        try:
            rental_products = Product.objects.filter(
                product_type='rental',
                is_active=True,
                is_archived=False
            )
            
            rented_product_ids = []
            rental_status_dict = {}
            
            for product in rental_products:
                is_rented = product.rental_status == 'rented'
                rental_status_dict[str(product.id)] = {
                    'is_rented': is_rented,
                    'rental_status': product.rental_status or 'available',
                    'has_current_order': bool(product.current_rental_order),
                    'current_order_id': product.current_rental_order.id if product.current_rental_order else None,
                    'rental_start_date': product.rental_start_date.isoformat() if product.rental_start_date else None,
                    'rental_due_date': product.rental_due_date.isoformat() if product.rental_due_date else None,
                    'is_overdue': product.is_overdue
                }
                
                if is_rented:
                    rented_product_ids.append(product.id)
            
            return JsonResponse({
                'success': True,
                'rental_status': rental_status_dict,
                'rented_products': rented_product_ids,
                'count': len(rented_product_ids)
            })
        except Exception as e2:
            return JsonResponse({
                'success': False,
                'error': str(e2),
                'rental_status': {},
                'rented_products': [],
                'count': 0
            })


@login_required
def api_rental_availability_check(request):
    """API endpoint to check rental availability"""
    if request.method == 'POST':
        try:
            import json
            data = json.loads(request.body)
            product_ids = data.get('product_ids', [])
            
            if not product_ids:
                return JsonResponse({
                    'success': False,
                    'error': 'No product IDs provided'
                })
            
            available_products = []
            unavailable_products = []
            
            for product_id in product_ids:
                try:
                    product = Product.objects.get(
                        id=product_id,
                        product_type='rental',
                        is_active=True,
                        is_archived=False
                    )
                    
                    if product.rental_status == 'available' and product.quantity > 0:
                        available_products.append(product_id)
                    else:
                        unavailable_products.append({
                            'id': product_id,
                            'reason': 'rented' if product.rental_status == 'rented' else 'unavailable'
                        })
                except Product.DoesNotExist:
                    unavailable_products.append({
                        'id': product_id,
                        'reason': 'not_found'
                    })
            
            return JsonResponse({
                'success': True,
                'available': available_products,
                'unavailable': unavailable_products
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            })
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})


@login_required
def api_check_overdue_orders(request):
    """API endpoint to check for overdue rental orders"""
    try:
        from datetime import timedelta

        from django.utils import timezone
        
        now = timezone.now()
        
        # Find overdue rental orders
        overdue_orders = Order.objects.filter(
            order_type__in=['rent', 'rental'],
            status__in=['rented', 'pending'],
            due_date__lt=now
        ).select_related('customer')
        
        overdue_list = []
        for order in overdue_orders:
            days_overdue = (now - order.due_date).days
            overdue_list.append({
                'order_id': order.id,
                'order_identifier': order.order_identifier,
                'customer_name': order.customer.name if order.customer else 'Unknown',
                'due_date': order.due_date.isoformat() if order.due_date else None,
                'days_overdue': days_overdue,
                'total_amount': float(order.total_amount) if order.total_amount else 0.0
            })
        
        return JsonResponse({
            'success': True,
            'overdue_orders': overdue_list,
            'count': len(overdue_list)
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e),
            'overdue_orders': [],
            'count': 0
        })


@login_required
def api_return_individual_items(request):
    """API endpoint to return individual rental items"""
    if request.method == 'POST':
        try:
            import json
            data = json.loads(request.body)
            product_ids = data.get('product_ids', [])
            order_id = data.get('order_id')
            
            if not product_ids or not order_id:
                return JsonResponse({
                    'success': False,
                    'error': 'Product IDs and order ID are required'
                })
            
            try:
                order = Order.objects.get(id=order_id)
            except Order.DoesNotExist:
                return JsonResponse({
                    'success': False,
                    'error': 'Order not found'
                })
            
            returned_products = []
            errors = []
            
            for product_id in product_ids:
                try:
                    product = Product.objects.get(
                        id=product_id,
                        product_type='rental',
                        current_rental_order=order
                    )
                    
                    # Return the product
                    product.rental_status = 'available'
                    product.current_rental_order = None
                    product.rental_start_date = None
                    product.rental_due_date = None
                    product.save()
                    
                    returned_products.append(product_id)
                except Product.DoesNotExist:
                    errors.append({
                        'product_id': product_id,
                        'error': 'Product not found or not part of this order'
                    })
                except Exception as e:
                    errors.append({
                        'product_id': product_id,
                        'error': str(e)
                    })
            
            return JsonResponse({
                'success': len(errors) == 0,
                'returned_products': returned_products,
                'errors': errors,
                'count': len(returned_products)
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            })
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})


@login_required
def check_order_balance(request, order_id):
    """Check the balance/remaining payment for an order"""
    try:
        order = get_object_or_404(Order, id=order_id)
        
        # Calculate total paid
        from django.db.models import Sum
        total_paid = order.payment_set.aggregate(total=Sum('amount'))['total'] or 0
        total_amount = float(order.total_amount) if order.total_amount else 0.0
        remaining_balance = total_amount - float(total_paid)
        
        return JsonResponse({
            'success': True,
            'order_id': order.id,
            'order_identifier': order.order_identifier,
            'total_amount': total_amount,
            'total_paid': float(total_paid),
            'remaining_balance': remaining_balance,
            'is_fully_paid': remaining_balance <= 0
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        })


@login_required
def add_customize_product(request):
    """Add a customize product (uniform/PE)"""
    if request.method == 'POST':
        import json

        from .models import Category

        # Get category name from POST data
        category_name = request.POST.get('category', '').strip()
        category_obj = None
        
        # Get or create category
        if category_name:
            category_obj, created = Category.objects.get_or_create(
                name=category_name,
                defaults={'description': f'Category for {category_name} products'}
            )
        
        # Validate required fields manually
        name = request.POST.get('name', '').strip()
        price = request.POST.get('price', '0')
        image = request.FILES.get('image')
        gender = request.POST.get('gender', '').strip()
        
        if not name:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or 'application/json' in request.headers.get('Accept', ''):
                return JsonResponse({'success': False, 'error': 'Product name is required.'})
            messages.error(request, 'Product name is required.')
            return redirect('inventory')
        
        if not image:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or 'application/json' in request.headers.get('Accept', ''):
                return JsonResponse({'success': False, 'error': 'Product image is required.'})
            messages.error(request, 'Product image is required.')
            return redirect('inventory')
        
        # Check for duplicate products
        import os
        from decimal import Decimal

        # Use the dedicated function to check for duplicates
        from business.customize_product_manager import ensure_no_duplicates

        from .models import Product

        # Reset file pointer to beginning before checking duplicates
        # This ensures the file is readable for duplicate checking
        if image and hasattr(image, 'file'):
            try:
                if hasattr(image.file, 'seek'):
                    image.file.seek(0)
            except (IOError, OSError, AttributeError):
                pass  # File might not support seek, that's okay
        
        # Create a temporary product object to check for duplicates
        temp_product = Product(
            name=name,
            product_type='service',
            category=category_obj,
            image=image
        )
        
        is_duplicate, existing_product = ensure_no_duplicates(temp_product)
        if is_duplicate and existing_product:
            error_msg = f'A customize product with the same image already exists: "{existing_product.name}". Please use a different image or modify the existing product.'
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or 'application/json' in request.headers.get('Accept', ''):
                return JsonResponse({
                    'success': False, 
                    'error': error_msg
                })
            messages.error(request, error_msg)
            return redirect('inventory')

        # Check for existing product with same name, category, and product type
        duplicate_check = Product.objects.filter(
            name__iexact=name,
            product_type='service',
            is_archived=False
        )
        
        # If category is provided, also check category
        if category_obj:
            duplicate_check = duplicate_check.filter(category=category_obj)
        
        # For fullset products, also check gender in description
        if category_name.lower() == 'fullset' and gender:
            # Check if any existing product has the same name, category, and gender
            for existing_product in duplicate_check:
                if existing_product.description:
                    existing_gender_match = re.search(r'Gender:\s*(\w+)', existing_product.description, re.IGNORECASE)
                    if existing_gender_match and existing_gender_match.group(1).lower() == gender.lower():
                        # Duplicate found
                        if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or 'application/json' in request.headers.get('Accept', ''):
                            return JsonResponse({
                                'success': False, 
                                'error': f'A customize product with the same name "{name}", category "{category_name}", and gender "{gender.title()}" already exists. Please use a different name or modify the existing product.'
                            })
                        messages.error(request, f'A customize product with the same name "{name}", category "{category_name}", and gender "{gender.title()}" already exists.')
                        return redirect('inventory')
        
        # Check for duplicates (excluding fullset with gender check above)
        if duplicate_check.exists():
            existing_product = duplicate_check.first()
            # If it's not a fullset or gender doesn't match, it's a duplicate
            if category_name.lower() != 'fullset' or not gender:
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or 'application/json' in request.headers.get('Accept', ''):
                    return JsonResponse({
                        'success': False, 
                        'error': f'A customize product with the name "{name}" and category "{category_name}" already exists. Please use a different name or modify the existing product.'
                    })
                messages.error(request, f'A customize product with the name "{name}" and category "{category_name}" already exists.')
                return redirect('inventory')
        
        try:
            product = Product(
                name=name,
                description=request.POST.get('description', '').strip(),
                category=category_obj,
                product_type='service',
                price=Decimal(price) if price else Decimal('0.00'),
                cost=Decimal('0.00'),
                quantity=0,
                min_quantity=0,
                image=image
            )
            product.save()
            
            # Collect measurements
            measurements = {}
            
            # Collect all measurement fields
            measurement_fields = [
                'measurement_length', 'measurement_crotch', 'measurement_waist', 
                'measurement_hips', 'measurement_thigh', 'measurement_knee', 
                'measurement_bottom', 'measurement_bust', 'measurement_shoulder',
                'measurement_sleeve', 'measurement_armhole', 'measurement_neckline',
                'measurement_hem_width'
            ]
            
            for field in measurement_fields:
                values = request.POST.getlist(field)
                selected_value = ''
                if values:
                    for raw in values:
                        if raw is None:
                            continue
                        value = raw.strip()
                        if value:
                            selected_value = value
                            break
                # Remove 'measurement_' prefix for cleaner JSON
                key = field.replace('measurement_', '')
                measurements[key] = selected_value
            
            # Get gender if provided (for fullset products)
            gender = request.POST.get('gender', '').strip()
            if gender:
                # Add gender information to description
                if product.description:
                    product.description += f"\n\nGender: {gender.title()}"
                else:
                    product.description = f"Gender: {gender.title()}"
            
            # Always save measurements JSON, even if all are empty
            # This ensures we can show all measurement fields later
            measurements_json = json.dumps(measurements, indent=2)
            if product.description:
                product.description += f"\n\nMeasurements:\n{measurements_json}"
            else:
                product.description = f"Measurements:\n{measurements_json}"
            
            # Validate gender is provided for fullset
            category_name = request.POST.get('category', '').strip()
            if category_name.lower() == 'fullset' and not gender:
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or 'application/json' in request.headers.get('Accept', ''):
                    return JsonResponse({'success': False, 'error': 'Gender is required when Category is Fullset.'})
                messages.error(request, 'Gender is required when Category is Fullset.')
                return redirect('inventory')
            
            product.save()
            
            # Return JSON response for AJAX
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or 'application/json' in request.headers.get('Accept', ''):
                return JsonResponse({'success': True, 'message': f'{product.name} added as customize product.'})
            
            messages.success(request, f'{product.name} added as customize product.')
            return redirect('inventory')
            
        except Exception as e:
            # Return JSON error for AJAX
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest' or 'application/json' in request.headers.get('Accept', ''):
                error_message = f'Error adding product: {str(e)}'
                return JsonResponse({'success': False, 'error': error_message})
            
            messages.error(request, f'Error adding product: {str(e)}')
            return redirect('inventory')
    else:
        form = ProductForm()
    
    return render(request, 'business/add_customize_product.html', {'form': form})


@login_required
def add_material_product(request):
    """Add a material product"""
    from .forms import MaterialProductForm
    
    if request.method == 'POST':
        form = MaterialProductForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                product = form.save()
                messages.success(request, f'{product.name} added to materials successfully.')
                return redirect('materials_management')
            except Exception as e:
                import logging
                logger = logging.getLogger(__name__)
                logger.error(f'Error saving material product: {str(e)}')
                messages.error(request, f'Error saving material: {str(e)}. Please check all required fields.')
        else:
            # Display form validation errors
            error_messages = []
            for field, errors in form.errors.items():
                for error in errors:
                    error_messages.append(f'{field}: {error}')
            if error_messages:
                messages.error(request, 'Please correct the following errors: ' + '; '.join(error_messages))
            else:
                messages.error(request, 'Please fill in all required fields correctly.')
    else:
        form = MaterialProductForm()
    
    return render(request, 'business/add_material_product.html', {'form': form})


@login_required
@csrf_protect
def delete_product(request, product_id):
    """Delete a product permanently"""
    product = get_object_or_404(Product, id=product_id)
    if request.method == 'POST':
        product_name = product.name
        product.delete()
        messages.success(request, f'{product_name} has been deleted permanently.')
        return redirect('inventory')
    
    return render(request, 'business/delete_product.html', {'product': product})


@login_required
def materials_management(request):
    """Materials management page"""
    # Get filter parameter - support both 'filter' and 'type' for compatibility
    material_type_filter = request.GET.get('filter') or request.GET.get('type', 'all')
    
    # Get all material products
    all_material_products = Product.objects.filter(
        product_type='material',
        is_archived=False
    ).select_related('material_type', 'material_pricing').order_by('-created_at')
    
    # Apply filter if specified
    if material_type_filter == 'all' or not material_type_filter:
        material_products = all_material_products
    elif material_type_filter == 'garter_locks':
        # Special filter for Garter & Locks together
        material_products = all_material_products.filter(
            Q(material_type__name__icontains='garter') | Q(material_type__name__icontains='lock') | Q(material_type__name__icontains='kawit')
        )
    elif material_type_filter.lower() == 'fabric':
        # Filter for Fabric
        material_products = all_material_products.filter(
            Q(material_type__name__icontains='fabric')
        )
    elif material_type_filter.lower() == 'buttons':
        # Filter for Buttons
        material_products = all_material_products.filter(
            Q(material_type__name__icontains='button')
        )
    elif material_type_filter.lower() == 'zippers':
        # Filter for Zippers
        material_products = all_material_products.filter(
            Q(material_type__name__icontains='zipper')
        )
    elif material_type_filter.lower() == 'thread':
        # Filter for Thread
        material_products = all_material_products.filter(
            Q(material_type__name__icontains='thread')
        )
    elif material_type_filter.lower() == 'patches':
        # Filter for Patches
        material_products = all_material_products.filter(
            Q(material_type__name__icontains='patch')
        )
    else:
        # Generic filter - match by material type name
        material_products = all_material_products.filter(
            material_type__name__icontains=material_type_filter
        )
    
    # Calculate statistics from all materials (unfiltered)
    from decimal import Decimal

    # Total materials count
    total_materials = all_material_products.count()
    
    # Out of stock: quantity = 0
    out_of_stock = all_material_products.filter(quantity=0).count()
    
    # Low stock: quantity > 0 AND quantity <= min_quantity AND min_quantity > 0
    low_stock_materials = all_material_products.filter(
        quantity__gt=0,
        min_quantity__gt=0
    ).filter(quantity__lte=F('min_quantity')).count()
    
    # In stock: quantity > 0 AND (min_quantity = 0 OR quantity > min_quantity)
    in_stock = all_material_products.filter(
        quantity__gt=0
    ).exclude(
        Q(min_quantity__gt=0) & Q(quantity__lte=F('min_quantity'))
    ).count()
    
    # Total value: sum of (quantity * cost) for all materials
    # Use cost if available, otherwise use price as fallback
    # Using database aggregation for better performance
    total_value_result = all_material_products.aggregate(
        total=Sum(
            Case(
                When(cost__gt=0, then=F('quantity') * F('cost')),
                When(price__gt=0, then=F('quantity') * F('price')),
                default=0,
                output_field=DecimalField(max_digits=20, decimal_places=2)
            )
        )
    )
    total_value = total_value_result['total'] or Decimal('0.00')
    
    # Get unique material types for filter dropdown
    material_types = MaterialType.objects.filter(
        products__product_type='material',
        products__is_archived=False
    ).distinct()
    
    # Pagination
    paginator = Paginator(material_products, 10)  # Show 10 materials per page
    page = request.GET.get('page')
    try:
        materials = paginator.page(page)
    except PageNotAnInteger:
        materials = paginator.page(1)
    except EmptyPage:
        materials = paginator.page(paginator.num_pages)
    
    context = {
        'materials': materials,
        'material_types': material_types,
        'current_filter': material_type_filter,
        'total_materials': total_materials,
        'in_stock': in_stock,
        'low_stock_materials': low_stock_materials,
        'out_of_stock': out_of_stock,
        'total_value': total_value,
    }
    
    return render(request, 'business/materials_management.html', context)


@login_required
def edit_material(request, product_id):
    """Edit a material product"""
    from decimal import Decimal, InvalidOperation

    from .forms import MaterialProductForm
    from .models import MaterialType
    
    product = get_object_or_404(Product, id=product_id, product_type='material')
    
    # Get all material types for the dropdown
    material_types = MaterialType.objects.exclude(name__iexact='locks').exclude(name__iexact='elastic').order_by('name')
    
    if request.method == 'POST':
        # For editing, we'll update fields directly to preserve user input
        # The form validation is still useful for basic checks
        
        # Get material type if changed
        material_type_id = request.POST.get('material_type')
        if material_type_id:
            try:
                new_material_type = MaterialType.objects.get(id=material_type_id)
                product.material_type = new_material_type
            except MaterialType.DoesNotExist:
                pass  # Keep existing material type
        
        # Update name
        if 'name' in request.POST and request.POST['name']:
            product.name = request.POST['name'].strip()
        
        # Update price (selling price)
        if 'price' in request.POST and request.POST['price']:
            try:
                product.price = Decimal(str(request.POST['price']))
            except (ValueError, TypeError, InvalidOperation):
                pass  # Keep existing value if invalid
        
        # Update cost price
        if 'cost_price' in request.POST and request.POST['cost_price']:
            try:
                product.cost = Decimal(str(request.POST['cost_price']))
            except (ValueError, TypeError, InvalidOperation):
                pass  # Keep existing value if invalid
        
        # Update quantity
        if 'quantity' in request.POST and request.POST['quantity']:
            try:
                product.quantity = Decimal(str(request.POST['quantity']))
                # Update current_quantity_in_stock to match
                product.current_quantity_in_stock = product.quantity
            except (ValueError, TypeError, InvalidOperation):
                pass  # Keep existing value if invalid
        
        # Update minimum quantity
        if 'min_quantity' in request.POST and request.POST['min_quantity']:
            try:
                product.min_quantity = int(request.POST['min_quantity'])
            except (ValueError, TypeError):
                pass  # Keep existing value if invalid
        
        # Update unit of measurement
        if 'unit_of_measurement' in request.POST and request.POST['unit_of_measurement']:
            product.unit_of_measurement = request.POST['unit_of_measurement'].strip()
        
        # Auto-set unit_of_measurement for thread materials
        if product.material_type and product.material_type.name.lower() == 'thread':
            product.unit_of_measurement = 'meters'
        
        # Update description
        if 'description' in request.POST:
            product.description = request.POST['description'].strip()
        
        # Handle image upload - only update if a new image is provided
        if 'image' in request.FILES and request.FILES['image']:
            product.image = request.FILES['image']
        
        # Validate required fields
        if not product.name:
            messages.error(request, 'Material name is required.')
        elif not product.material_type:
            messages.error(request, 'Material type is required.')
        else:
            # Save all changes
            product.save()
            
            # Log the update (note: signal will also create a log entry, but this one includes more details)
            from .models import ActivityLog
            try:
                ActivityLog.objects.create(
                    activity_type='product_updated',
                    description=f"Material '{product.name}' updated - Price: ₱{float(product.price)}, Cost: ₱{float(product.cost)}, Quantity: {float(product.quantity)} {product.unit_of_measurement}",
                    user=request.user,
                    product=product,
                    metadata={
                        'product_name': product.name,
                        'product_type': product.product_type,
                        'price': float(product.price) if product.price else 0.0,
                        'cost': float(product.cost) if product.cost else 0.0,
                        'quantity': float(product.quantity) if product.quantity else 0.0,
                        'unit_of_measurement': product.unit_of_measurement or '',
                        'is_archived': product.is_archived
                    }
                )
            except Exception as e:
                # If logging fails, don't break the save operation
                import logging
                logger = logging.getLogger(__name__)
                logger.warning(f'Failed to log material update: {str(e)}')
            
            messages.success(request, f'{product.name} has been updated successfully.')
            return redirect('materials_management')
    else:
        form = MaterialProductForm(instance=product)
    
    return render(request, 'business/edit_material.html', {
        'form': form, 
        'product': product,
        'material_types': material_types
    })


@login_required
def restock_material(request, product_id):
    """Restock a material product"""
    product = get_object_or_404(Product, id=product_id, product_type='material')
    
    if request.method == 'POST':
        restock_quantity = request.POST.get('quantity')
        try:
            restock_quantity = int(restock_quantity)
            if restock_quantity > 0:
                product.quantity += restock_quantity
                product.current_quantity_in_stock = product.quantity
                product.save()
                
                # Create inventory transaction
                InventoryTransaction.objects.create(
                    product=product,
                    transaction_type='in',
                    quantity=restock_quantity,
                    notes=f'Restocked {restock_quantity} {product.unit_of_measurement or "units"}'
                )
                
                messages.success(request, f'{restock_quantity} units added to {product.name}. New quantity: {product.quantity}')
                return redirect('materials_management')
            else:
                messages.error(request, 'Restock quantity must be greater than 0.')
        except ValueError:
            messages.error(request, 'Invalid quantity entered.')
    
    return render(request, 'business/restock_material.html', {'product': product})


@login_required
def api_customize_products_list(request):
    """API endpoint to get list of customize products"""
    if request.method == 'GET':
        # Get customize products (service type products with images, excluding repair products)
        # Only products added via "Add Customize Product" or uploaded in order form
        customize_products = Product.objects.filter(
            product_type='service',
            is_active=True,
            is_archived=False,
            image__isnull=False  # Only products with images
        ).exclude(
            name__istartswith='Repair -'  # Exclude repair products
        ).exclude(
            name__istartswith='repair -'  # Exclude repair products (case-insensitive)
        ).select_related('category').order_by('name')
        
        products_data = []
        for product in customize_products:
            products_data.append({
                'id': product.id,
                'name': product.name,
                'description': product.description or '',
                'price': float(product.price) if product.price else 0.0,
                'image_url': product.image.url if product.image else None,
                'category': product.category.name if product.category else None,
                'created_at': product.created_at.isoformat() if product.created_at else None
            })
        
        return JsonResponse({
            'success': True,
            'products': products_data,
            'count': len(products_data)
        })
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})


@login_required
def upload_customize_image_immediate(request):
    """Upload customize product image and create product immediately"""
    if request.method == 'POST':
        try:
            import hashlib
            import os
            from datetime import datetime, timedelta

            from django.core.files.base import ContentFile
            from django.core.files.storage import default_storage
            
            image_file = request.FILES.get('image')
            type_of_customize = request.POST.get('type_of_customize', 'customize')
            customize_type = request.POST.get('customize_type', '')
            
            if not image_file:
                return JsonResponse({
                    'success': False,
                    'error': 'No image file provided'
                })
            
            # Use the dedicated function to check for duplicates
            from business.customize_product_manager import ensure_no_duplicates

            # Reset file pointer to beginning before checking duplicates
            # This ensures the file is readable for duplicate checking
            if image_file and hasattr(image_file, 'file'):
                try:
                    if hasattr(image_file.file, 'seek'):
                        image_file.file.seek(0)
                except (IOError, OSError, AttributeError):
                    pass  # File might not support seek, that's okay
            
            # Create a temporary product to check for duplicates
            temp_product = Product(
                product_type='service',
                image=image_file
            )
            
            is_duplicate, existing_product = ensure_no_duplicates(temp_product)
            if is_duplicate and existing_product:
                # Return existing product instead of creating duplicate
                return JsonResponse({
                    'success': True,
                    'message': f'Product already exists: "{existing_product.name}"',
                    'product_id': existing_product.id,
                    'product_name': existing_product.name,
                    'image_url': existing_product.image.url if existing_product.image else None,
                    'duplicate': True
                })
            
            # Create product name based on type
            timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
            if customize_type:
                product_name = f"{customize_type.replace('_', ' ').title()} - {timestamp}"
            else:
                product_name = f"Customize Product - {timestamp}"
            
            # Create the product
            product = Product.objects.create(
                name=product_name,
                product_type='service',
                price=0.00,  # Default price, can be updated later
                quantity=0,
                image=image_file,
                is_active=True,
                is_archived=False
            )
            
            # Set category if available
            if type_of_customize == 'uniform' and customize_type:
                category_name = f"Uniform {customize_type.replace('_', ' ').title()}"
                category, _ = Category.objects.get_or_create(name=category_name)
                product.category = category
                product.save()
            
            return JsonResponse({
                'success': True,
                'message': f'Product "{product_name}" created successfully',
                'product_id': product.id,
                'product_name': product.name,
                'image_url': product.image.url if product.image else None
            })
        except Exception as e:
            import traceback
            traceback.print_exc()
            return JsonResponse({
                'success': False,
                'error': str(e)
            })
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})


# Material Management API endpoints
@login_required
def api_get_material_pricing_options(request):
    """API endpoint to get material pricing options"""
    if request.method == 'GET':
        try:
            from .models import MaterialPricing, MaterialType
            pricing_options = MaterialPricing.objects.select_related('material_type').all()
            
            options_data = []
            for option in pricing_options:
                options_data.append({
                    'id': option.id,
                    'material_type': option.material_type.name,
                    'pricing_type': option.pricing_type,
                    'pricing_type_display': option.get_pricing_type_display(),
                    'bundle_size': option.bundle_size,
                    'buy_price_per_unit': float(option.buy_price_per_unit),
                    'sell_price_per_unit': float(option.sell_price_per_unit),
                    'is_default': option.is_default
                })
            
            return JsonResponse({
                'success': True,
                'pricing_options': options_data,
                'count': len(options_data)
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            })
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})


@login_required
def api_material_availability_tracking(request):
    """API endpoint to track material availability"""
    if request.method == 'GET':
        try:
            materials = Product.objects.filter(
                product_type='material',
                is_archived=False
            ).select_related('material_type')
            
            tracking_data = []
            for material in materials:
                tracking_data.append({
                    'id': material.id,
                    'name': material.name,
                    'material_type': material.material_type.name if material.material_type else None,
                    'quantity': material.quantity,
                    'min_quantity': material.min_quantity,
                    'is_available': material.is_available,
                    'is_low_stock': material.is_low_stock,
                    'unit_of_measurement': material.unit_of_measurement or 'units'
                })
            
            return JsonResponse({
                'success': True,
                'materials': tracking_data,
                'count': len(tracking_data)
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            })
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})


@login_required
def api_materials_details(request):
    """API endpoint to get detailed materials information"""
    if request.method == 'GET':
        try:
            materials = Product.objects.filter(
                product_type='material',
                is_archived=False
            ).select_related('material_type', 'material_pricing')
            
            materials_data = []
            for material in materials:
                # Determine unit of measurement dynamically
                unit_of_measurement = material.unit_of_measurement or 'units'
                if material.material_type and material.material_type.name.lower() == 'garter':
                    unit_of_measurement = 'inch'
                
                materials_data.append({
                    'id': material.id,
                    'name': material.name,
                    'description': material.description or '',
                    'material_type': material.material_type.name if material.material_type else None,
                    'quantity': material.quantity,
                    'min_quantity': material.min_quantity,
                    'unit_of_measurement': unit_of_measurement,
                    'price': float(material.price) if material.price else 0.0,
                    'cost': float(material.cost) if material.cost else 0.0,
                    'image_url': material.image.url if material.image else None,
                    'is_available': material.is_available,
                    'is_low_stock': material.is_low_stock,
                    'created_at': material.created_at.isoformat() if material.created_at else None
                })
            
            return JsonResponse({
                'success': True,
                'materials': materials_data,
                'count': len(materials_data)
            })
        except Exception as e:
            import traceback
            traceback.print_exc()
            return JsonResponse({
                'success': False,
                'error': str(e)
            })
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})


@login_required
def api_material_detail(request, product_id):
    """API endpoint to get detailed information about a specific material"""
    if request.method == 'GET':
        try:
            material = get_object_or_404(Product, id=product_id, product_type='material')
            
            # Determine unit of measurement dynamically
            unit_of_measurement = material.unit_of_measurement or 'units'
            if material.material_type and material.material_type.name.lower() == 'garter':
                unit_of_measurement = 'inch'
            
            # Determine status and status_class
            status = 'UNDEFINED'
            status_class = 'status-undefined'
            
            if material.quantity == 0:
                status = 'OUT OF STOCK'
                status_class = 'status-out-of-stock'
            elif material.is_low_stock:
                status = 'LOW STOCK'
                status_class = 'status-low-stock'
            else:
                status = 'IN STOCK'
                status_class = 'status-in-stock'
            
            # Build images array (similar to api_product_detail)
            images = []
            if material.image:
                images.append({
                    'url': material.image.url,
                    'alt': material.name,
                    'is_primary': True
                })
            
            return JsonResponse({
                'success': True,
                'material': {
                    'id': material.id,
                    'name': material.name,
                    'description': material.description or '',
                    'material_type': material.material_type.name if material.material_type else None,
                    'quantity': material.quantity,
                    'min_quantity': material.min_quantity,
                    'unit_of_measurement': unit_of_measurement,
                    'price': float(material.price) if material.price else 0.0,
                    'cost': float(material.cost) if material.cost else 0.0,
                    'image_url': material.image.url if material.image else None,
                    'images': images,  # Add images array for frontend
                    'status': status,
                    'status_class': status_class,
                    'is_available': material.is_available,
                    'is_low_stock': material.is_low_stock,
                    'created_at': material.created_at.isoformat() if material.created_at else None
                }
            })
        except Exception as e:
            import traceback
            traceback.print_exc()
            return JsonResponse({
                'success': False,
                'error': str(e)
            })
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})


@login_required
def api_material_usage_history(request):
    """API endpoint to get material usage history from orders - ensures completeness"""
    if request.method == 'GET':
        try:
            from collections import defaultdict

            from django.db.models import Q

            # Get limit from query params (default 50)
            limit = int(request.GET.get('limit', 50))
            
            # Get date range from query params (optional)
            date_from = request.GET.get('date_from')
            date_to = request.GET.get('date_to')
            
            # Build base query for material transactions
            transaction_query = InventoryTransaction.objects.filter(
                transaction_type='out',
                reference_order__isnull=False,
                product__product_type='material'
            ).select_related('product', 'product__material_type', 'reference_order', 'reference_order__customer')
            
            # Apply date filters if provided
            if date_from:
                try:
                    from datetime import datetime
                    date_from_obj = datetime.strptime(date_from, '%Y-%m-%d')
                    transaction_query = transaction_query.filter(created_at__gte=date_from_obj)
                except ValueError:
                    pass  # Invalid date format, ignore
            
            if date_to:
                try:
                    from datetime import datetime, timedelta
                    date_to_obj = datetime.strptime(date_to, '%Y-%m-%d') + timedelta(days=1)
                    transaction_query = transaction_query.filter(created_at__lt=date_to_obj)
                except ValueError:
                    pass  # Invalid date format, ignore
            
            # Get material transactions that are 'out' type (deductions) with reference orders
            transactions = transaction_query.order_by('-created_at')[:limit * 2]  # Get more to ensure completeness
            
            # Group transactions by order
            order_history = defaultdict(lambda: {
                'order_id': None,
                'order_identifier': None,
                'order_type': None,
                'order_type_display': None,
                'created_at': None,
                'created_at_timestamp': None,
                'materials': []
            })
            
            for transaction in transactions:
                order = transaction.reference_order
                if not order:
                    continue
                
                order_key = order.id
                
                # Initialize order data if first time seeing this order
                if not order_history[order_key]['order_id']:
                    order_history[order_key]['order_id'] = order.id
                    order_history[order_key]['order_identifier'] = order.order_identifier
                    order_history[order_key]['order_type'] = order.order_type
                    
                    # Format order type display
                    if order.order_type == 'repair':
                        order_history[order_key]['order_type_display'] = 'Repair Order'
                    elif order.order_type == 'customize':
                        order_history[order_key]['order_type_display'] = 'Customize Order'
                    elif order.order_type == 'rent':
                        order_history[order_key]['order_type_display'] = 'Rental Order'
                    else:
                        order_history[order_key]['order_type_display'] = order.order_type.title() + ' Order'
                    
                    # Use order's created_at for proper sorting (latest orders first)
                    order_history[order_key]['created_at'] = order.created_at.isoformat()
                    order_history[order_key]['created_at_timestamp'] = order.created_at
                
                # Add material to the order's materials list
                material_name = transaction.product.name
                quantity = abs(float(transaction.quantity))  # Make positive for display
                unit = transaction.product.unit_of_measurement or 'piece'
                
                # Get material type
                material_type = ''
                if transaction.product.material_type:
                    material_type = transaction.product.material_type.name
                elif transaction.product.product_type == 'material':
                    # Try to infer from name or description if no material_type is set
                    name_lower = material_name.lower()
                    if 'thread' in name_lower:
                        material_type = 'Thread'
                    elif 'fabric' in name_lower or 'cotton' in name_lower or 'polyester' in name_lower or 'linen' in name_lower or 'denim' in name_lower:
                        material_type = 'Fabric'
                    elif 'patch' in name_lower:
                        material_type = 'Patches'
                    elif 'zipper' in name_lower:
                        material_type = 'Zipper'
                    elif 'button' in name_lower:
                        material_type = 'Buttons'
                    elif 'garter' in name_lower or 'elastic' in name_lower:
                        material_type = 'Garter'
                    elif 'lock' in name_lower or 'kawit' in name_lower:
                        material_type = 'Locks'
                
                # Extract thread color and length from transaction notes if available
                thread_color = None
                thread_length = None
                
                # Check if this is a thread material (by material_type, product name, description, or transaction notes)
                is_thread = (material_type == 'Thread' or 
                            'thread' in material_name.lower() or
                            (transaction.product.description and 'thread' in transaction.product.description.lower()) or
                            (transaction.notes and 'thread' in transaction.notes.lower()))
                
                if is_thread and transaction.notes:
                    import re
                    notes = transaction.notes
                    
                    # Try to extract thread length/meters - look for patterns like "Used X.XXm" or "X.XXm"
                    # Handle both formats: "Used 2.50m" and "2.50m (250cm)"
                    # Pattern: "Used 5.00m" or "5.00m" -> captures "5.00"
                    # \d+\.?\d* matches: 5, 5.0, 5.00, 50.00, etc.
                    thread_length_match = re.search(r'Used\s+(\d+\.?\d*)\s*m\b|(?:^|\s)(\d+\.?\d*)\s*m(?:\s|\(|$)', notes, re.IGNORECASE)
                    if thread_length_match:
                        thread_length_str = thread_length_match.group(1) or thread_length_match.group(2)
                        try:
                            thread_length = float(thread_length_str.strip())
                        except (ValueError, AttributeError):
                            thread_length = quantity
                    else:
                        # If not found in notes, use the quantity converted to meters if needed
                        if unit.lower() in ['cm', 'centimeter', 'centimeters']:
                            thread_length = quantity / 100.0  # Convert cm to meters
                        elif unit.lower() in ['meter', 'meters', 'm']:
                            thread_length = quantity
                        else:
                            thread_length = quantity
                    
                    # Try to extract thread color - look for patterns like:
                    # "Used X.XXm color thread" - format: "Used 2.50m red thread for repair/customize"
                    # Pattern examples: "Used 2.50m red thread" or "2.50m (250cm) navy blue thread"
                    # Note: Format is "Used {meters}m {color} thread for {type} - Order {id}"
                    
                    # First try: "Used X.XXm COLOR thread for" - most common format (works for both repair and customize)
                    # Pattern: "Used 5.00m brown thread for customize" -> captures "brown"
                    # Note: \d+\.?\d* matches numbers like 5, 5.0, 5.00, 50.00
                    thread_color_match = re.search(r'Used\s+\d+\.?\d*\s*m\s+([a-zA-Z]+(?:\s+[a-zA-Z]+)?)\s+thread\s+for', notes, re.IGNORECASE)
                    if thread_color_match:
                        thread_color = thread_color_match.group(1).strip()
                    else:
                        # Second try: "Used X.XXm COLOR thread" (without "for" at the end)
                        thread_color_match = re.search(r'Used\s+\d+\.?\d*\s*m\s+([a-zA-Z]+(?:\s+[a-zA-Z]+)?)\s+thread\b', notes, re.IGNORECASE)
                        if thread_color_match:
                            thread_color = thread_color_match.group(1).strip()
                        else:
                            # Third try: "(Xcm) COLOR thread" format
                            thread_color_match = re.search(r'\([^)]+\)\s+([a-zA-Z]+(?:\s+[a-zA-Z]+)?)\s+thread', notes, re.IGNORECASE)
                            if thread_color_match:
                                thread_color = thread_color_match.group(1).strip()
                            else:
                                # Fourth try: "X.XXm COLOR thread" (without "Used")
                                thread_color_match = re.search(r'\d+\.?\d*\s*m\s+([a-zA-Z]+(?:\s+[a-zA-Z]+)?)\s+thread\b', notes, re.IGNORECASE)
                                if thread_color_match:
                                    thread_color = thread_color_match.group(1).strip()
                    
                    # Clean up thread color - remove common non-color words
                    if thread_color:
                        # Remove words that aren't colors
                        thread_color = re.sub(r'\s+(for|thread|order|m|meter|meters|customize|repair)$', '', thread_color, flags=re.IGNORECASE)
                        thread_color = thread_color.strip()
                    
                    # If still no color found, try to extract from product name or description
                    if not thread_color:
                        # Check product name for color indicators
                        name_parts = material_name.lower().split()
                        common_colors = ['red', 'blue', 'green', 'yellow', 'black', 'white', 'brown', 'gray', 'grey', 
                                       'pink', 'purple', 'orange', 'navy', 'beige', 'tan', 'cream', 'ivory', 'maroon',
                                       'burgundy', 'khaki', 'olive', 'teal', 'cyan', 'magenta', 'lime', 'coral']
                        for color in common_colors:
                            if color in name_parts:
                                thread_color = color.title()
                                break
                
                # If this is thread but no color was extracted, set material_type to Thread
                if is_thread and not material_type:
                    material_type = 'Thread'
                
                # Format quantity based on unit type
                if unit.lower() in ['yard', 'yards']:
                    quantity_display = f"{quantity:.2f} yard{'s' if quantity != 1 else ''}"
                elif unit.lower() in ['meter', 'meters', 'm']:
                    quantity_display = f"{quantity:.2f} meter{'s' if quantity != 1 else ''}"
                elif unit.lower() in ['piece', 'pieces']:
                    quantity_display = f"{quantity:.1f} piece{'s' if quantity != 1 else ''}"
                elif unit.lower() in ['group', 'groups']:
                    quantity_display = f"{quantity:.1f} group{'s' if quantity != 1 else ''}"
                elif unit.lower() in ['inch', 'inches']:
                    quantity_display = f"{quantity:.2f} inch{'es' if quantity != 1 else ''}"
                elif unit.lower() in ['cm', 'centimeter', 'centimeters']:
                    quantity_display = f"{quantity:.2f} cm"
                else:
                    quantity_display = f"{quantity:.1f} {unit}"
                
                # For thread, always use meters and show color/length if available
                if is_thread or material_type == 'Thread':
                    # Use the extracted length if available, otherwise use quantity converted to meters
                    if thread_length is not None:
                        thread_length_display = f"{thread_length:.2f} meter{'s' if thread_length != 1 else ''}"
                    else:
                        # Convert quantity to meters if needed
                        if unit.lower() in ['cm', 'centimeter', 'centimeters']:
                            thread_length_display = f"{(quantity / 100.0):.2f} meter{'s' if quantity / 100.0 != 1 else ''}"
                        elif unit.lower() in ['meter', 'meters', 'm']:
                            thread_length_display = f"{quantity:.2f} meter{'s' if quantity != 1 else ''}"
                        else:
                            thread_length_display = f"{quantity:.2f} meter{'s' if quantity != 1 else ''}"
                    
                    # Format thread display - always show length, and color if available
                    if thread_color:
                        display_text = f"{material_name} ({material_type or 'Thread'}, {thread_length_display}, Color: {thread_color})"
                    else:
                        # Show thread with length even if color is not available
                        display_text = f"{material_name} ({material_type or 'Thread'}, {thread_length_display})"
                elif material_type:
                    display_text = f"{material_name} ({material_type}, {quantity_display})"
                else:
                    display_text = f"{material_name} ({quantity_display})"
                
                material_info = {
                    'name': material_name,
                    'quantity': quantity,
                    'unit': unit,
                    'material_type': material_type,
                    'display': display_text
                }
                
                # Add thread-specific information if available
                if material_type == 'Thread':
                    if thread_color:
                        material_info['thread_color'] = thread_color
                    if thread_length is not None:
                        material_info['thread_length'] = thread_length
                        material_info['thread_length_display'] = f"{thread_length:.2f} meter{'s' if thread_length != 1 else ''}"
                
                order_history[order_key]['materials'].append(material_info)
            
            # Convert to list and format for response
            history_list = []
            for order_data in order_history.values():
                materials_list = order_data['materials']
                materials_display = []
                
                # Format materials string according to user's format
                # Example: "Spun Polyester (Qty: -101.0), and "Magic Zipper (Qty: -9.0)"
                if len(materials_list) == 1:
                    materials_display.append(f'"{materials_list[0]["display"]}"')
                    used_text = 'used a'
                elif len(materials_list) == 2:
                    materials_display.append(f'"{materials_list[0]["display"]}", and "{materials_list[1]["display"]}"')
                    used_text = 'used a'
                else:
                    # More than 2 materials
                    used_text = 'used'
                    for i, mat in enumerate(materials_list):
                        if i == 0:
                            materials_display.append(f'"{mat["display"]}"')
                        elif i == len(materials_list) - 1:
                            materials_display.append(f', and "{mat["display"]}"')
                        else:
                            materials_display.append(f', "{mat["display"]}"')
                
                materials_str = ''.join(materials_display)
                
                # Get all unique material names for the "was deducted to" part
                material_names = [mat['name'] for mat in materials_list]
                if len(material_names) == 1:
                    materials_summary = f'"{material_names[0]}"'
                elif len(material_names) == 2:
                    materials_summary = f'"{material_names[0]}" and "{material_names[1]}"'
                else:
                    materials_summary = f'"{material_names[0]}" and {len(material_names) - 1} other material(s)'
                
                history_list.append({
                    'order_id': order_data['order_id'],
                    'order_identifier': order_data['order_identifier'],
                    'order_type': order_data['order_type'],
                    'order_type_display': order_data['order_type_display'],
                    'created_at': order_data['created_at'],
                    'created_at_timestamp': order_data.get('created_at_timestamp'),
                    'materials': order_data['materials'],
                    'materials_display': materials_str,
                    'materials_summary': materials_summary,
                    'used_text': used_text,
                    'formatted_text': f'Order "{order_data["order_identifier"]}", "{order_data["order_type_display"]}" {used_text} {materials_str}. was deducted to the total number of the {materials_summary}'
                })
            
            # Sort by created_at_timestamp descending (most recent first)
            # If timestamp is not available, fall back to string comparison
            history_list.sort(key=lambda x: x.get('created_at_timestamp') or x['created_at'], reverse=True)
            
            # Ensure we have complete material records for all orders
            # Check if any repair/customize orders are missing material transactions
            order_ids_with_transactions = set([t.reference_order.id for t in transactions if t.reference_order])
            
            # Get all repair and customize orders in the same date range
            order_query = Order.objects.filter(
                order_type__in=['repair', 'customize'],
                status__in=['completed', 'in_progress', 'ready_to_pick_up', 'repair_done']
            ).select_related('customer')
            
            # Apply same date filters
            if date_from:
                try:
                    from datetime import datetime
                    date_from_obj = datetime.strptime(date_from, '%Y-%m-%d')
                    order_query = order_query.filter(created_at__gte=date_from_obj)
                except ValueError:
                    pass
            
            if date_to:
                try:
                    from datetime import datetime, timedelta
                    date_to_obj = datetime.strptime(date_to, '%Y-%m-%d') + timedelta(days=1)
                    order_query = order_query.filter(created_at__lt=date_to_obj)
                except ValueError:
                    pass
            
            # Get orders that might be missing material transactions
            orders_without_transactions = order_query.exclude(id__in=order_ids_with_transactions).order_by('-created_at')[:limit]
            
            # For orders without transactions, try to create them retroactively
            for order in orders_without_transactions:
                # Check if order has any material transactions at all
                has_transactions = InventoryTransaction.objects.filter(
                    reference_order=order,
                    transaction_type='out',
                    product__product_type='material'
                ).exists()
                
                if not has_transactions:
                    # Try to retroactively record materials for this order
                    # This is a best-effort attempt - we can't always reconstruct order_data
                    # But we can at least log that materials might be missing
                    try:
                        # Try to get order data from order notes or other sources
                        # For now, we'll just note that this order might have missing materials
                        import logging
                        logger = logging.getLogger(__name__)
                        logger.warning(f'Order {order.order_identifier} ({order.order_type}) may have missing material transactions')
                    except Exception:
                        pass
            
            return JsonResponse({
                'success': True,
                'history': history_list[:limit],  # Apply limit to final results
                'count': len(history_list[:limit]),
                'total_found': len(history_list),
                'orders_checked': len(orders_without_transactions)
            })
            
        except Exception as e:
            import traceback
            traceback.print_exc()
            return JsonResponse({
                'success': False,
                'error': str(e),
                'history': [],
                'count': 0
            })
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})


@login_required
def api_product_detail(request, product_id):
    """API endpoint to get detailed information about a specific product"""
    if request.method == 'GET':
        try:
            product = get_object_or_404(Product, id=product_id)
            
            # Determine product type display
            product_type_display_map = {
                'rental': 'Rental Item',
                'material': 'Material',
                'service': 'Service'
            }
            product_type_display = product_type_display_map.get(product.product_type, product.product_type.title())
            
            # Determine status and status_class based on product type
            status = 'UNDEFINED'
            status_class = 'status-undefined'
            
            if product.product_type == 'rental':
                if product.rental_status == 'available':
                    status = 'AVAILABLE'
                    status_class = 'status-available'
                elif product.rental_status == 'rented':
                    if product.is_overdue:
                        status = 'OVERDUE'
                        status_class = 'status-overdue'
                    else:
                        status = 'RENTED'
                        status_class = 'status-rented'
                elif product.rental_status == 'maintenance':
                    status = 'MAINTENANCE'
                    status_class = 'status-maintenance'
            elif product.product_type == 'material':
                if product.quantity == 0:
                    status = 'OUT OF STOCK'
                    status_class = 'status-out-of-stock'
                elif product.is_low_stock:
                    status = 'LOW STOCK'
                    status_class = 'status-low-stock'
                else:
                    status = 'IN STOCK'
                    status_class = 'status-in-stock'
            elif product.product_type == 'service':
                if product.is_active:
                    status = 'ACTIVE'
                    status_class = 'status-active'
                else:
                    status = 'INACTIVE'
                    status_class = 'status-inactive'
            
            # Build images array
            images = []
            if product.image:
                images.append({
                    'url': product.image.url,
                    'alt': product.name
                })
            
            # Parse measurements from description if present
            measurements = {}
            description_text = product.description or ''
            
            # Try multiple patterns to find measurements
            measurements_found = False
            
            # Pattern 1: Look for "Measurements:" followed by JSON
            if 'Measurements:' in description_text:
                try:
                    # Extract JSON from description - handle multi-line JSON
                    measurements_start = description_text.find('Measurements:')
                    if measurements_start != -1:
                        measurements_json = description_text[measurements_start + len('Measurements:'):].strip()
                        
                        # Try to parse the JSON - handle both single-line and multi-line JSON
                        # First, try direct parsing (works for both single-line and properly formatted multi-line)
                        try:
                            measurements = json.loads(measurements_json)
                            measurements_found = True
                        except json.JSONDecodeError as e:
                            # If that fails, try to extract JSON object using regex (handles multi-line with indentation)
                            import re

                            # Match JSON object from first { to last } including all content
                            # Use non-greedy match to get the complete JSON object
                            json_pattern = r'\{[\s\S]*?\}'
                            json_match = re.search(json_pattern, measurements_json, re.MULTILINE | re.DOTALL)
                            if json_match:
                                try:
                                    measurements = json.loads(json_match.group())
                                    measurements_found = True
                                except json.JSONDecodeError:
                                    # Try to clean up the JSON string
                                    # Remove any trailing text after the closing brace
                                    cleaned_json = json_match.group()
                                    # Find the last closing brace
                                    last_brace = cleaned_json.rfind('}')
                                    if last_brace != -1:
                                        cleaned_json = cleaned_json[:last_brace + 1]
                                        try:
                                            measurements = json.loads(cleaned_json)
                                            measurements_found = True
                                        except:
                                            pass
                        
                        if measurements_found and isinstance(measurements, dict):
                            def _normalize_key(k):
                                k = str(k).strip().lower().replace(' ', '_')
                                if k.startswith('measurement_'):
                                    k = k[len('measurement_'):]
                                d = {
                                    'len': 'length',
                                    'length': 'length',
                                    'chest': 'bust',
                                    'bust': 'bust',
                                    'waist': 'waist',
                                    'hip': 'hips',
                                    'hips': 'hips',
                                    'shoulders': 'shoulder',
                                    'shoulder': 'shoulder',
                                    'sleeve_length': 'sleeve',
                                    'sleeve': 'sleeve',
                                    'arm_hole': 'armhole',
                                    'armhole': 'armhole',
                                    'neck_line': 'neckline',
                                    'neck': 'neckline',
                                    'neckline': 'neckline',
                                    'crotch': 'crotch',
                                    'thigh': 'thigh',
                                    'knee': 'knee',
                                    'bottom': 'bottom',
                                    'leg_opening': 'bottom',
                                    'hemwidth': 'hem_width',
                                    'hem_width': 'hem_width'
                                }
                                return d.get(k, k)
                            normalized = {}
                            for key, val in measurements.items():
                                nk = _normalize_key(key)
                                # Only include non-empty values, but keep the key for display
                                normalized[nk] = (str(val).strip() if val is not None and str(val).strip() else '')
                            measurements = normalized
                            description_text = description_text[:measurements_start].strip()
                except Exception as e:
                    # If all parsing fails, keep original description
                    # Log error for debugging but don't break
                    import logging
                    logger = logging.getLogger(__name__)
                    logger.warning(f"Failed to parse measurements for product {product.id}: {str(e)}")
                    pass
            
            # Pattern 2: If no measurements found, try to look for JSON object anywhere in description
            if not measurements_found and description_text:
                try:
                    import re

                    # Look for any JSON object in the description
                    json_pattern = r'\{[\s\S]*?\}'
                    json_match = re.search(json_pattern, description_text, re.MULTILINE | re.DOTALL)
                    if json_match:
                        try:
                            potential_measurements = json.loads(json_match.group())
                            # Check if it looks like measurements (has common measurement keys)
                            measurement_keys = ['length', 'bust', 'waist', 'hips', 'shoulder', 'sleeve', 'armhole', 'neckline', 'crotch', 'thigh', 'knee', 'bottom', 'hem_width']
                            if any(key.lower() in str(k).lower() for k in potential_measurements.keys() for key in measurement_keys):
                                measurements = potential_measurements
                                measurements_found = True
                        except:
                            pass
                except:
                    pass
            
            # Get transaction history
            transaction_history = []
            try:
                from .models import InventoryTransaction
                transactions = InventoryTransaction.objects.filter(product=product).order_by('-created_at')[:10]
                for trans in transactions:
                    transaction_history.append({
                        'id': trans.id,
                        'transaction_type': trans.transaction_type,
                        'quantity': trans.quantity,
                        'notes': trans.notes or '',
                        'created_at': trans.created_at.isoformat() if trans.created_at else None
                    })
            except Exception:
                pass
            
            # Ensure measurements is always a dict (not None)
            if not isinstance(measurements, dict):
                measurements = {}
            
            # Ensure all measurement keys are present, even if empty
            # This ensures the frontend always displays all expected measurements
            # Normalize empty values to empty strings for consistency
            all_measurement_keys = ['length', 'bust', 'waist', 'hips', 'shoulder', 'sleeve', 'armhole', 'neckline', 
                                   'crotch', 'thigh', 'knee', 'bottom', 'hem_width']
            normalized_measurements = {}
            
            # First, add all existing measurements (normalize values)
            for key, val in measurements.items():
                normalized_key = key.lower().strip().replace(' ', '_')
                # Ensure value is a string, normalize empty/null to empty string
                if val is None:
                    normalized_measurements[normalized_key] = ''
                else:
                    normalized_val = str(val).strip()
                    normalized_measurements[normalized_key] = normalized_val if normalized_val else ''
            
            # Add missing keys as empty strings so frontend can display them
            for key in all_measurement_keys:
                if key not in normalized_measurements:
                    normalized_measurements[key] = ''
            
            # Only keep keys that are in the all_measurement_keys list or were in original measurements
            final_measurements = {}
            for key in all_measurement_keys:
                if key in normalized_measurements:
                    final_measurements[key] = normalized_measurements[key]
            
            # Also include any additional measurement keys that aren't in the standard list
            for key, val in normalized_measurements.items():
                if key not in final_measurements:
                    final_measurements[key] = val
            
            measurements = final_measurements
            
            # Debug: Log measurements for troubleshooting
            import logging
            logger = logging.getLogger(__name__)
            logger.info(f"Product {product.id} measurements: {measurements}")
            
            return JsonResponse({
                'success': True,
                'product': {
                    'id': product.id,
                    'name': product.name,
                    'description': description_text,
                    'measurements': measurements,
                    'product_type': product.product_type,
                    'product_type_display': product_type_display,
                    'category': product.category.name if product.category else None,
                    'quantity': product.quantity,
                    'min_quantity': product.min_quantity,
                    'price': float(product.price) if product.price else 0.0,
                    'cost': float(product.cost) if product.cost else 0.0,
                    'image_url': product.image.url if product.image else None,
                    'images': images,
                    'is_available': product.is_available,
                    'is_active': product.is_active,
                    'is_archived': product.is_archived,
                    'is_low_stock': product.is_low_stock,
                    'rental_status': product.rental_status if product.product_type == 'rental' else None,
                    'status': status,
                    'status_class': status_class,
                    'created_at': product.created_at.isoformat() if product.created_at else None,
                    'transaction_history': transaction_history
                }
            })
        except Exception as e:
            import traceback
            traceback.print_exc()
            return JsonResponse({
                'success': False,
                'error': str(e)
            })
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})


@login_required
def api_edit_product(request, product_id):
    """API endpoint to edit a product"""
    if request.method == 'POST':
        try:
            import json
            product = get_object_or_404(Product, id=product_id)
            data = json.loads(request.body)
            
            # Update product fields
            if 'name' in data:
                product.name = data['name']
            if 'description' in data:
                product.description = data['description']
            if 'price' in data:
                product.price = float(data['price'])
            if 'cost' in data:
                product.cost = float(data['cost'])
            if 'quantity' in data:
                product.quantity = int(data['quantity'])
            
            product.save()
            
            return JsonResponse({
                'success': True,
                'message': 'Product updated successfully',
                'product_id': product.id
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            })
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})


@login_required
def api_adjust_stock(request, product_id):
    """API endpoint to adjust product stock"""
    if request.method == 'POST':
        try:
            import json
            product = get_object_or_404(Product, id=product_id)
            data = json.loads(request.body)
            
            adjustment = int(data.get('adjustment', 0))
            notes = data.get('notes', '')
            
            old_quantity = product.quantity
            product.quantity = max(0, product.quantity + adjustment)
            product.save()
            
            # Create inventory transaction
            InventoryTransaction.objects.create(
                product=product,
                transaction_type='in' if adjustment > 0 else 'out',
                quantity=abs(adjustment),
                notes=notes or f'Stock adjustment: {adjustment:+d}'
            )
            
            return JsonResponse({
                'success': True,
                'message': 'Stock adjusted successfully',
                'old_quantity': old_quantity,
                'new_quantity': product.quantity,
                'adjustment': adjustment
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            })
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})


@login_required
def api_accounting_details(request):
    """API endpoint to get accounting details"""
    if request.method == 'GET':
        try:
            from decimal import Decimal

            from .models import OrderItem, Sales
            
            metric_type = request.GET.get('type', '')
            
            # Get all completed sales
            completed_sales = Sales.objects.filter(order__status='completed').select_related('order__customer')
            
            # Get all completed orders
            completed_orders = Order.objects.filter(status='completed').select_related('customer')
            
            if metric_type == 'cash_on_hand':
                # Cash on hand = all completed sales (represents payments received)
                total = completed_sales.aggregate(total=Sum('amount'))['total'] or Decimal('0')
                
                # Get breakdown of sales (last 50) - these represent payments received
                breakdown = []
                for sale in completed_sales.order_by('-created_at')[:50]:
                    order = sale.order
                    breakdown.append({
                        'order_id': order.order_identifier if order else 'N/A',
                        'customer': order.customer.name if order and order.customer else 'N/A',
                        'order_type': order.get_order_type_display() if order else 'N/A',
                        'payment_method': sale.payment_method or 'Cash',
                        'amount': float(sale.amount),
                        'date': sale.created_at.strftime('%Y-%m-%d %H:%M:%S') if sale.created_at else 'N/A'
                    })
                
                return JsonResponse({
                    'success': True,
                    'type': 'cash_on_hand',
                    'title': 'Cash on Hand Details',
                    'description': 'Total amount received from all payments. This represents actual cash collected.',
                    'total': float(total),
                    'breakdown': breakdown
                })
            
            elif metric_type == 'accounts_receivable':
                # Accounts receivable = completed orders total - sales (paid amounts)
                total_orders = completed_orders.aggregate(total=Sum('total_amount'))['total'] or Decimal('0')
                total_sales_amount = completed_sales.aggregate(total=Sum('amount'))['total'] or Decimal('0')
                total = total_orders - total_sales_amount
                if total < 0:
                    total = Decimal('0')
                
                # Get breakdown of unpaid orders (last 50)
                breakdown = []
                
                for order in completed_orders.order_by('-created_at')[:50]:
                    # Get sales for this order (represents payments received)
                    order_sales = completed_sales.filter(order=order).aggregate(total=Sum('amount'))['total'] or Decimal('0')
                    balance = order.total_amount - order_sales
                    
                    if balance > 0:
                        breakdown.append({
                            'order_id': order.order_identifier,
                            'customer': order.customer.name if order.customer else 'N/A',
                            'order_type': order.get_order_type_display(),
                            'total_amount': float(order.total_amount),
                            'paid_amount': float(order_sales),
                            'balance': float(balance),
                            'date': order.created_at.strftime('%Y-%m-%d %H:%M:%S') if order.created_at else 'N/A'
                        })
                
                return JsonResponse({
                    'success': True,
                    'type': 'accounts_receivable',
                    'title': 'Accounts Receivable Details',
                    'description': 'Amounts owed by customers for completed orders that haven\'t been fully paid.',
                    'total': float(total),
                    'breakdown': breakdown
                })
            
            elif metric_type == 'total_profit':
                # Total profit = total sales - total costs
                total_revenue = completed_sales.aggregate(total=Sum('amount'))['total'] or Decimal('0')
                total_costs = total_revenue * Decimal('0.3')  # 30% cost estimate
                total = total_revenue - total_costs
                profit_margin = (total / total_revenue * 100) if total_revenue > 0 else Decimal('0')
                
                # Get breakdown of profitable orders (last 50)
                breakdown = []
                for sale in completed_sales.order_by('-amount')[:50]:
                    order = sale.order
                    revenue = sale.amount
                    cost = revenue * Decimal('0.3')
                    profit = revenue - cost
                    margin = (profit / revenue * 100) if revenue > 0 else Decimal('0')
                    
                    breakdown.append({
                        'order_id': order.order_identifier if order else 'N/A',
                        'customer': order.customer.name if order and order.customer else 'N/A',
                        'order_type': order.get_order_type_display() if order else 'N/A',
                        'revenue': float(revenue),
                        'cost': float(cost),
                        'profit': float(profit),
                        'margin': float(margin),
                        'date': sale.created_at.strftime('%Y-%m-%d %H:%M:%S') if sale.created_at else 'N/A'
                    })
                
                return JsonResponse({
                    'success': True,
                    'type': 'total_profit',
                    'title': 'Total Profit Details',
                    'description': 'Total profit calculated as revenue minus estimated costs (30% cost ratio).',
                    'total': float(total),
                    'total_revenue': float(total_revenue),
                    'total_costs': float(total_costs),
                    'profit_margin': float(profit_margin),
                    'breakdown': breakdown
                })
            
            elif metric_type == 'total_costs':
                # Total costs = 30% of total revenue (estimate)
                total_revenue = completed_sales.aggregate(total=Sum('amount'))['total'] or Decimal('0')
                total = total_revenue * Decimal('0.3')
                
                # Get breakdown of costs by order (last 50) - use select_related to avoid N+1 queries
                breakdown = []
                sales_with_orders = completed_sales.select_related('order', 'order__customer').order_by('-amount')[:50]
                for sale in sales_with_orders:
                    try:
                        # Skip if sale has no order (shouldn't happen but safety check)
                        if not hasattr(sale, 'order') or sale.order is None:
                            continue
                            
                        order = sale.order
                        material_cost = sale.amount * Decimal('0.25')  # 25% material
                        labor_cost = sale.amount * Decimal('0.05')  # 5% labor
                        total_cost = material_cost + labor_cost
                        
                        # Get order items for breakdown
                        items = []
                        if order and order.id:
                            try:
                                order_items = OrderItem.objects.filter(order=order).select_related('product')
                                for item in order_items:
                                    if item.product:
                                        unit_cost = item.product.cost_price or Decimal('0')
                                        items.append({
                                            'product': item.product.name or 'Unknown Product',
                                            'quantity': item.quantity or 0,
                                            'unit_cost': float(unit_cost),
                                            'total_cost': float(unit_cost * item.quantity) if item.quantity else 0.0
                                        })
                            except Exception as e:
                                # If there's an error getting items, continue without them
                                items = []
                        
                        # Safely get order identifier
                        order_id = 'N/A'
                        try:
                            if order and hasattr(order, 'order_identifier') and order.order_identifier:
                                order_id = order.order_identifier
                            elif order and hasattr(order, 'id'):
                                order_id = f"Order-{order.id}"
                        except:
                            order_id = 'N/A'
                        
                        # Safely get customer name
                        customer_name = 'N/A'
                        try:
                            if order and order.customer and hasattr(order.customer, 'name'):
                                customer_name = order.customer.name
                        except:
                            customer_name = 'N/A'
                        
                        # Safely get order type
                        order_type_display = 'N/A'
                        try:
                            if order and hasattr(order, 'get_order_type_display'):
                                order_type_display = order.get_order_type_display()
                        except:
                            order_type_display = 'N/A'
                        
                        breakdown.append({
                            'order_id': order_id,
                            'customer': customer_name,
                            'order_type': order_type_display,
                            'material_costs': float(material_cost),
                            'labor_costs': float(labor_cost),
                            'total_cost': float(total_cost),
                            'items': items,
                            'date': sale.created_at.strftime('%Y-%m-%d %H:%M:%S') if sale.created_at else 'N/A'
                        })
                    except Exception as e:
                        # Skip this sale if there's an error processing it
                        import logging
                        logger = logging.getLogger(__name__)
                        logger.warning(f"Error processing sale {sale.id} for total_costs: {str(e)}")
                        continue
                
                return JsonResponse({
                    'success': True,
                    'type': 'total_costs',
                    'title': 'Total Costs Details',
                    'description': 'Total costs including material costs (25%) and labor/other costs (5%) estimated at 30% of revenue.',
                    'total': float(total),
                    'breakdown': breakdown if breakdown else []
                })
            
            else:
                # Default: sales-based stats (total_sales, monthly_sales, etc.)
                total = completed_sales.aggregate(total=Sum('amount'))['total'] or Decimal('0')
                
                # Get breakdown (last 50)
                breakdown = []
                for sale in completed_sales.order_by('-created_at')[:50]:
                    order = sale.order
                    breakdown.append({
                        'order_id': order.order_identifier if order else 'N/A',
                        'customer': order.customer.name if order and order.customer else 'N/A',
                        'order_type': order.get_order_type_display() if order else 'N/A',
                        'payment_method': sale.payment_method or 'Cash',
                        'amount': float(sale.amount),
                        'date': sale.created_at.strftime('%Y-%m-%d %H:%M:%S') if sale.created_at else 'N/A'
                    })
                
                return JsonResponse({
                    'success': True,
                    'type': metric_type,
                    'title': f'{metric_type.replace("_", " ").title()} Details',
                    'description': f'Detailed breakdown of {metric_type.replace("_", " ")}.',
                    'total': float(total),
                    'breakdown': breakdown
                })
                
        except Exception as e:
            import traceback
            return JsonResponse({
                'success': False,
                'error': str(e),
                'traceback': traceback.format_exc()
            })
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})


@login_required
def reports_dashboard(request):
    """Reports dashboard page with date filtering and pagination"""
    # Get date range from request, default to last 30 days
    date_from_str = request.GET.get('from', '')
    date_to_str = request.GET.get('to', '')
    
    # Default to last 30 days if no dates provided
    if not date_from_str:
        date_from = timezone.now() - timedelta(days=30)
    else:
        try:
            date_from = datetime.strptime(date_from_str, '%Y-%m-%d').date()
            date_from = timezone.make_aware(datetime.combine(date_from, datetime.min.time()))
        except ValueError:
            date_from = timezone.now() - timedelta(days=30)
    
    if not date_to_str:
        date_to = timezone.now()
    else:
        try:
            date_to = datetime.strptime(date_to_str, '%Y-%m-%d').date()
            date_to = timezone.make_aware(datetime.combine(date_to, datetime.max.time()))
        except ValueError:
            date_to = timezone.now()
    
    # Format dates for template display
    if date_from_str:
        date_from_display = date_from_str
    else:
        default_from = timezone.now() - timedelta(days=30)
        date_from_display = default_from.strftime('%Y-%m-%d')
    
    if date_to_str:
        date_to_display = date_to_str
    else:
        date_to_display = timezone.now().strftime('%Y-%m-%d')
    
    # Overall statistics (all time)
    total_orders = Order.objects.filter(is_archived=False).count()
    completed_orders = Order.objects.filter(status='completed', is_archived=False).count()
    pending_orders = Order.objects.filter(status='pending', is_archived=False).count()
    
    # Total revenue (all time)
    total_revenue = Sales.objects.filter(order__status='completed').aggregate(
        total=Sum('amount')
    )['total'] or 0
    
    # Completed orders in date range
    completed_orders_range = Order.objects.filter(
        status='completed',
        is_archived=False,
        created_at__gte=date_from,
        created_at__lte=date_to
    )
    completed_orders_count_range = completed_orders_range.count()
    
    # Calculate revenue breakdown for each order (40% staff, 60% owner for repair/customize)
    from decimal import Decimal
    orders_with_revenue = []
    for order in completed_orders_range:
        order_dict = {
            'order': order,
            'staff_revenue': Decimal('0'),
            'owner_revenue': Decimal('0')
        }
        # Only repair and customize orders have revenue split
        if order.order_type in ['repair', 'customize']:
            order_dict['staff_revenue'] = order.total_amount * Decimal('0.4')
            order_dict['owner_revenue'] = order.total_amount * Decimal('0.6')
        else:
            # Rental orders: owner gets 100%
            order_dict['owner_revenue'] = order.total_amount
        orders_with_revenue.append(order_dict)
    
    # Pagination for completed orders
    orders_page = request.GET.get('orders_page', 1)
    orders_paginator = Paginator(orders_with_revenue, 10)
    try:
        orders = orders_paginator.page(orders_page)
    except PageNotAnInteger:
        orders = orders_paginator.page(1)
    except EmptyPage:
        orders = orders_paginator.page(orders_paginator.num_pages)
    
    # Sales in date range with revenue breakdown
    sales_range = Sales.objects.filter(
        order__status='completed',
        created_at__gte=date_from,
        created_at__lte=date_to
    ).select_related('order').order_by('-created_at')
    
    # Calculate revenue breakdown for each sale
    sales_with_revenue = []
    for sale in sales_range:
        sale_dict = {
            'sale': sale,
            'staff_revenue': Decimal('0'),
            'owner_revenue': Decimal('0')
        }
        # Only repair and customize orders have revenue split
        if sale.order.order_type in ['repair', 'customize']:
            sale_dict['staff_revenue'] = sale.amount * Decimal('0.4')
            sale_dict['owner_revenue'] = sale.amount * Decimal('0.6')
        else:
            # Rental orders: owner gets 100%
            sale_dict['owner_revenue'] = sale.amount
        sales_with_revenue.append(sale_dict)
    
    # Pagination for sales
    sales_page = request.GET.get('sales_page', 1)
    sales_paginator = Paginator(sales_with_revenue, 10)
    try:
        sales_list = sales_paginator.page(sales_page)
    except PageNotAnInteger:
        sales_list = sales_paginator.page(1)
    except EmptyPage:
        sales_list = sales_paginator.page(sales_paginator.num_pages)
    
    # Staff revenue & salary summary (bucket since last withdrawal, within date range)
    staff_reports = []
    staff_chart_data = []
    staff_qs = User.objects.filter(is_staff=True, is_active=True).order_by('username')
    for staff in staff_qs:
        staff_completed_qs = Order.objects.filter(
            assigned_staff=staff,
            status='completed',
            order_type__in=['repair', 'customize'],
            created_at__gte=date_from,
            created_at__lte=date_to
        )
        last_withdrawal = StaffWithdrawal.objects.filter(staff=staff).order_by('-created_at').first()
        if last_withdrawal:
            staff_completed_qs = staff_completed_qs.filter(created_at__gt=last_withdrawal.created_at)

        staff_total_revenue = staff_completed_qs.aggregate(total=Sum('total_amount'))['total'] or Decimal('0')
        staff_owner_share = staff_total_revenue * Decimal('0.6')
        staff_share = staff_total_revenue * Decimal('0.4')

        staff_reports.append({
            'staff': staff,
            'completed_orders': staff_completed_qs.count(),
            'total_revenue': staff_total_revenue,
            'owner_revenue': staff_owner_share,
            'staff_revenue': staff_share,
            'last_withdrawal': last_withdrawal,
            'bucket_start': last_withdrawal.created_at if last_withdrawal else None,
        })

        staff_chart_data.append({
            'label': staff.get_full_name() or staff.username,
            'staff_revenue': float(staff_share),
            'owner_revenue': float(staff_owner_share),
            'total_revenue': float(staff_total_revenue),
        })

    # Materials inventory transactions in date range
    materials_logs_range = InventoryTransaction.objects.filter(
        created_at__gte=date_from,
        created_at__lte=date_to,
        product__product_type='material'
    ).select_related('product', 'reference_order').order_by('-created_at')
    
    # Helper functions for formatting repair and customize types (same as orders_list)
    def format_repair_type(repair_type_str):
        """Format repair type from snake_case to Title Case"""
        if not repair_type_str:
            return None
        # Remove "Repair - " prefix if present
        if "Repair - " in repair_type_str:
            repair_type_str = repair_type_str[10:].strip()
        # Remove "repair - " prefix if present (lowercase)
        if "repair - " in repair_type_str.lower():
            repair_type_str = repair_type_str[repair_type_str.lower().find("repair - ") + 9:].strip()
        # Remove class suffix like "(Class standard)"
        if " (Class " in repair_type_str:
            repair_type_str = repair_type_str.split(" (Class ")[0].strip()
        # Replace underscores with spaces and title case
        formatted = repair_type_str.replace('_', ' ').title()
        return formatted
    
    def format_customize_type(customize_type_str):
        """Format customize type name"""
        if not customize_type_str:
            return None
        # Remove "Customize - " prefix if present
        if "Customize - " in customize_type_str:
            customize_type_str = customize_type_str[11:].strip()
        # Remove "customize - " prefix if present (lowercase)
        if "customize - " in customize_type_str.lower():
            customize_type_str = customize_type_str[customize_type_str.lower().find("customize - ") + 12:].strip()
        # Remove class suffix like "(Class standard)"
        if " (Class " in customize_type_str:
            customize_type_str = customize_type_str.split(" (Class ")[0].strip()
        # Handle common customize types
        customize_type_lower = customize_type_str.lower()
        if 'uniform' in customize_type_lower:
            return "Uniform"
        elif 'pe' in customize_type_lower and len(customize_type_lower) <= 5:
            return "PE"
        # Replace underscores with spaces and title case
        formatted = customize_type_str.replace('_', ' ').title()
        return formatted
    
    # Helper function to format materials_used according to user requirements
    def format_materials_used(trans, order_type, category):
        """Format materials used in the specific format requested by user"""
        import json
        from decimal import Decimal
        
        materials_parts = []
        
        # Get order data if available - try to extract from transaction notes or order items
        order_data = None
        # Since Order model doesn't have order_data field, we'll extract info from notes and order items
        # Try to parse order data from transaction notes or infer from order items
        if trans.reference_order:
            # Try to get repair type from order items
            order_items = trans.reference_order.items.select_related('product').all()
            repair_type = None
            for item in order_items:
                if item.product and item.product.name:
                    product_name = item.product.name.lower()
                    if 'repair' in product_name:
                        # Extract repair type from product name
                        if 'button' in product_name:
                            repair_type = 'buttons'
                        elif 'zipper' in product_name:
                            repair_type = 'zipper_replacement'
                        elif 'lock' in product_name or 'kawit' in product_name:
                            repair_type = 'lock_repair'
                        elif 'patch' in product_name:
                            repair_type = 'patches'
                        elif 'bewang' in product_name:
                            repair_type = 'bewang'
                        elif 'elastic' in product_name:
                            repair_type = 'elastic'
                        break
            
            # Build order_data from available information
            order_data = {}
            if repair_type:
                order_data['repair_type'] = repair_type
            
            # Try to extract information from transaction notes
            if trans.notes:
                import re
                notes = trans.notes.lower()
                
                # Extract button quantity
                button_match = re.search(r'(\d+)\s*(?:pieces?|pcs)', notes)
                if button_match:
                    order_data['selected_button_quantity_used'] = int(button_match.group(1))
                
                # Extract lock groups
                lock_match = re.search(r'(\d+)\s*(?:groups?|set)', notes)
                if lock_match:
                    order_data['selected_lock_groups_used'] = int(lock_match.group(1))
                
                # Extract thread color
                thread_color_match = re.search(r'(\w+)\s+thread', notes)
                if thread_color_match:
                    order_data['thread_color'] = thread_color_match.group(1)
                
                # Extract thread meters
                thread_meters_match = re.search(r'(\d+\.?\d*)\s*m', notes)
                if thread_meters_match:
                    order_data['thread_meters'] = float(thread_meters_match.group(1))
                
                # Extract zipper inches
                zipper_match = re.search(r'(\d+\.?\d*)\s*inches?', notes)
                if zipper_match:
                    order_data['selected_zipper_inches_used'] = float(zipper_match.group(1))
                
                # Extract patch quantity
                patch_match = re.search(r'(\d+)\s*(?:patches?|pcs)', notes)
                if patch_match:
                    order_data['patch_quantity'] = int(patch_match.group(1))
        
        # Get product info
        product_name = trans.product.name if trans.product else ''
        quantity = abs(float(trans.quantity)) if trans.quantity else 0
        unit = trans.product.unit_of_measurement if trans.product else 'pcs'
        
        # Format based on order type and category
        if order_type == 'Repair':
            if 'Button' in category or category == 'Buttons Replacement':
                # Buttons repair: "1 set/ 8pcs, Metal Buttons, & Black Thread 200(cm)"
                if order_data:
                    button_qty = order_data.get('selected_button_quantity_used') or order_data.get('button_quantity', 0)
                    thread_color = order_data.get('thread_color') or order_data.get('selected_thread_color_buttons', '')
                    thread_color_other = order_data.get('thread_color_other') or order_data.get('selected_thread_color_other_buttons', '')
                    
                    if button_qty:
                        pieces = int(button_qty)
                        sets = pieces / 8.0
                        if sets >= 1:
                            materials_parts.append(f"{int(sets)} set/ {pieces}pcs")
                        else:
                            materials_parts.append(f"{pieces}pcs")
                        
                        # Get button type from product name or order data
                        button_type = 'Metal Buttons'
                        if trans.product and trans.product.name:
                            if 'metal' in trans.product.name.lower():
                                button_type = 'Metal Buttons'
                            elif 'plastic' in trans.product.name.lower():
                                button_type = 'Plastic Buttons'
                            else:
                                button_type = trans.product.name
                        
                        materials_parts.append(button_type)
                        
                        # Add thread info
                        final_thread_color = thread_color_other if thread_color == 'other' else (thread_color or 'Black')
                        if final_thread_color:
                            # Calculate thread meters from notes or use default
                            thread_meters = 0
                            if trans.notes:
                                # Try to extract thread meters from notes
                                import re
                                match = re.search(r'(\d+\.?\d*)\s*m', trans.notes, re.IGNORECASE)
                                if match:
                                    thread_meters = float(match.group(1))
                            
                            if thread_meters == 0:
                                # Default: 2.2 meters per group for buttons
                                thread_meters = sets * 2.2
                            
                            thread_cm = int(thread_meters * 100)
                            materials_parts.append(f"& {final_thread_color.title()} Thread {thread_cm}(cm)")
            
            elif 'KAWIT' in category.upper() or 'LOCKS' in category.upper() or 'Lock' in category:
                # Kawit/Locks: "1 set/ 8pcs, Kawit/locks, & Black Thread 200(cm)"
                lock_groups = 0
                if order_data:
                    lock_groups = order_data.get('selected_lock_groups_used') or order_data.get('lock_groups', 0)
                
                # If not in order_data, try to extract from transaction quantity
                if not lock_groups and trans.quantity:
                    lock_groups = int(abs(float(trans.quantity)))
                
                if lock_groups:
                    groups = int(lock_groups)
                    pieces = groups * 4  # 4 pieces per group
                    sets = pieces / 8.0
                    if sets >= 1:
                        materials_parts.append(f"{int(sets)} set/ {pieces}pcs")
                    else:
                        materials_parts.append(f"{pieces}pcs")
                    
                    materials_parts.append("Kawit/locks")
                    
                    # Add thread info
                    thread_color = 'Black'
                    if order_data:
                        thread_color = order_data.get('thread_color', 'Black')
                    
                    # Try to extract from notes
                    if thread_color == 'Black' and trans.notes:
                        import re
                        color_match = re.search(r'(\w+)\s+thread', trans.notes, re.IGNORECASE)
                        if color_match:
                            thread_color = color_match.group(1)
                    
                    # Calculate thread meters
                    thread_meters = 0
                    if trans.notes:
                        import re
                        match = re.search(r'(\d+\.?\d*)\s*m', trans.notes, re.IGNORECASE)
                        if match:
                            thread_meters = float(match.group(1))
                    
                    if thread_meters == 0:
                        # Default: 2 meters per group for locks
                        thread_meters = groups * 2.0
                    
                    thread_cm = int(thread_meters * 100)
                    materials_parts.append(f"& {thread_color.title()} Thread {thread_cm}(cm)")
            
            elif 'Zipper' in category:
                # Zipper replacement - keep simple format
                materials_parts.append(f"{trans.product.name}: {quantity} {unit}")
                
                # Extract thread info from notes or order_data
                thread_color = 'Black'
                zipper_inches = 0
                
                if order_data:
                    thread_color = order_data.get('thread_color', 'Black')
                    zipper_inches = order_data.get('selected_zipper_inches_used', 0)
                
                # Try to extract from notes
                if trans.notes:
                    import re
                    color_match = re.search(r'(\w+)\s+thread', trans.notes, re.IGNORECASE)
                    if color_match:
                        thread_color = color_match.group(1)
                    
                    inches_match = re.search(r'(\d+\.?\d*)\s*inches?', trans.notes, re.IGNORECASE)
                    if inches_match:
                        zipper_inches = float(inches_match.group(1))
                
                if thread_color and zipper_inches:
                    thread_meters = calculate_thread_for_zipper(float(zipper_inches))
                    thread_cm = int(thread_meters * 100)
                    materials_parts.append(f"& {thread_color.title()} Thread {thread_cm}(cm)")
                elif thread_color:
                    # Try to extract thread meters from notes
                    if trans.notes:
                        import re
                        meters_match = re.search(r'(\d+\.?\d*)\s*m', trans.notes, re.IGNORECASE)
                        if meters_match:
                            thread_meters = float(meters_match.group(1))
                            thread_cm = int(thread_meters * 100)
                            materials_parts.append(f"& {thread_color.title()} Thread {thread_cm}(cm)")
            
            elif 'Patch' in category or category == 'Patches':
                # Patches repair
                patch_qty = 0
                if order_data:
                    patch_qty = order_data.get('patch_quantity', 0)
                
                # If not in order_data, try to extract from transaction quantity
                if not patch_qty and trans.quantity:
                    patch_qty = int(abs(float(trans.quantity)))
                
                if patch_qty:
                    materials_parts.append(f"{patch_qty} pcs, {trans.product.name if trans.product else 'Patches'}")
                    
                    # Get thread color
                    thread_color = 'Black'
                    if order_data:
                        thread_color = order_data.get('thread_color', 'Black')
                    
                    # Try to extract from notes
                    if thread_color == 'Black' and trans.notes:
                        import re
                        color_match = re.search(r'(\w+)\s+thread', trans.notes, re.IGNORECASE)
                        if color_match:
                            thread_color = color_match.group(1)
                    
                    # Calculate thread meters
                    thread_meters = 0
                    if trans.notes:
                        import re
                        match = re.search(r'(\d+\.?\d*)\s*m', trans.notes, re.IGNORECASE)
                        if match:
                            thread_meters = float(match.group(1))
                    
                    if thread_meters == 0:
                        # Default thread calculation for patches
                        patch_size = 'medium'
                        if order_data:
                            patch_size = order_data.get('patch_size', 'medium').lower()
                        thread_per_patch = {'small': 0.5, 'medium': 1.0, 'large': 1.25, 'xl': 1.75}.get(patch_size, 1.0)
                        thread_meters = int(patch_qty) * thread_per_patch
                    
                    thread_cm = int(thread_meters * 100)
                    materials_parts.append(f"& {thread_color.title()} Thread {thread_cm}(cm)")
            
            else:
                # Other repair types - use default format
                materials_parts.append(f"{trans.product.name}: {quantity} {unit}")
                if order_data:
                    thread_color = order_data.get('thread_color', '')
                    thread_color_other = order_data.get('thread_color_other', '')
                    if thread_color or thread_color_other:
                        final_thread_color = thread_color_other if thread_color == 'other' else (thread_color or 'Black')
                        thread_meters = order_data.get('thread_meters', 0)
                        if thread_meters:
                            thread_cm = int(float(thread_meters) * 100)
                            materials_parts.append(f"& {final_thread_color.title()} Thread {thread_cm}(cm)")
        
        elif order_type == 'Customize':
            # Customize orders
            if 'Uniform' in category or 'Wholeset' in category:
                # Uniform/Wholeset: "1 set/ 8pcs, Metal Buttons, & Assorted Thread 400(cm) & 10 yards Cotton fabric/white"
                # Check for buttons in transaction notes or order items
                buttons_needed = 0
                if order_data:
                    buttons_needed = order_data.get('uniform_buttons_needed', 0)
                
                # Try to extract from notes
                if not buttons_needed and trans.notes:
                    import re
                    button_match = re.search(r'(\d+)\s*(?:buttons?|pieces?|pcs)', trans.notes, re.IGNORECASE)
                    if button_match:
                        buttons_needed = int(button_match.group(1))
                
                if buttons_needed:
                    pieces = int(buttons_needed)
                    sets = pieces / 8.0
                    if sets >= 1:
                        materials_parts.append(f"{int(sets)} set/ {pieces}pcs")
                    else:
                        materials_parts.append(f"{pieces}pcs")
                    
                    button_type = 'Metal Buttons'
                    if order_data:
                        button_type = order_data.get('uniform_button_type', 'Metal Buttons')
                    materials_parts.append(button_type)
                
                # Thread - extract from notes
                thread_color = ''
                thread_cm = 0
                if order_data:
                    thread_color = order_data.get('customize_thread_color', '')
                    thread_cm = order_data.get('customize_thread_cm', 0)
                
                if not thread_color and trans.notes:
                    import re
                    color_match = re.search(r'(\w+)\s+thread', trans.notes, re.IGNORECASE)
                    if color_match:
                        thread_color = color_match.group(1)
                
                if not thread_cm and trans.notes:
                    import re
                    cm_match = re.search(r'(\d+)\s*cm', trans.notes, re.IGNORECASE)
                    if cm_match:
                        thread_cm = int(cm_match.group(1))
                
                if thread_color and thread_cm:
                    materials_parts.append(f"& {thread_color.title()} Thread {thread_cm}(cm)")
                elif thread_color:
                    materials_parts.append(f"& {thread_color.title()} Thread")
                
                # Fabric - extract from notes
                fabric = ''
                fabric_yards = 0
                fabric_color = 'white'
                if order_data:
                    fabric = order_data.get('customize_fabric', '')
                    fabric_yards = order_data.get('customize_fabric_yards', 0)
                    fabric_color = order_data.get('customize_fabric_color', 'white')
                
                if not fabric and trans.notes:
                    import re
                    fabric_match = re.search(r'(\w+)\s+fabric', trans.notes, re.IGNORECASE)
                    if fabric_match:
                        fabric = fabric_match.group(1)
                
                if not fabric_yards and trans.notes:
                    import re
                    yards_match = re.search(r'(\d+)\s*yards?', trans.notes, re.IGNORECASE)
                    if yards_match:
                        fabric_yards = int(yards_match.group(1))
                
                if fabric and fabric_yards:
                    materials_parts.append(f"& {fabric_yards} yards {fabric.title()} fabric/{fabric_color}")
            
            elif 'PE' in category:
                # PE: "Garter 12 (cm) & Black Thread 200(cm) & 10 yards Cotton fabric/white"
                # Garter - extract from notes
                garter_cm = 0
                if order_data:
                    garter_cm = order_data.get('garter_cm', 0) or order_data.get('customize_garter_cm', 0)
                
                if not garter_cm and trans.notes:
                    import re
                    garter_match = re.search(r'garter.*?(\d+)\s*cm', trans.notes, re.IGNORECASE)
                    if garter_match:
                        garter_cm = int(garter_match.group(1))
                
                if garter_cm:
                    materials_parts.append(f"Garter {garter_cm} (cm)")
                
                # Thread - extract from notes
                thread_color = ''
                thread_cm = 0
                if order_data:
                    thread_color = order_data.get('customize_thread_color', '')
                    thread_cm = order_data.get('customize_thread_cm', 0)
                
                if not thread_color and trans.notes:
                    import re
                    color_match = re.search(r'(\w+)\s+thread', trans.notes, re.IGNORECASE)
                    if color_match:
                        thread_color = color_match.group(1)
                
                if not thread_cm and trans.notes:
                    import re
                    cm_match = re.search(r'(\d+)\s*cm', trans.notes, re.IGNORECASE)
                    if cm_match:
                        thread_cm = int(cm_match.group(1))
                
                if thread_color and thread_cm:
                    materials_parts.append(f"& {thread_color.title()} Thread {thread_cm}(cm)")
                elif thread_color:
                    materials_parts.append(f"& {thread_color.title()} Thread")
                
                # Fabric - extract from notes
                fabric = ''
                fabric_yards = 0
                fabric_color = 'white'
                if order_data:
                    fabric = order_data.get('customize_fabric', '')
                    fabric_yards = order_data.get('customize_fabric_yards', 0)
                    fabric_color = order_data.get('customize_fabric_color', 'white')
                
                if not fabric and trans.notes:
                    import re
                    fabric_match = re.search(r'(\w+)\s+fabric', trans.notes, re.IGNORECASE)
                    if fabric_match:
                        fabric = fabric_match.group(1)
                
                if not fabric_yards and trans.notes:
                    import re
                    yards_match = re.search(r'(\d+)\s*yards?', trans.notes, re.IGNORECASE)
                    if yards_match:
                        fabric_yards = int(yards_match.group(1))
                
                if fabric and fabric_yards:
                    materials_parts.append(f"& {fabric_yards} yards {fabric.title()} fabric/{fabric_color}")
            
            else:
                # Other customize types
                materials_parts.append(f"{trans.product.name}: {quantity} {unit}")
        
        else:
            # Default format for other order types
            materials_parts.append(f"{trans.product.name}: {quantity} {unit}")
        
        # Join all parts with commas and proper spacing
        if materials_parts:
            return ', '.join(materials_parts)
        else:
            # Fallback to original format
            return f"{trans.product.name}: {quantity} {unit}" + (f"\nNotes: {trans.notes}" if trans.notes else "")
    
    # Format materials logs for template - fetch actual categories from order table
    materials_logs_list = []
    for trans in materials_logs_range:
        # Get order type - never show N/A, always infer from available data
        order_type = None
        order_identifier = None
        
        if trans.reference_order:
            # Use order type from reference order
            order_type_raw = trans.reference_order.order_type
            order_identifier = trans.reference_order.order_identifier
            
            # Format order type for display
            if order_type_raw == 'rent':
                order_type = 'Rent'
            elif order_type_raw == 'repair':
                order_type = 'Repair'
            elif order_type_raw == 'customize':
                order_type = 'Customize'
            else:
                order_type = order_type_raw.title() if order_type_raw else 'General Service'
        else:
            # For transactions without reference_order, try multiple methods to infer order type
            
            # Method 1: Check transaction notes for order references
            if trans.notes:
                notes_lower = trans.notes.lower()
                if 'repair' in notes_lower or 'repair order' in notes_lower:
                    order_type = 'repair'
                elif 'customize' in notes_lower or 'customize order' in notes_lower:
                    order_type = 'customize'
                elif 'rent' in notes_lower or 'rental' in notes_lower or 'rental order' in notes_lower:
                    order_type = 'rent'
                elif 'manual adjustment' in notes_lower or 'adjustment' in notes_lower:
                    # For manual adjustments, find the most recent order using this product
                    from .models import OrderItem
                    recent_order_item = OrderItem.objects.filter(
                        product=trans.product
                    ).select_related('order').order_by('-order__created_at').first()
                    if recent_order_item and recent_order_item.order:
                        order_type = recent_order_item.order.order_type
                        order_identifier = recent_order_item.order.order_identifier
                    else:
                        # If no order found, check if product is typically used for specific order types
                        order_type = 'adjustment'
            
            # Method 2: Check product type
            if not order_type and trans.product:
                if trans.product.product_type == 'rental':
                    order_type = 'rent'
                elif 'repair' in trans.product.name.lower():
                    order_type = 'repair'
                elif 'customize' in trans.product.name.lower():
                    order_type = 'customize'
            
            # Method 3: Use transaction type to infer
            if not order_type:
                if trans.transaction_type == 'rental_out' or trans.transaction_type == 'rental_in':
                    order_type = 'rent'
                elif trans.transaction_type == 'adjustment':
                    # For adjustments, try to find related orders
                    from .models import OrderItem
                    recent_order_item = OrderItem.objects.filter(
                        product=trans.product
                    ).select_related('order').order_by('-order__created_at').first()
                    if recent_order_item and recent_order_item.order:
                        order_type = recent_order_item.order.order_type
                        order_identifier = recent_order_item.order.order_identifier
                    else:
                        order_type = 'adjustment'
                elif trans.transaction_type in ['in', 'out']:
                    # For stock in/out, check if product is used in any orders
                    from .models import OrderItem
                    recent_order_item = OrderItem.objects.filter(
                        product=trans.product
                    ).select_related('order').order_by('-order__created_at').first()
                    if recent_order_item and recent_order_item.order:
                        order_type = recent_order_item.order.order_type
                        order_identifier = recent_order_item.order.order_identifier
                    else:
                        order_type = 'general'
                else:
                    order_type = 'general'
            
            # Format order type for display (use get_order_type_display format)
            # Ensure we never have None or empty - always provide a value
            if order_type:
                if order_type == 'rent':
                    order_type = 'Rent'
                elif order_type == 'repair':
                    order_type = 'Repair'
                elif order_type == 'customize':
                    order_type = 'Customize'
                elif order_type == 'adjustment':
                    order_type = 'Adjustment'
                elif order_type == 'general':
                    order_type = 'General Service'
                else:
                    order_type = order_type.title() if order_type else 'General Service'
            else:
                order_type = 'General Service'
        
        # Final safety check - ensure order_type is never None, empty, or 'N/A'
        if not order_type or str(order_type) == 'N/A' or (isinstance(order_type, str) and order_type.strip() == ''):
            order_type = 'General Service'
        
        # Ensure order_type is a string
        order_type = str(order_type) if order_type else 'General Service'
        
        # Get category from order items using same logic as orders table
        category = None
        if trans.reference_order:
            order_items = trans.reference_order.items.select_related('product__category').all()
            
            # For rent/rental orders, group by category like in orders table
            if trans.reference_order.order_type in ['rent', 'rental']:
                # Group rental items by category (same logic as format_rental_categories filter)
                categories = []
                for item in order_items:
                    if item.product and item.product.category:
                        cat_name = item.product.category.name
                        if cat_name and cat_name not in categories:
                            categories.append(cat_name)
                
                if categories:
                    # If only one category, return it; otherwise comma-separated
                    if len(categories) == 1:
                        category = categories[0]
                    else:
                        category = ', '.join(categories)
            
            # For repair orders, use same logic as orders table
            elif trans.reference_order.order_type == 'repair':
                categories = []
                for item in order_items:
                    category_name = None
                    
                    # Try to get repair type from product name (same as orders_list)
                    if item.product and item.product.name:
                        product_name = item.product.name
                        if "Repair - " in product_name or "repair - " in product_name.lower():
                            if "Repair - " in product_name:
                                repair_part = product_name.split("Repair - ")[1].strip()
                            else:
                                repair_part = product_name.split("repair - ")[1].strip()
                            if " (Class " in repair_part:
                                repair_part = repair_part.split(" (Class ")[0].strip()
                            category_name = format_repair_type(repair_part)
                        elif item.product.category:
                            category_name = item.product.category.name
                        else:
                            category_name = format_repair_type(product_name)
                    
                    # Fallback to product category
                    if not category_name and item.product and item.product.category:
                        category_name = item.product.category.name
                    
                    # Final fallback
                    if not category_name:
                        category_name = "Repair Service"
                    
                    if category_name and category_name not in categories:
                        categories.append(category_name)
                
                if categories:
                    category = ', '.join(categories)
            
            # For customize orders, use same logic as orders table
            elif trans.reference_order.order_type == 'customize':
                categories = []
                for item in order_items:
                    category_name = None
                    
                    # Try to get customize type from product name (same as orders_list)
                    if item.product and item.product.name:
                        product_name = item.product.name
                        if "Customize - " in product_name or "customize - " in product_name.lower():
                            if "Customize - " in product_name:
                                customize_part = product_name.split("Customize - ")[1].strip()
                            else:
                                customize_part = product_name.split("customize - ")[1].strip()
                            if " (Class " in customize_part:
                                customize_part = customize_part.split(" (Class ")[0].strip()
                            if " - " in customize_part:
                                customize_part = customize_part.split(" - ")[0].strip()
                            category_name = format_customize_type(customize_part)
                        elif item.product.description and "Type: " in item.product.description:
                            desc = item.product.description
                            if "Type: Uniform" in desc or "Type: uniform" in desc:
                                category_name = "Uniform"
                            elif "Type: PE" in desc or "Type: Pe" in desc or "Type: pe" in desc:
                                category_name = "PE"
                            else:
                                category_name = "Customize Service"
                        elif item.product.category:
                            category_name = item.product.category.name
                        else:
                            category_name = format_customize_type(product_name)
                    
                    # Fallback to product category
                    if not category_name and item.product and item.product.category:
                        category_name = item.product.category.name
                    
                    # Final fallback
                    if not category_name:
                        category_name = "Customize Service"
                    
                    if category_name and category_name not in categories:
                        categories.append(category_name)
                
                if categories:
                    category = ', '.join(categories)
            
            # For other order types, get unique categories
            else:
                categories = set()
                for item in order_items:
                    if item.product and item.product.category:
                        categories.add(item.product.category.name)
                if categories:
                    category = ', '.join(sorted(categories))
        
        # Fallback to product category if order category not found
        if not category and trans.product and trans.product.category:
            category = trans.product.category.name
        
        # Additional fallback - try to get from any related order item
        if not category and trans.reference_order:
            from .models import OrderItem
            order_item = OrderItem.objects.filter(
                order=trans.reference_order
            ).select_related('product__category').first()
            if order_item and order_item.product and order_item.product.category:
                category = order_item.product.category.name
        
        # Try to extract category from product name if it's a repair/customize product
        if not category and trans.product and trans.product.name:
            product_name = trans.product.name
            if "Repair - " in product_name or "repair - " in product_name.lower():
                if "Repair - " in product_name:
                    repair_part = product_name.split("Repair - ")[1].strip()
                else:
                    repair_part = product_name.split("repair - ")[1].strip()
                if " (Class " in repair_part:
                    repair_part = repair_part.split(" (Class ")[0].strip()
                category = format_repair_type(repair_part)
            elif "Customize - " in product_name or "customize - " in product_name.lower():
                if "Customize - " in product_name:
                    customize_part = product_name.split("Customize - ")[1].strip()
                else:
                    customize_part = product_name.split("customize - ")[1].strip()
                if " (Class " in customize_part:
                    customize_part = customize_part.split(" (Class ")[0].strip()
                if " - " in customize_part:
                    customize_part = customize_part.split(" - ")[0].strip()
                category = format_customize_type(customize_part)
        
        # Final fallback based on order type - never use "Uncategorized" or "N/A"
        if not category:
            if trans.reference_order:
                if trans.reference_order.order_type == 'repair':
                    category = 'Repair Service'
                elif trans.reference_order.order_type == 'customize':
                    category = 'Customize Service'
                elif trans.reference_order.order_type in ['rent', 'rental']:
                    category = 'Rental Service'
                else:
                    category = 'General Service'
            elif order_type:
                # Use the inferred order type to determine category
                # Handle both raw order_type (like 'rent') and formatted (like 'Rent')
                order_type_lower = str(order_type).lower() if order_type else ''
                if order_type_lower in ['rent', 'rental']:
                    category = 'Rental Service'
                elif order_type_lower == 'repair':
                    category = 'Repair Service'
                elif order_type_lower == 'customize':
                    category = 'Customize Service'
                elif order_type_lower == 'adjustment':
                    category = 'Adjustment'
                else:
                    category = 'General Service'
            else:
                category = 'General Service'
        
        # Final safety check - ensure category is never None, empty, or 'N/A'
        if not category or str(category) == 'N/A' or str(category) == 'Uncategorized' or (isinstance(category, str) and category.strip() == ''):
            category = 'General Service'
        
        # Ensure category is a string
        category = str(category) if category else 'General Service'
        
        # Format materials_used according to user requirements
        materials_used = format_materials_used(trans, order_type, category)
        
        materials_logs_list.append({
            'id': trans.id,
            'order_type': order_type,
            'order_identifier': order_identifier,
            'category': category,
            'materials_used': materials_used,
            'date_time': trans.created_at
        })
    
    # Pagination for materials logs
    materials_page = request.GET.get('materials_page', 1)
    materials_paginator = Paginator(materials_logs_list, 10)
    try:
        materials_logs = materials_paginator.page(materials_page)
    except PageNotAnInteger:
        materials_logs = materials_paginator.page(1)
    except EmptyPage:
        materials_logs = materials_paginator.page(materials_paginator.num_pages)
    
    # Rental items in date range (orders with rental products)
    rental_orders_range = Order.objects.filter(
        order_type='rent',
        is_archived=False,
        created_at__gte=date_from,
        created_at__lte=date_to
    ).select_related('customer').prefetch_related('items__product').order_by('-created_at')
    
    # Format rental items for template
    rental_items_list = []
    for order in rental_orders_range:
        # Get order items
        order_items = order.items.all()
        product_names = []
        total_quantity = 0
        added_items_list = []
        
        for item in order_items:
            if item.product:
                product_names.append(item.product.name)
                total_quantity += item.quantity
                added_items_list.append(f"{item.product.name} (x{item.quantity})")
        
        product_name = ", ".join(product_names) if product_names else "N/A"
        added_items = "\n".join(added_items_list) if added_items_list else "N/A"
        
        # Get rental dates (assuming order created_at is rental date)
        rented_date = order.created_at
        
        # Get returned date (if order is completed and has a completion date)
        returned_date = None
        if order.status == 'completed':
            # Try to find return date from related transactions or use updated_at
            returned_date = order.updated_at if order.updated_at else None
        
        # Calculate penalty (simplified - can be enhanced)
        penalty = None
        if order.status == 'overdue':
            # Calculate penalty based on overdue days
            penalty = "Calculate based on overdue days"
        
        rental_items_list.append({
            'order_number': order.order_identifier or f"Order #{order.id}",
            'order_id': order.id,
            'product_name': product_name,
            'quantity': total_quantity,
            'added_items': added_items,
            'rented_date': rented_date,
            'returned_date': returned_date,
            'penalty': penalty or 'None'
        })
    
    # Pagination for rental items
    rental_page = request.GET.get('rental_page', 1)
    rental_paginator = Paginator(rental_items_list, 10)
    try:
        rental_items = rental_paginator.page(rental_page)
    except PageNotAnInteger:
        rental_items = rental_paginator.page(1)
    except EmptyPage:
        rental_items = rental_paginator.page(rental_paginator.num_pages)
    
    # Flags for export readiness (20+ items)
    orders_ready = completed_orders_count_range >= 20
    sales_ready = sales_range.count() >= 20
    
    context = {
        'date_from': date_from_display,
        'date_to': date_to_display,
        'total_orders': total_orders,
        'completed_orders': completed_orders,
        'pending_orders': pending_orders,
        'total_revenue': total_revenue,
        'completed_orders_count_range': completed_orders_count_range,
        'orders': orders,
        'sales_list': sales_list,
        'staff_reports': staff_reports,
        'staff_chart_data': json.dumps(staff_chart_data),
        'materials_logs': materials_logs,
        'rental_items': rental_items,
        'orders_ready': orders_ready,
        'sales_ready': sales_ready,
    }
    
    return render(request, 'business/reports.html', context)


@login_required
@csrf_exempt  # Allow AJAX callers even if CSRF token/cookie is missing
def send_sms(request):
    """API endpoint to send SMS"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            phone_number = data.get('phone_number') or data.get('phone')
            message = data.get('message')
            order_id = data.get('order_id')
            
            if not phone_number or not message:
                return JsonResponse({
                    'success': False,
                    'error': 'Phone number and message are required'
                }, status=400)
            
            if not order_id:
                return JsonResponse({
                    'success': False,
                    'error': 'Order ID is required to log the SMS'
                }, status=400)
            
            try:
                order = Order.objects.get(id=order_id)
            except Order.DoesNotExist:
                return JsonResponse({
                    'success': False,
                    'error': 'Order not found'
                }, status=404)
            
            normalized_phone = normalize_phone_number(phone_number)
            if not normalized_phone.startswith('+'):
                return JsonResponse({
                    'success': False,
                    'error': 'Phone number must include the country code (e.g. +639XXXXXXXXX)'
                }, status=400)
            
            sms = SMSNotification.objects.create(
                order=order,
                phone_number=normalized_phone,
                message=message,
                status='pending'
            )
            
            try:
                
                # twilio_message = client.messages.create(from_=settings.TWILIO_PHONE_NUMBER, to=normalized_phone, body=message)
                # curl -X POST -u <username>:<password> \
                # -H "Content-Type: application/json" \
                # -d '{ "textMessage": { "text": "Hello, world!"}, "phoneNumbers": ["+79990001234", "+79995556677"] }' \
                # http://<device_local_ip>:8080/message

                url = "http://10.27.63.146:8080/message"

                payload = {
                    "textMessage": {
                        "text": message
                    },
                    "phoneNumbers": [
                        normalized_phone,
                    ]
                }

                response = requests.post(
                    url,
                    json=payload,
                    auth=HTTPBasicAuth("sms", "Q4L35hBv"),
                    headers={"Content-Type": "application/json"}
                )

                print("Status code:", response.status_code)
                print("Response:", response.text)

                sms.status = 'sent'
                sms.sent_at = timezone.now()
                sms.save(update_fields=['status', 'sent_at'])
                
                return JsonResponse({
                    'success': True,
                    'message': 'SMS sent successfully',
                    'sms_id': sms.id,
                    # 'twilio_sid': twilio_message.sid
                })
            except Exception as exc:
                sms.status = 'failed'
                sms.save(update_fields=['status'])
                raise exc
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=500)
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})


@login_required
def api_rental_availability(request):
    """API endpoint to check rental availability"""
    if request.method == 'POST':
        try:
            import json
            data = json.loads(request.body)
            product_ids = data.get('product_ids', [])
            
            available_products = []
            for product_id in product_ids:
                try:
                    product = Product.objects.get(
                        id=product_id,
                        product_type='rental',
                        is_active=True,
                        is_archived=False
                    )
                    if product.rental_status == 'available':
                        available_products.append(product_id)
                except Product.DoesNotExist:
                    pass
            
            return JsonResponse({
                'success': True,
                'available_products': available_products
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            })
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})


@login_required
def api_return_rental(request):
    """API endpoint to return rental items"""
    if request.method == 'POST':
        try:
            import json
            data = json.loads(request.body)
            order_id = data.get('order_id')
            
            if not order_id:
                return JsonResponse({
                    'success': False,
                    'error': 'Order ID is required'
                })
            
            try:
                order = Order.objects.get(id=order_id)
            except Order.DoesNotExist:
                return JsonResponse({
                    'success': False,
                    'error': 'Order not found'
                })
            
            # Return all rental products in the order
            returned_count = 0
            for product in order.items.filter(product__product_type='rental').values_list('product', flat=True):
                try:
                    product_obj = Product.objects.get(id=product)
                    if product_obj.rental_status == 'rented':
                        product_obj.rental_status = 'available'
                        product_obj.current_rental_order = None
                        product_obj.save()
                        returned_count += 1
                except Product.DoesNotExist:
                    pass
            
            return JsonResponse({
                'success': True,
                'message': f'{returned_count} items returned',
                'returned_count': returned_count
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            })
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})


@login_required
def api_rental_items(request):
    """API endpoint to get rental items"""
    if request.method == 'GET':
        try:
            rental_products = Product.objects.filter(
                product_type='rental',
                is_active=True,
                is_archived=False
            )
            
            items = []
            for product in rental_products:
                items.append({
                    'id': product.id,
                    'name': product.name,
                    'rental_status': product.rental_status,
                    'price': float(product.price) if product.price else 0.0,
                    'image_url': product.image.url if product.image else None
                })
            
            return JsonResponse({
                'success': True,
                'items': items,
                'count': len(items)
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            })
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})


@login_required
def api_rental_status_update(request):
    """API endpoint to update rental status"""
    if request.method == 'POST':
        try:
            import json
            data = json.loads(request.body)
            product_id = data.get('product_id')
            status = data.get('status')
            
            if not product_id or not status:
                return JsonResponse({
                    'success': False,
                    'error': 'Product ID and status are required'
                })
            
            product = get_object_or_404(Product, id=product_id)
            product.rental_status = status
            product.save()
            
            return JsonResponse({
                'success': True,
                'message': 'Rental status updated'
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            })
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})


@login_required
def archive(request):
    """Archive page to view archived items"""
    # Get filter type and page numbers from query parameters
    filter_type = request.GET.get('type', 'all')
    items_per_page = 10  # Number of items per page
    
    # Get archived products
    archived_products_qs = Product.objects.filter(is_archived=True).select_related('category', 'material_type').order_by('-updated_at')
    
    # Get archived/cancelled orders
    cancelled_orders_qs = Order.objects.filter(
        Q(is_archived=True) | Q(status='cancelled')
    ).select_related('customer').order_by('-updated_at')
    
    # Get archived customers
    archived_customers_qs = Customer.objects.none()  # Empty queryset for now since Customer doesn't have is_archived
    
    # Apply filters
    if filter_type == 'products':
        archived_products_qs = archived_products_qs
        cancelled_orders_qs = Order.objects.none()
        archived_customers_qs = Customer.objects.none()
    elif filter_type == 'orders':
        archived_products_qs = Product.objects.none()
        cancelled_orders_qs = cancelled_orders_qs
        archived_customers_qs = Customer.objects.none()
    elif filter_type == 'customers':
        archived_products_qs = Product.objects.none()
        cancelled_orders_qs = Order.objects.none()
        # archived_customers_qs = archived_customers_qs (keep as is)
    # else 'all' - show all
    
    # Paginate products
    products_page = request.GET.get('products_page', 1)
    products_paginator = Paginator(archived_products_qs, items_per_page)
    try:
        archived_products = products_paginator.page(products_page)
    except PageNotAnInteger:
        archived_products = products_paginator.page(1)
    except EmptyPage:
        if products_paginator.num_pages > 0:
            archived_products = products_paginator.page(products_paginator.num_pages)
        else:
            archived_products = products_paginator.page(1)
    
    # Paginate orders
    orders_page = request.GET.get('orders_page', 1)
    orders_paginator = Paginator(cancelled_orders_qs, items_per_page)
    try:
        cancelled_orders = orders_paginator.page(orders_page)
    except PageNotAnInteger:
        cancelled_orders = orders_paginator.page(1)
    except EmptyPage:
        if orders_paginator.num_pages > 0:
            cancelled_orders = orders_paginator.page(orders_paginator.num_pages)
        else:
            cancelled_orders = orders_paginator.page(1)
    
    # Paginate customers
    customers_page = request.GET.get('customers_page', 1)
    customers_paginator = Paginator(archived_customers_qs, items_per_page)
    try:
        archived_customers = customers_paginator.page(customers_page)
    except PageNotAnInteger:
        archived_customers = customers_paginator.page(1)
    except EmptyPage:
        if customers_paginator.num_pages > 0:
            archived_customers = customers_paginator.page(customers_paginator.num_pages)
        else:
            archived_customers = customers_paginator.page(1)
    
    # Calculate counts (use original querysets, not paginated)
    total_products_count = Product.objects.filter(is_archived=True).count()
    total_orders_count = Order.objects.filter(Q(is_archived=True) | Q(status='cancelled')).count()
    total_customers_count = 0  # Customer model doesn't have is_archived field
    total_count = total_products_count + total_orders_count + total_customers_count
    
    context = {
        'archived_products': archived_products,
        'cancelled_orders': cancelled_orders,
        'archived_customers': archived_customers,
        'products_paginator': products_paginator,
        'orders_paginator': orders_paginator,
        'customers_paginator': customers_paginator,
        'total_count': total_count,
        'products_count': total_products_count,
        'orders_count': total_orders_count,
        'customers_count': total_customers_count,
        'current_filter': filter_type,
    }
    
    return render(request, 'business/archive.html', context)


@login_required
def restore_item(request, item_type, item_id):
    """Restore an archived item"""
    if request.method == 'POST':
        try:
            if item_type == 'product':
                item = get_object_or_404(Product, id=item_id)
                item.is_archived = False
                item.is_active = True
                item.save()
            elif item_type == 'order':
                item = get_object_or_404(Order, id=item_id)
                item.is_archived = False
                item.save()
            elif item_type == 'customer':
                # Customer model doesn't have is_archived field, so restoration isn't applicable
                messages.warning(request, 'Customer restoration is not available. Customer model does not support archiving.')
                return redirect('archive')
            
            messages.success(request, f'{item_type.title()} restored successfully')
        except Exception as e:
            messages.error(request, f'Error restoring {item_type}: {str(e)}')
    
    return redirect('archive')


@login_required
def delete_permanent(request, item_type, item_id):
    """Permanently delete an item"""
    if request.method == 'POST':
        try:
            if item_type == 'product':
                item = get_object_or_404(Product, id=item_id)
                item.delete()
            elif item_type == 'order':
                item = get_object_or_404(Order, id=item_id)
                item.delete()
            elif item_type == 'customer':
                item = get_object_or_404(Customer, id=item_id)
                item.delete()
            
            messages.success(request, f'{item_type.title()} deleted permanently')
        except Exception as e:
            messages.error(request, f'Error deleting {item_type}: {str(e)}')
    
    return redirect('archive')


@login_required
def bulk_delete_permanent(request):
    """Permanently delete multiple items"""
    if request.method == 'POST':
        try:
            import json
            data = json.loads(request.body)
            item_type = data.get('item_type')
            item_ids = data.get('item_ids', [])
            
            if not item_type or not item_ids:
                return JsonResponse({
                    'success': False,
                    'error': 'Item type and IDs are required'
                })
            
            deleted_count = 0
            for item_id in item_ids:
                try:
                    if item_type == 'product':
                        item = Product.objects.get(id=item_id)
                        item.delete()
                        deleted_count += 1
                    elif item_type == 'order':
                        item = Order.objects.get(id=item_id)
                        item.delete()
                        deleted_count += 1
                    elif item_type == 'customer':
                        item = Customer.objects.get(id=item_id)
                        item.delete()
                        deleted_count += 1
                except (Product.DoesNotExist, Order.DoesNotExist, Customer.DoesNotExist):
                    pass
            
            return JsonResponse({
                'success': True,
                'message': f'{deleted_count} items deleted permanently',
                'deleted_count': deleted_count
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            })
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})


# Additional missing API functions - stubs for now
@login_required
def api_check_inventory_availability(request):
    """API endpoint to check inventory availability"""
    return JsonResponse({'success': False, 'error': 'Not implemented'})


@login_required
def api_update_rental_status(request):
    """API endpoint to update rental status"""
    return JsonResponse({'success': False, 'error': 'Not implemented'})


@login_required
def api_fix_stuck_rental_products(request):
    """API endpoint to fix stuck rental products"""
    return JsonResponse({'success': False, 'error': 'Not implemented'})


@login_required
def api_get_order_customer_data(request, order_id):
    """API endpoint to get order customer data"""
    return JsonResponse({'success': False, 'error': 'Not implemented'})


@login_required
def api_return_details(request, order_id):
    """API endpoint to get return details"""
    return JsonResponse({'success': False, 'error': 'Not implemented'})


@login_required
def api_generate_qr_code(request):
    """API endpoint to generate QR code"""
    return JsonResponse({'success': False, 'error': 'Not implemented'})


@login_required
def api_generate_qr_code_for_order(request):
    """API endpoint to generate QR code for order"""
    if request.method == 'POST':
        try:
            import json
            data = json.loads(request.body)
            order_id = data.get('order_id')
            order_identifier = data.get('order_identifier')
            
            if not order_id:
                return JsonResponse({
                    'success': False,
                    'error': 'Order ID is required'
                })
            
            # Get the order
            try:
                order = Order.objects.get(id=order_id)
            except Order.DoesNotExist:
                return JsonResponse({
                    'success': False,
                    'error': 'Order not found'
                })
            
            # Generate QR code using the existing function
            generate_qr_code(order)
            
            # Get the QR code object
            qr_code_obj = QRCode.objects.filter(order=order).first()
            
            if qr_code_obj and qr_code_obj.qr_code_image:
                return JsonResponse({
                    'success': True,
                    'qr_code': qr_code_obj.qr_code_image.url,
                    'message': 'QR code generated successfully'
                })
            else:
                return JsonResponse({
                    'success': False,
                    'error': 'Failed to generate QR code'
                })
                
        except Exception as e:
            import traceback
            traceback.print_exc()
            return JsonResponse({
                'success': False,
                'error': str(e)
            })
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})


@login_required
def api_fix_order_identifiers(request):
    """API endpoint to fix order identifiers"""
    return JsonResponse({'success': False, 'error': 'Not implemented'})


@login_required
def api_fix_static_orders(request):
    """API endpoint to fix static orders"""
    from .static_data_manager import get_static_orders, remove_static_data
    
    if request.method == 'POST':
        try:
            dry_run = request.POST.get('dry_run', 'true').lower() == 'true'
            counts = remove_static_data(dry_run=dry_run)
            
            return safe_json_response({
                'success': True,
                'dry_run': dry_run,
                'counts': counts,
                'message': f'Found {counts["products"]} static products, {counts["orders"]} static orders' + 
                          ('. Deleted.' if not dry_run else '. Run with dry_run=false to delete.')
            })
        except Exception as e:
            return safe_json_response({
                'success': False,
                'error': str(e)
            })
    
    return safe_json_response({'success': False, 'error': 'Invalid request method'})


@login_required
def api_cleanup_all_static_orders(request):
    """API endpoint to cleanup all static orders and products"""
    from .static_data_manager import (get_static_orders, get_static_products,
                                      remove_static_data)
    
    if request.method == 'POST':
        try:
            # First show what will be deleted (dry run)
            dry_run_counts = remove_static_data(dry_run=True)
            
            # Actually delete
            actual_counts = remove_static_data(dry_run=False)
            
            return safe_json_response({
                'success': True,
                'deleted': actual_counts,
                'message': f'Successfully removed {actual_counts["products"]} static products, {actual_counts["orders"]} static orders, {actual_counts["order_items"]} order items, and {actual_counts["sales"]} sales records.'
            })
        except Exception as e:
            return safe_json_response({
                'success': False,
                'error': str(e)
            })
    
    return safe_json_response({'success': False, 'error': 'Invalid request method'})


@login_required
def api_cleanup_duplicate_customize_products(request):
    """API endpoint to cleanup duplicate customize products"""
    from business.customize_product_manager import (
        find_duplicate_customize_products, remove_duplicate_customize_products)
    
    if request.method == 'POST':
        try:
            # Check if it's a dry run request
            dry_run = request.POST.get('dry_run', 'false').lower() == 'true'
            keep_newest = request.POST.get('keep_newest', 'false').lower() == 'true'
            
            if dry_run:
                # Just find and report duplicates
                duplicates = find_duplicate_customize_products()
                duplicate_list = []
                for key, duplicate_info in duplicates.items():
                    duplicate_list.append({
                        'method': duplicate_info['method'],
                        'identifier': duplicate_info['identifier'],
                        'product_count': len(duplicate_info['products']),
                        'products': [
                            {
                                'id': p.id,
                                'name': p.name,
                                'created_at': p.created_at.isoformat() if p.created_at else None,
                                'image_url': p.image.url if p.image else None
                            }
                            for p in duplicate_info['products']
                        ]
                    })
                
                return safe_json_response({
                    'success': True,
                    'dry_run': True,
                    'duplicate_groups': len(duplicates),
                    'duplicates': duplicate_list,
                    'message': f'Found {len(duplicates)} duplicate groups'
                })
            else:
                # Actually remove duplicates
                keep_oldest = not keep_newest
                result = remove_duplicate_customize_products(dry_run=False, keep_oldest=keep_oldest)
                
                return safe_json_response({
                    'success': True,
                    'dry_run': False,
                    'result': result,
                    'message': f'Archived {result["products_to_delete"]} duplicate products, kept {result["products_to_keep"]} unique products'
                })
        except Exception as e:
            import traceback
            traceback.print_exc()
            return safe_json_response({
                'success': False,
                'error': str(e)
            })
    
    return safe_json_response({'success': False, 'error': 'Invalid request method'})


@login_required
def api_ensure_only_real_orders(request):
    """API endpoint to ensure only real orders"""
    return JsonResponse({'success': False, 'error': 'Not implemented'})


@login_required
def api_sync_frontend_backend_orders(request):
    """API endpoint to sync frontend backend orders"""
    return JsonResponse({'success': False, 'error': 'Not implemented'})


@login_required
def api_sync_rental_orders(request):
    """API endpoint to sync rental orders"""
    return JsonResponse({'success': False, 'error': 'Not implemented'})


@login_required
def api_sync_repair_orders(request):
    """API endpoint to sync repair orders"""
    return JsonResponse({'success': False, 'error': 'Not implemented'})


@login_required
def api_sync_custom_orders(request):
    """API endpoint to sync custom orders"""
    return JsonResponse({'success': False, 'error': 'Not implemented'})


@login_required
def api_backfill_repair_order_categories(request):
    """API endpoint to backfill repair order categories"""
    return JsonResponse({'success': False, 'error': 'Not implemented'})


@login_required
def admin_settings(request):
    """Admin settings page"""
    return render(request, 'business/admin_settings.html', {})


@login_required
def help_support(request):
    """Help support page"""
    return render(request, 'business/help_support.html', {})


@login_required
def display_accessibility(request):
    """Display accessibility page"""
    return render(request, 'business/display_accessibility.html', {})


@login_required
def feedback(request):
    """Feedback page"""
    return render(request, 'business/feedback.html', {})


@login_required
def uniform_measurement_form(request):
    """Uniform measurement form page"""
    return render(request, 'business/uniform_measurement_form.html', {})


@login_required
def navigation_health_check(request):
    """API endpoint for navigation health check"""
    return JsonResponse({'success': True, 'status': 'ok'})


@login_required
def quick_nav_check(request):
    """API endpoint for quick navigation check"""
    return JsonResponse({'success': True, 'status': 'ok'})


@login_required
def api_autosave_sync(request):
    """
    API endpoint to sync auto-saved data from frontend to backend.
    Ensures all changes are persisted.
    """
    if request.method == 'POST':
        try:
            import json

            from .persistence_manager import ChangeTracker
            
            # Handle empty request body
            if not request.body:
                return JsonResponse({'success': True, 'synced': 0, 'message': 'No data to sync'})
            
            try:
                data = json.loads(request.body)
            except json.JSONDecodeError:
                return JsonResponse({'success': True, 'synced': 0, 'message': 'Invalid or empty JSON'})
            
            sync_data = data.get('data', {})
            timestamp = data.get('timestamp')
            
            # Process each saved item
            synced_count = 0
            logger = logging.getLogger(__name__)
            for key, value in sync_data.items():
                try:
                    # Only track sync for non-modal forms (modal forms are temporary)
                    # Skip autosave forms from modals to reduce log noise
                    if 'addStaffForm' in key or 'editStaffForm' in key:
                        # Skip logging for modal forms - they're temporary
                        synced_count += 1
                        continue
                    
                    # Log the sync for tracking (only for non-modal forms)
                    ChangeTracker.track_change(
                        model_name='AutoSave',
                        object_id=key,
                        action='sync',
                        changes={'data': value, 'timestamp': timestamp}
                    )
                    synced_count += 1
                except Exception as e:
                    logger.warning(f"Failed to sync item {key}: {str(e)}")
            
            return JsonResponse({
                'success': True,
                'message': f'Synced {synced_count} items',
                'synced_count': synced_count,
                'timestamp': timezone.now().isoformat()
            })
            
        except Exception as e:
            import traceback
            traceback.print_exc()
            return JsonResponse({
                'success': False,
                'error': str(e)
            })
    
    return JsonResponse({'success': False, 'error': 'Invalid request method'})
    #     try:
    #                 synced_count += 1
    #             except Exception as e:
    #                 import logging
    #                 logging.getLogger(__name__).warning(f"Failed to sync item {key}: {str(e)}")
            
    #         return JsonResponse({
    #             'success': True,
    #             'message': f'Synced {synced_count} items',
    #             'synced_count': synced_count,
    #             'timestamp': timezone.now().isoformat()
    #         })
            
    #     except Exception as e:
    #         import traceback
    #         traceback.print_exc()
    #         return JsonResponse({
    #             'success': False,
    #             'error': str(e)
    #         })
    
    # return JsonResponse({'success': False, 'error': 'Invalid request method'})
    #                 )
    #                 synced_count += 1
    #             except Exception as e:
    #                 import logging
    #                 logging.getLogger(__name__).warning(f"Failed to sync item {key}: {str(e)}")
            
    #         return JsonResponse({
    #             'success': True,
    #             'message': f'Synced {synced_count} items',
    #             'synced_count': synced_count,
    #             'timestamp': timezone.now().isoformat()
    #         })
            
    #     except Exception as e:
    #         import traceback
    #         traceback.print_exc()
    #         return JsonResponse({
    #             'success': False,
    #             'error': str(e)
    #         })
    # return JsonResponse({'success': False, 'error': 'Invalid request method'})


# ============================================
# STAFF MANAGEMENT VIEWS
# ============================================

@login_required
@ensure_csrf_cookie
def staff_management(request):
    """Staff management page with performance tracking and efficiency ratings"""
    from datetime import timedelta
    from decimal import Decimal

    from django.db.models import Avg, Count, F, Q, Sum

    # Get all staff members (users with is_staff=True) with profiles
    staff_members = User.objects.filter(is_staff=True, is_active=True).select_related('staff_profile').order_by('first_name', 'username')
    
    # Calculate statistics for each staff member
    staff_stats = []
    now = timezone.now()
    
    for staff in staff_members:
        # Get orders assigned to this staff
        assigned_orders = Order.objects.filter(assigned_staff=staff)
        
        # Get active assigned orders (not completed, not cancelled)
        active_orders = assigned_orders.exclude(status__in=['completed', 'cancelled'])
        
        # Check if staff is busy (has active assigned orders)
        is_busy = active_orders.exists()
        busy_status = 'available'
        busy_status_class = 'success'  # Green for available
        time_since_assignment = None
        active_order_count = 0
        
        if is_busy:
            # Get the most recent active order
            most_recent_order = active_orders.order_by('-staff_assigned_at').first()
            if most_recent_order and most_recent_order.staff_assigned_at:
                time_diff = now - most_recent_order.staff_assigned_at
                time_since_assignment_minutes = int(time_diff.total_seconds() / 60)
                time_since_assignment = time_since_assignment_minutes
                active_order_count = active_orders.count()
                
                # Determine status based on time
                if time_since_assignment_minutes < 20:
                    busy_status = 'busy'
                    busy_status_class = 'danger'  # Red for busy < 20 min
                else:
                    busy_status = 'overdue'
                    busy_status_class = 'warning'  # Orange for busy > 20 min
        
        # Total orders
        total_orders = assigned_orders.count()
        
        # Completed orders
        completed_orders = assigned_orders.filter(status='completed').count()
        
        # Calculate salary: 40% of total amount from completed repair and customize orders
        completed_repair_customize_orders = assigned_orders.filter(
            status='completed',
            order_type__in=['repair', 'customize']
        )
        total_revenue = completed_repair_customize_orders.aggregate(
            total=Sum('total_amount')
        )['total'] or Decimal('0')
        staff_salary = total_revenue * Decimal('0.4')  # 40% to staff
        owner_share = total_revenue * Decimal('0.6')  # 60% to owner
        
        # Pending orders
        pending_orders = assigned_orders.filter(status__in=['pending', 'in_progress', 'repair_done', 'ready_to_pick_up']).count()
        
        # Calculate completion rate
        completion_rate = (completed_orders / total_orders * 100) if total_orders > 0 else 0
        
        # Calculate average completion time (for completed orders)
        avg_completion_time = None
        completed_order_list = assigned_orders.filter(status='completed', created_at__isnull=False, updated_at__isnull=False)
        
        if completed_order_list.exists():
            total_time = timedelta()
            count = 0
            for order in completed_order_list:
                if order.created_at and order.updated_at:
                    time_diff = order.updated_at - order.created_at
                    total_time += time_diff
                    count += 1
            
            if count > 0:
                avg_seconds = total_time.total_seconds() / count
                avg_hours = avg_seconds / 3600
                avg_completion_time = round(avg_hours, 1)
        
        # Recent activity (orders in last 7 days)
        recent_orders = assigned_orders.filter(created_at__gte=now - timedelta(days=7)).count()
        
        # Calculate efficiency score (0-100)
        # Based on: completion rate (40%), speed (30%), recent activity (30%)
        efficiency_score = 0
        
        # Completion rate component (40 points max)
        efficiency_score += (completion_rate * 0.4)
        
        # Speed component (30 points max) - faster is better
        # Assuming ideal completion time is 24 hours (1 day)
        if avg_completion_time:
            if avg_completion_time <= 24:
                speed_score = 30
            elif avg_completion_time <= 48:
                speed_score = 25
            elif avg_completion_time <= 72:
                speed_score = 20
            else:
                speed_score = 15
            efficiency_score += speed_score
        
        # Recent activity component (30 points max)
        if recent_orders >= 5:
            activity_score = 30
        elif recent_orders >= 3:
            activity_score = 25
        elif recent_orders >= 1:
            activity_score = 20
        else:
            activity_score = 10
        efficiency_score += activity_score
        
        # Determine rating level
        if efficiency_score >= 80:
            rating = 'Excellent'
            rating_class = 'success'
        elif efficiency_score >= 60:
            rating = 'Good'
            rating_class = 'info'
        elif efficiency_score >= 40:
            rating = 'Average'
            rating_class = 'warning'
        else:
            rating = 'Needs Improvement'
            rating_class = 'danger'
        
        staff_stats.append({
            'staff': staff,
            'total_orders': total_orders,
            'completed_orders': completed_orders,
            'completed_repair_customize_count': completed_repair_customize_orders.count(),
            'pending_orders': pending_orders,
            'completion_rate': round(completion_rate, 1),
            'avg_completion_time': avg_completion_time,
            'recent_orders': recent_orders,
            'efficiency_score': round(efficiency_score, 1),
            'rating': rating,
            'rating_class': rating_class,
            'is_busy': is_busy,
            'busy_status': busy_status,
            'busy_status_class': busy_status_class,
            'time_since_assignment': time_since_assignment,
            'active_order_count': active_order_count,
            'active_orders': list(active_orders.values('id', 'order_identifier', 'staff_assigned_at', 'order_type', 'status')[:5]) if is_busy else [],
            'total_revenue': float(total_revenue),
            'staff_salary': float(staff_salary),
            'owner_share': float(owner_share),
            'completed_orders_details': list(completed_repair_customize_orders.values('id', 'order_identifier', 'total_amount', 'order_type', 'created_at', 'updated_at')[:50]),
        })
    
    # Sort by efficiency score (highest first)
    staff_stats.sort(key=lambda x: x['efficiency_score'], reverse=True)
    
    # Overall statistics
    total_staff = staff_members.count()
    total_assigned_orders = Order.objects.filter(assigned_staff__isnull=False).count()
    total_completed_by_staff = Order.objects.filter(assigned_staff__isnull=False, status='completed').count()
    
    context = {
        'staff_stats': staff_stats,
        'total_staff': total_staff,
        'total_assigned_orders': total_assigned_orders,
        'total_completed_by_staff': total_completed_by_staff,
    }
    
    return render(request, 'business/staff_management.html', context)


@login_required
def staff_report_pdf(request):
    """Generate comprehensive PDF report of all staff with detailed information"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="staff_detailed_report.pdf"'
    
    doc = SimpleDocTemplate(response, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    story = []
    
    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#1e3a8a'),
        spaceAfter=30,
        alignment=TA_CENTER
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor('#2c3e50'),
        spaceAfter=12,
        spaceBefore=12
    )
    
    # Title
    story.append(Paragraph("TopStyle Business - Staff Detailed Report", title_style))
    story.append(Paragraph(f"Generated on: {timezone.now().strftime('%B %d, %Y at %I:%M %p')}", styles['Normal']))
    story.append(Spacer(1, 20))
    
    # Get all staff members
    staff_members = User.objects.filter(is_staff=True, is_active=True).select_related('staff_profile').order_by('first_name', 'username')
    
    for staff in staff_members:
        # Staff Information Section
        story.append(Paragraph(f"<b>Staff: {staff.get_full_name() or staff.username}</b>", heading_style))
        
        # Staff Details Table
        staff_data = [
            ['Username:', staff.username],
            ['Email:', staff.email or 'N/A'],
            ['Phone:', staff.staff_profile.phone if hasattr(staff, 'staff_profile') and staff.staff_profile.phone else 'N/A'],
            ['Status:', 'Active' if staff.is_active else 'Inactive'],
        ]
        staff_table = Table(staff_data, colWidths=[2*inch, 4*inch])
        staff_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f8f9fa')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ]))
        story.append(staff_table)
        story.append(Spacer(1, 15))
        
        # Get completed orders
        completed_orders = Order.objects.filter(
            assigned_staff=staff,
            status='completed',
            order_type__in=['repair', 'customize']
        ).select_related('customer').order_by('-created_at')
        
        # Calculate totals
        from django.db.models import Sum
        from decimal import Decimal
        total_revenue = completed_orders.aggregate(total=Sum('total_amount'))['total'] or Decimal('0')
        staff_salary = total_revenue * Decimal('0.4')
        owner_share = total_revenue * Decimal('0.6')
        
        # Get withdrawals
        from .models import StaffWithdrawal
        withdrawals = StaffWithdrawal.objects.filter(staff=staff).order_by('-created_at')
        total_withdrawn = withdrawals.aggregate(total=Sum('withdrawal_amount'))['total'] or Decimal('0')
        available_salary = staff_salary - total_withdrawn
        
        # Summary Table
        summary_data = [
            ['Total Completed Orders:', str(completed_orders.count())],
            ['Total Revenue:', f'₱{float(total_revenue):,.2f}'],
            ['Staff Share (40%):', f'₱{float(staff_salary):,.2f}'],
            ['Owner Share (60%):', f'₱{float(owner_share):,.2f}'],
            ['Total Withdrawn:', f'₱{float(total_withdrawn):,.2f}'],
            ['Available Salary:', f'₱{float(available_salary):,.2f}'],
        ]
        summary_table = Table(summary_data, colWidths=[2.5*inch, 3.5*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#e9ecef')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ]))
        story.append(Paragraph("<b>Summary</b>", styles['Heading3']))
        story.append(summary_table)
        story.append(Spacer(1, 15))
        
        # Completed Orders Table
        if completed_orders.exists():
            story.append(Paragraph("<b>Completed Orders</b>", styles['Heading3']))
            orders_data = [['Order ID', 'Customer', 'Order Type', 'Amount', 'Created', 'Completed']]
            
            for order in completed_orders:
                orders_data.append([
                    order.order_identifier,
                    order.customer.name if order.customer else 'N/A',
                    order.get_order_type_display(),
                    f'₱{float(order.total_amount):,.2f}',
                    order.created_at.strftime('%Y-%m-%d %H:%M') if order.created_at else 'N/A',
                    order.updated_at.strftime('%Y-%m-%d %H:%M') if order.updated_at else 'N/A',
                ])
            
            orders_table = Table(orders_data, colWidths=[1.2*inch, 1.5*inch, 0.8*inch, 1*inch, 1.2*inch, 1.2*inch])
            orders_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c3e50')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
            ]))
            story.append(orders_table)
            story.append(Spacer(1, 15))
        
        # Withdrawals Table
        if withdrawals.exists():
            story.append(Paragraph("<b>Salary Withdrawals</b>", styles['Heading3']))
            withdrawals_data = [['Date & Time', 'Amount', 'Total Revenue', 'Owner Share', 'Orders Count', 'Notes']]
            
            for withdrawal in withdrawals:
                withdrawals_data.append([
                    withdrawal.created_at.strftime('%Y-%m-%d %H:%M') if withdrawal.created_at else 'N/A',
                    f'₱{float(withdrawal.withdrawal_amount):,.2f}',
                    f'₱{float(withdrawal.total_revenue):,.2f}',
                    f'₱{float(withdrawal.owner_share):,.2f}',
                    str(withdrawal.completed_orders_count),
                    withdrawal.notes or 'N/A',
                ])
            
            withdrawals_table = Table(withdrawals_data, colWidths=[1.2*inch, 1*inch, 1*inch, 1*inch, 0.8*inch, 1.5*inch])
            withdrawals_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c3e50')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
            ]))
            story.append(withdrawals_table)
            story.append(Spacer(1, 15))
        else:
            story.append(Paragraph("<b>Salary Withdrawals:</b> No withdrawals recorded", styles['Normal']))
            story.append(Spacer(1, 10))
        
        # Non-withdrawn Salary
        if available_salary > 0:
            story.append(Paragraph(f"<b>Available (Non-Withdrawn) Salary:</b> ₱{float(available_salary):,.2f}", styles['Normal']))
            story.append(Spacer(1, 10))
        
        story.append(PageBreak())
    
    doc.build(story)
    return response


@login_required
def staff_report_excel(request):
    """Generate comprehensive Excel report of all staff with detailed information"""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="staff_detailed_report.xlsx"'
    
    wb = Workbook()
    
    # Styles
    header_fill = PatternFill(start_color='2c3e50', end_color='2c3e50', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    title_fill = PatternFill(start_color='1e3a8a', end_color='1e3a8a', fill_type='solid')
    title_font = Font(bold=True, color='FFFFFF', size=14)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    
    # Get all staff members
    staff_members = User.objects.filter(is_staff=True, is_active=True).select_related('staff_profile').order_by('first_name', 'username')
    
    # Create summary sheet
    summary_ws = wb.active
    summary_ws.title = "Summary"
    
    # Summary header
    summary_ws['A1'] = 'TopStyle Business - Staff Detailed Report'
    summary_ws['A1'].font = title_font
    summary_ws['A1'].fill = title_fill
    summary_ws['A1'].alignment = center_align
    summary_ws.merge_cells('A1:H1')
    summary_ws['A2'] = f'Generated on: {timezone.now().strftime("%B %d, %Y at %I:%M %p")}'
    summary_ws.merge_cells('A2:H2')
    summary_ws['A2'].alignment = center_align
    
    # Summary headers
    summary_headers = ['Staff Name', 'Username', 'Email', 'Total Orders', 'Total Revenue', 'Staff Salary', 'Withdrawn', 'Available']
    for col, header in enumerate(summary_headers, 1):
        cell = summary_ws.cell(row=4, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
    
    # Summary data
    from django.db.models import Sum
    from decimal import Decimal
    from .models import StaffWithdrawal
    
    row = 5
    for staff in staff_members:
        completed_orders = Order.objects.filter(
            assigned_staff=staff,
            status='completed',
            order_type__in=['repair', 'customize']
        )
        total_revenue = completed_orders.aggregate(total=Sum('total_amount'))['total'] or Decimal('0')
        staff_salary = total_revenue * Decimal('0.4')
        total_withdrawn = StaffWithdrawal.objects.filter(staff=staff).aggregate(
            total=Sum('withdrawal_amount')
        )['total'] or Decimal('0')
        available_salary = staff_salary - total_withdrawn
        
        summary_ws.cell(row=row, column=1, value=staff.get_full_name() or staff.username).border = border
        summary_ws.cell(row=row, column=2, value=staff.username).border = border
        summary_ws.cell(row=row, column=3, value=staff.email or 'N/A').border = border
        summary_ws.cell(row=row, column=4, value=completed_orders.count()).border = border
        summary_ws.cell(row=row, column=5, value=float(total_revenue)).border = border
        summary_ws.cell(row=row, column=6, value=float(staff_salary)).border = border
        summary_ws.cell(row=row, column=7, value=float(total_withdrawn)).border = border
        summary_ws.cell(row=row, column=8, value=float(available_salary)).border = border
        row += 1
    
    # Adjust column widths
    for col in range(1, 9):
        summary_ws.column_dimensions[get_column_letter(col)].width = 18
    
    # Create detailed sheet for each staff
    for staff in staff_members:
        ws = wb.create_sheet(title=staff.username[:31])  # Excel sheet name limit
        
        # Title
        ws['A1'] = f'Staff: {staff.get_full_name() or staff.username}'
        ws['A1'].font = title_font
        ws['A1'].fill = title_fill
        ws.merge_cells('A1:F1')
        ws['A1'].alignment = center_align
        
        ws['A2'] = f'Generated on: {timezone.now().strftime("%B %d, %Y at %I:%M %p")}'
        ws.merge_cells('A2:F2')
        ws['A2'].alignment = center_align
        
        # Staff Information
        row = 4
        ws.cell(row=row, column=1, value='Username:').font = Font(bold=True)
        ws.cell(row=row, column=2, value=staff.username)
        row += 1
        ws.cell(row=row, column=1, value='Email:').font = Font(bold=True)
        ws.cell(row=row, column=2, value=staff.email or 'N/A')
        row += 1
        ws.cell(row=row, column=1, value='Phone:').font = Font(bold=True)
        ws.cell(row=row, column=2, value=staff.staff_profile.phone if hasattr(staff, 'staff_profile') and staff.staff_profile.phone else 'N/A')
        row += 1
        ws.cell(row=row, column=1, value='Status:').font = Font(bold=True)
        ws.cell(row=row, column=2, value='Active' if staff.is_active else 'Inactive')
        row += 2
        
        # Get data
        completed_orders = Order.objects.filter(
            assigned_staff=staff,
            status='completed',
            order_type__in=['repair', 'customize']
        ).select_related('customer').order_by('-created_at')
        
        total_revenue = completed_orders.aggregate(total=Sum('total_amount'))['total'] or Decimal('0')
        staff_salary = total_revenue * Decimal('0.4')
        owner_share = total_revenue * Decimal('0.6')
        
        withdrawals = StaffWithdrawal.objects.filter(staff=staff).order_by('-created_at')
        total_withdrawn = withdrawals.aggregate(total=Sum('withdrawal_amount'))['total'] or Decimal('0')
        available_salary = staff_salary - total_withdrawn
        
        # Summary Section
        row += 1
        ws.cell(row=row, column=1, value='SUMMARY').font = Font(bold=True, size=12)
        row += 1
        summary_info = [
            ['Total Completed Orders:', str(completed_orders.count())],
            ['Total Revenue:', f'₱{float(total_revenue):,.2f}'],
            ['Staff Share (40%):', f'₱{float(staff_salary):,.2f}'],
            ['Owner Share (60%):', f'₱{float(owner_share):,.2f}'],
            ['Total Withdrawn:', f'₱{float(total_withdrawn):,.2f}'],
            ['Available Salary:', f'₱{float(available_salary):,.2f}'],
        ]
        for info_row in summary_info:
            ws.cell(row=row, column=1, value=info_row[0]).font = Font(bold=True)
            ws.cell(row=row, column=2, value=info_row[1])
            row += 1
        
        row += 2
        
        # Completed Orders
        if completed_orders.exists():
            ws.cell(row=row, column=1, value='COMPLETED ORDERS').font = Font(bold=True, size=12)
            row += 1
            
            order_headers = ['Order ID', 'Customer', 'Order Type', 'Amount', 'Created Date', 'Created Time', 'Completed Date', 'Completed Time']
            for col, header in enumerate(order_headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = center_align
            row += 1
            
            for order in completed_orders:
                ws.cell(row=row, column=1, value=order.order_identifier).border = border
                ws.cell(row=row, column=2, value=order.customer.name if order.customer else 'N/A').border = border
                ws.cell(row=row, column=3, value=order.get_order_type_display()).border = border
                ws.cell(row=row, column=4, value=float(order.total_amount)).border = border
                if order.created_at:
                    ws.cell(row=row, column=5, value=order.created_at.strftime('%Y-%m-%d')).border = border
                    ws.cell(row=row, column=6, value=order.created_at.strftime('%H:%M:%S')).border = border
                else:
                    ws.cell(row=row, column=5, value='N/A').border = border
                    ws.cell(row=row, column=6, value='N/A').border = border
                if order.updated_at:
                    ws.cell(row=row, column=7, value=order.updated_at.strftime('%Y-%m-%d')).border = border
                    ws.cell(row=row, column=8, value=order.updated_at.strftime('%H:%M:%S')).border = border
                else:
                    ws.cell(row=row, column=7, value='N/A').border = border
                    ws.cell(row=row, column=8, value='N/A').border = border
                row += 1
            
            row += 1
        
        # Withdrawals
        if withdrawals.exists():
            ws.cell(row=row, column=1, value='SALARY WITHDRAWALS').font = Font(bold=True, size=12)
            row += 1
            
            withdrawal_headers = ['Date', 'Time', 'Amount', 'Total Revenue', 'Owner Share', 'Orders Count', 'Notes']
            for col, header in enumerate(withdrawal_headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = center_align
            row += 1
            
            for withdrawal in withdrawals:
                if withdrawal.created_at:
                    ws.cell(row=row, column=1, value=withdrawal.created_at.strftime('%Y-%m-%d')).border = border
                    ws.cell(row=row, column=2, value=withdrawal.created_at.strftime('%H:%M:%S')).border = border
                else:
                    ws.cell(row=row, column=1, value='N/A').border = border
                    ws.cell(row=row, column=2, value='N/A').border = border
                ws.cell(row=row, column=3, value=float(withdrawal.withdrawal_amount)).border = border
                ws.cell(row=row, column=4, value=float(withdrawal.total_revenue)).border = border
                ws.cell(row=row, column=5, value=float(withdrawal.owner_share)).border = border
                ws.cell(row=row, column=6, value=withdrawal.completed_orders_count).border = border
                ws.cell(row=row, column=7, value=withdrawal.notes or 'N/A').border = border
                row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 15
    
    wb.save(response)
    return response


@login_required
@require_http_methods(["GET"])
def staff_salary(request, staff_id):
    """Return salary breakdown for a staff member as JSON.

    Calculates completed repair/customize total revenue, staff share (40%), and owner share (60%).
    """
    try:
        staff = User.objects.get(id=staff_id, is_staff=True)
        from decimal import Decimal

        from django.db.models import Sum

        completed_qs = Order.objects.filter(
            assigned_staff=staff,
            status='completed',
            order_type__in=['repair', 'customize']
        )
        total_revenue = completed_qs.aggregate(total=Sum('total_amount'))['total'] or Decimal('0')
        
        # Calculate total withdrawn amount
        from .models import StaffWithdrawal
        total_withdrawn = StaffWithdrawal.objects.filter(staff=staff).aggregate(
            total=Sum('withdrawal_amount')
        )['total'] or Decimal('0')
        
        # New bucket logic: only count revenue after the last withdrawal
        last_withdrawal = StaffWithdrawal.objects.filter(staff=staff).order_by('-created_at').first()
        if last_withdrawal:
            completed_qs = completed_qs.filter(created_at__gt=last_withdrawal.created_at)
            total_revenue = completed_qs.aggregate(total=Sum('total_amount'))['total'] or Decimal('0')
        
        # Calculate current available amounts (bucket-based)
        owner_share = total_revenue * Decimal('0.6')
        staff_salary = total_revenue * Decimal('0.4')
        available_revenue = total_revenue  # full bucket; reset after withdrawal
        completed_count = completed_qs.count()

        return JsonResponse({
            'success': True,
            'salary_data': {
                'staff_id': staff_id,
                'staff_name': staff.get_full_name() or staff.username,
                'completed_orders': completed_count,
                'total_revenue': float(total_revenue),
                'total_withdrawn': float(total_withdrawn),
                'available_revenue': float(available_revenue),
                'staff_salary': float(staff_salary),
                'owner_share': float(owner_share),
            }
        })
    except User.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Staff member not found'}, status=404)
    except Exception as e:
        logger = logging.getLogger(__name__)
        logger.exception(f"Error computing salary for staff {staff_id}: {e}")
        return JsonResponse({'success': False, 'error': 'Failed to calculate salary data'}, status=500)


@login_required
@csrf_protect
@require_http_methods(["POST"])
def withdraw_staff_revenue(request, staff_id):
    """Process staff revenue withdrawal and create withdrawal history record"""
    try:
        staff = User.objects.get(id=staff_id, is_staff=True)
        from decimal import Decimal
        from django.db.models import Sum
        from django.db import transaction
        from .models import StaffWithdrawal, ActivityLog

        # Get latest withdrawal to implement "bucket reset" behavior
        last_withdrawal = StaffWithdrawal.objects.filter(staff=staff).order_by('-created_at').first()
        
        # Get current salary data for orders completed AFTER last withdrawal
        completed_qs = Order.objects.filter(
            assigned_staff=staff,
            status='completed',
            order_type__in=['repair', 'customize']
        )
        if last_withdrawal:
            completed_qs = completed_qs.filter(created_at__gt=last_withdrawal.created_at)

        total_revenue = completed_qs.aggregate(total=Sum('total_amount'))['total'] or Decimal('0')
        
        # Bucket reset: after a withdrawal, totals start from zero and accumulate again
        owner_share = total_revenue * Decimal('0.6')  # Owner share on post-withdrawal revenue
        staff_salary = total_revenue * Decimal('0.4')  # Staff salary on post-withdrawal revenue
        completed_count = completed_qs.count()
        
        # Validate withdrawal amount (force full bucket withdrawal)
        withdrawal_amount = staff_salary
        
        if withdrawal_amount <= 0:
            return JsonResponse({
                'success': False,
                'error': 'Withdrawal amount must be greater than zero'
            }, status=400)
        
        if withdrawal_amount > staff_salary:
            return JsonResponse({
                'success': False,
                'error': f'Withdrawal amount (₱{withdrawal_amount}) exceeds available staff revenue (₱{staff_salary})'
            }, status=400)
        
        # Create withdrawal record
        with transaction.atomic():
            withdrawal = StaffWithdrawal.objects.create(
                staff=staff,
                withdrawal_amount=withdrawal_amount,
                total_revenue=total_revenue,
                owner_share=owner_share,
                completed_orders_count=completed_count,
                withdrawn_by=request.user,
                notes=request.POST.get('notes', '')
            )
            
            # Log activity
            ActivityLog.objects.create(
                activity_type='other',
                description=f'Staff revenue withdrawal: {staff.get_full_name() or staff.username} - ₱{withdrawal_amount} (Completed Orders: {completed_count}, Total Revenue: ₱{total_revenue})',
                user=request.user
            )
        
        return JsonResponse({
            'success': True,
            'message': f'Successfully withdrew ₱{withdrawal_amount} for {staff.get_full_name() or staff.username}',
            'withdrawal': {
                'id': withdrawal.id,
                'amount': float(withdrawal_amount),
                'completed_orders': completed_count,
                'total_revenue': float(total_revenue),
                'owner_share': float(owner_share),
                'date': withdrawal.created_at.isoformat()
            }
        })
        
    except User.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Staff member not found'}, status=404)
    except Exception as e:
        logger = logging.getLogger(__name__)
        logger.exception(f"Error processing withdrawal for staff {staff_id}: {e}")
        return JsonResponse({'success': False, 'error': f'Failed to process withdrawal: {str(e)}'}, status=500)


@login_required
def staff_withdrawal_history(request, staff_id):
    """Get withdrawal history for a staff member"""
    try:
        staff = User.objects.get(id=staff_id, is_staff=True)
        from .models import StaffWithdrawal
        
        withdrawals = StaffWithdrawal.objects.filter(staff=staff).order_by('-created_at')
        
        withdrawal_list = []
        for w in withdrawals:
            withdrawal_list.append({
                'id': w.id,
                'amount': float(w.withdrawal_amount),
                'total_revenue': float(w.total_revenue),
                'owner_share': float(w.owner_share),
                'completed_orders': w.completed_orders_count,
                'date': w.created_at.strftime('%Y-%m-%d'),
                'time': w.created_at.strftime('%H:%M:%S'),
                'datetime': w.created_at.isoformat(),
                'withdrawn_by': w.withdrawn_by.get_full_name() or w.withdrawn_by.username if w.withdrawn_by else 'System',
                'notes': w.notes
            })
        
        return JsonResponse({
            'success': True,
            'withdrawals': withdrawal_list,
            'total_withdrawn': float(sum(w.withdrawal_amount for w in withdrawals))
        })
        
    except User.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Staff member not found'}, status=404)
    except Exception as e:
        logger = logging.getLogger(__name__)
        logger.exception(f"Error fetching withdrawal history for staff {staff_id}: {e}")
        return JsonResponse({'success': False, 'error': 'Failed to fetch withdrawal history'}, status=500)


@login_required
@csrf_protect
@require_http_methods(["POST"])
def add_staff(request):
    """
    Add a new staff member to the system.
    
    This function creates a new user account with staff privileges and
    associates it with a StaffProfile containing additional information.
    
    Required fields:
    - username: Unique username for login
    - password: Password for the account (optional - will be auto-generated if not provided)
    - first_name: Staff member's first name
    - last_name: Staff member's last name
    - email: Email address (must be unique)
    - phone: Phone number
    
    Optional fields:
    - profile_image: Profile picture file
    
    Returns JSON response with success status and message.
    """
    import logging
    logger = logging.getLogger(__name__)
    
    # Only allow POST requests
    if request.method != 'POST':
        return JsonResponse({
            'success': False,
            'error': 'Only POST method is allowed'
        }, status=405)
    
    # Log the request for debugging
    logger.info(f"Add staff request received from user: {request.user.username if request.user.is_authenticated else 'Anonymous'}")
    logger.info(f"Request method: {request.method}")
    logger.info(f"POST data keys: {list(request.POST.keys())}")
    logger.info(f"FILES data keys: {list(request.FILES.keys())}")
    
    try:
        from django.db import transaction

        from .models import ActivityLog, StaffProfile

        # Get form data
        username = request.POST.get('username', '').strip()
        email = request.POST.get('email', '').strip()
        first_name = request.POST.get('first_name', '').strip()
        last_name = request.POST.get('last_name', '').strip()
        phone = request.POST.get('phone', '').strip()
        password = request.POST.get('password', '').strip()
        profile_image = request.FILES.get('profile_image')
        
        # Validate required fields with detailed error messages
        validation_errors = []
        
        if not first_name:
            validation_errors.append('First name is required')
        
        if not last_name:
            validation_errors.append('Last name is required')
        
        if not email:
            validation_errors.append('Email address is required')
        elif '@' not in email:
            validation_errors.append('Please enter a valid email address')
        
        if not phone:
            validation_errors.append('Phone number is required')
        elif len(phone) < 10:
            validation_errors.append('Phone number must be at least 10 digits')
        
        if not username:
            validation_errors.append('Username is required')
        elif len(username) < 3:
            validation_errors.append('Username must be at least 3 characters')
        elif len(username) > 150:
            validation_errors.append('Username must be 150 characters or less')
        
        # Normalize username: replace spaces with underscores and remove invalid characters
        # Django User model doesn't allow spaces, so we'll convert them
        original_username = username
        username = username.replace(' ', '_')  # Replace spaces with underscores
        username = ''.join(c for c in username if c.isalnum() or c in ['_', '-'])  # Remove invalid chars
        
        # Check if normalized username is still valid
        if len(username) < 3:
            validation_errors.append('Username must contain at least 3 valid characters (spaces will be converted to underscores)')
        
        # Generate password if not provided
        if not password:
            # Generate a secure random password
            import secrets
            import string
            alphabet = string.ascii_letters + string.digits + "!@#$%^&*"
            password = ''.join(secrets.choice(alphabet) for i in range(12))
            logger.info(f"Generated random password for new staff: {username}")
        elif len(password) < 8:
            validation_errors.append('Password must be at least 8 characters')
        
        # Helper function to return error response
        def error_response(error_msg, status_code=400):
            is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest' or \
                      'application/json' in request.headers.get('Accept', '')
            if is_ajax:
                return JsonResponse({'success': False, 'error': error_msg}, status=status_code)
            else:
                messages.error(request, error_msg)
                return redirect('staff_management')
        
        # Return validation errors if any
        if validation_errors:
            return error_response('Validation failed: ' + '; '.join(validation_errors))
        
        # Check for duplicate username
        if User.objects.filter(username=username).exists():
            return error_response(f'Username "{username}" already exists. Please choose a different username.')
        
        # Check for duplicate email
        if User.objects.filter(email=email).exists():
            return error_response(f'Email "{email}" is already registered. Please use a different email address.')
        
        # Use transaction to ensure data consistency
        with transaction.atomic():
            # Create new staff user
            try:
                user = User.objects.create_user(
                    username=username,
                    email=email,
                    password=password,
                    first_name=first_name,
                    last_name=last_name,
                    is_staff=True,
                    is_active=True
                )
                logger.info(f"Created staff user: {username} (ID: {user.id})")
            except Exception as e:
                logger.error(f"Failed to create user: {str(e)}")
                return error_response(f'Failed to create user account: {str(e)}', 500)
            
            # Create staff profile with phone and image if provided
            try:
                profile = StaffProfile.objects.create(
                    user=user,
                    phone=phone
                )
                
                # Handle profile image upload if provided
                if profile_image:
                    # Validate image file
                    if profile_image.size > 5 * 1024 * 1024:  # 5MB limit
                        return error_response('Profile image is too large. Maximum size is 5MB.')
                    
                    # Validate image format
                    allowed_formats = ['image/jpeg', 'image/jpg', 'image/png', 'image/gif']
                    if profile_image.content_type not in allowed_formats:
                        return error_response('Invalid image format. Please upload a JPEG, PNG, or GIF image.')
                    
                    profile.profile_image = profile_image
                    profile.save()
                    logger.info(f"Uploaded profile image for staff: {username}")
                
                logger.info(f"Created staff profile for: {username}")
            except Exception as e:
                logger.error(f"Failed to create staff profile: {str(e)}")
                # Rollback user creation if profile creation fails
                user.delete()
                return error_response(f'Failed to create staff profile: {str(e)}', 500)
            
            # Log activity
            try:
                ActivityLog.objects.create(
                    activity_type='other',
                    description=f'New staff member added: {user.get_full_name() or user.username} (Email: {email}, Phone: {phone})',
                    user=request.user
                )
            except Exception as e:
                # Don't fail if activity log fails, just log it
                logger.warning(f"Failed to log activity for new staff: {str(e)}")
            
            # Send welcome email with login credentials (optional - don't fail if email fails)
            try:
                from django.conf import settings
                staff_name = user.get_full_name() or user.username
                subject = f'Welcome to TopStyle Business - Your Staff Account'
                message = f"""Hello {staff_name},

Your staff account has been created successfully!

Login Credentials:
- Username: {username}
- Password: {password}

Please log in at: {request.build_absolute_uri('/login/')}

For security reasons, we recommend changing your password after your first login.

If you have any questions, please contact the administrator.

Best regards,
TopStyle Business Team"""
                
                from_email = settings.DEFAULT_FROM_EMAIL or settings.EMAIL_HOST_USER
                send_mail(
                    subject,
                    message,
                    from_email,
                    [email],
                    fail_silently=True  # Don't fail the whole operation if email fails
                )
                logger.info(f"Welcome email sent to new staff: {email}")
            except Exception as e:
                # Don't fail if email sending fails, just log it
                logger.warning(f"Failed to send welcome email to new staff: {str(e)}")
        
        # Success response
        staff_name = user.get_full_name() or user.username
        
        # Check if this is an AJAX request
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest' or \
                  'application/json' in request.headers.get('Accept', '')
        
        if is_ajax:
            return JsonResponse({
                'success': True,
                'message': f'Staff member "{staff_name}" has been added successfully! Login credentials have been sent to {email}.',
                'staff_id': user.id,
                'staff_name': staff_name,
                'username': username,
                'email': email
            }, status=201)
        else:
            # Regular form submission - redirect back to staff page with success message
            messages.success(request, f'Staff member "{staff_name}" has been added successfully!')
            return redirect('staff_management')
        
    except Exception as e:
        import traceback
        error_trace = traceback.format_exc()
        logger.error(f"Error adding staff: {str(e)}\n{error_trace}")
        
        # Check if this is an AJAX request
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest' or \
                  'application/json' in request.headers.get('Accept', '')
        
        if is_ajax:
            return JsonResponse({
                'success': False,
                'error': f'An unexpected error occurred: {str(e)}'
            }, status=500)
        else:
            # Regular form submission - redirect back with error message
            messages.error(request, f'Error adding staff: {str(e)}')
            return redirect('staff_management')


@login_required
@csrf_protect
@require_http_methods(["POST"])
def edit_staff(request, staff_id):
    """Edit staff member details"""
    try:
        from .models import StaffProfile
        
        staff = User.objects.get(id=staff_id, is_staff=True)
        profile_image = request.FILES.get('profile_image')
        
        email = request.POST.get('email')
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        is_active = request.POST.get('is_active') == 'true'
        
        # Update fields
        if email:
            # Check if email is already used by another user
            if User.objects.filter(email=email).exclude(id=staff_id).exists():
                return JsonResponse({'success': False, 'error': 'Email already exists'})
            staff.email = email
        
        if first_name:
            staff.first_name = first_name
        if last_name:
            staff.last_name = last_name
        
        staff.is_active = is_active
        staff.save()
        
        # Update or create staff profile with image if provided
        profile, created = StaffProfile.objects.get_or_create(user=staff)
        if profile_image:
            profile.profile_image = profile_image
            profile.save()
        
        # Log activity
        ActivityLog.objects.create(
            activity_type='other',
            description=f'Staff member updated: {staff.get_full_name() or staff.username}'
        )
        
        return JsonResponse({
            'success': True,
            'message': f'Staff member {staff.get_full_name() or staff.username} updated successfully'
        })
        
    except User.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Staff member not found'})
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error editing staff: {e}")
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
@csrf_protect
@require_http_methods(["POST"])
def delete_staff(request, staff_id):
    """Delete a staff member from the system"""
    try:
        from django.db import transaction, IntegrityError

        from .models import ActivityLog, StaffProfile, StaffWithdrawal
        
        staff = User.objects.get(id=staff_id, is_staff=True)
        staff_name = staff.get_full_name() or staff.username
        
        # Check if staff has active assigned orders
        active_orders = Order.objects.filter(
            assigned_staff=staff
        ).exclude(status__in=['completed', 'cancelled'])
        
        if active_orders.exists():
            return JsonResponse({
                'success': False,
                'error': f'Cannot delete {staff_name}. They have {active_orders.count()} active assigned order(s). Please reassign or complete these orders first.'
            }, status=400)
        
        # Use transaction to ensure data consistency
        with transaction.atomic():
            from django.db import connection
            # Temporarily disable constraint checks while we clean up references
            with connection.constraint_checks_disabled():
                # Log activity before deletion
                ActivityLog.objects.create(
                    activity_type='other',
                    description=f'Staff member deleted: {staff_name}'
                )
                
                # Delete staff profile if it exists
                StaffProfile.objects.filter(user=staff).delete()
                
                # Remove staff assignment from all orders (set to None)
                Order.objects.filter(assigned_staff=staff).update(assigned_staff=None)

                # Clean up withdrawals referencing this staff
                StaffWithdrawal.objects.filter(staff=staff).delete()
                StaffWithdrawal.objects.filter(withdrawn_by=staff).update(withdrawn_by=None)

                # Clear any activity logs pointing to this staff to avoid FK issues
                ActivityLog.objects.filter(user=staff).update(user=None)

                # Clean up any other models with FK to the user
                from django.apps import apps
                from django.db.models import ForeignKey, OneToOneField
                from django.contrib.auth import get_user_model
                
                user_model = get_user_model()
                for model in apps.get_models():
                    if model is user_model:
                        continue
                    
                    fk_fields = [
                        field for field in model._meta.fields
                        if isinstance(field, ForeignKey)
                        and field.remote_field
                        and field.remote_field.model == user_model
                    ]
                    
                    if not fk_fields:
                        continue
                    
                    manager = model._default_manager
                    for field in fk_fields:
                        related_qs = manager.filter(**{field.name: staff})
                        if not related_qs.exists():
                            continue
                        
                        if field.null and not isinstance(field, OneToOneField):
                            related_qs.update(**{field.name: None})
                        else:
                            related_qs.delete()
                
                # Delete the user account
                staff.delete()
            
            # Validate constraints after cleanup
            connection.check_constraints()
        
        return JsonResponse({
            'success': True,
            'message': f'Staff member {staff_name} has been deleted successfully. All assigned orders have been unassigned.'
        })
        
    except User.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Staff member not found'
        }, status=404)
    except IntegrityError as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Integrity error deleting staff {staff_id}: {e}", exc_info=True)
        return JsonResponse({
            'success': False,
            'error': 'Cannot delete staff due to related records. Please clear related withdrawals/orders/logs and try again.'
        }, status=500)
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error deleting staff: {e}", exc_info=True)
        return JsonResponse({
            'success': False,
            'error': f'An error occurred while deleting staff: {str(e)}'
        }, status=500)


@login_required
def staff_detail(request, staff_id):
    """View detailed performance of a specific staff member"""
    try:
        staff = User.objects.get(id=staff_id, is_staff=True)
        
        # Get all orders assigned to this staff
        assigned_orders = Order.objects.filter(assigned_staff=staff).order_by('-created_at')
        
        # Get order activities related to this staff
        activities = ActivityLog.objects.filter(
            Q(activity_type='staff_assigned') | Q(order__assigned_staff=staff)
        ).order_by('-created_at')[:20]
        
        context = {
            'staff': staff,
            'assigned_orders': assigned_orders,
            'activities': activities,
        }
        
        return render(request, 'business/staff_detail.html', context)
        
    except User.DoesNotExist:
        messages.error(request, 'Staff member not found')
        return redirect('staff_management')
